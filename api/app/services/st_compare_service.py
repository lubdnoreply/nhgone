"""ST Files verification: our stored numbers vs each property's own
"<Name>-ST" Google Sheet, which is the ground truth for what actually gets
filed.

This module is the comparison and its rendering only. The recipient, send
time, subject and surrounding email design live in Admin > Email Template >
System Email > Test ST File, like every other scheduled mail in this app -
see email_service.ST_COMPARE_TEMPLATE_KEY. What this module hands the
template is the two finished tables (render_tokens), so an edit to the
template's wording can never silently produce a check with no numbers in it.

Two things about the source sheets drive the whole design:

1. Each workbook holds exactly ONE pasted MEWS export, for whatever day was
   last put into it - it is not a running log. So the date is READ FROM THE
   SHEETS rather than assumed, and a day they no longer hold simply cannot be
   compared.
2. The Master tab mixes TWO different MEWS exports. Spaces/Occupied/House
   uses/Out of order/Availability/Customers/Arrivals/Departures come from the
   Availability report (`Parameters` tab); Complimentary is computed off the
   `Reservation` tab, a separate Reservation report. Only the first group is
   compared cell-for-cell here; Complimentary is compared as a total.
"""
import asyncio
import io
import json
import logging
import re
from collections import OrderedDict
from datetime import datetime, timedelta, timezone

import httpx
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Same IDs as the links the front-office team maintains.
SHEETS = OrderedDict([
    ("Lub d Bangkok Chinatown",       ("Chinatown", "1npf-d74wYYwsQk9LrNUyYJxnCHH0dsLVuxELJo_CqUM")),
    ("Lub d Bangkok Siam",            ("Siam",      "1JFEQcs1lz62KSIYPuzJhY8o_nP_4PEGdZxmeKrqELVc")),
    ("Lub d Koh Samui Chaweng Beach", ("Samui",     "1DwDsPAajjFH5Fbe_-6Jcl-3ds0q-DCkyGVZr2sRpSw4")),
    ("Lub d Koh Tao Tanote Bay",      ("Koh Tao",   "1-48rztEk1J0TmOiFgPfTE9I_UB0Ba9-RvxOKj8aJyG0")),
    ("Lub d Philippines Makati",      ("Makati",    "1ykTLblIv9dIXEDzdivzPr-z6IWDr_asdeCTn0evut3k")),
    ("Lub d Phuket Patong",           ("Patong",    "1X8mh5Hlvcl6hZhO4Z-eyqdZeGQSFwbYjwHTjzMxcEEs")),
    ("Lub d Siem Reap",               ("Siem Reap", "17siV7sMIT5GsW9x8LdsWBVjh1DcSUTchX-WwiS09FpY")),
    ("Marasca Samui",                 ("Marasca",   "1r2i50lPT8VOFKsjqnSgphawAQvgb3S8yWafyGpVIOCs")),
])

# (Master row label, our st_files_sync key). Siam and Siem Reap label the last
# row "Complimentary Room" where the others say "Complimentary".
METRICS = [
    ("Spaces",        "spaces"),
    ("Occupied",      "occupied"),
    ("House uses",    "house_use"),
    ("Out of order",  "out_of_order"),
    ("Availability",  "availability"),
    ("Customers",     "customers"),
    ("Arrivals",      "arrivals"),
    ("Departures",    "departures"),
    ("Complimentary", "complimentary"),
]

# Manila is UTC+8; every other property is UTC+7.
TZ_OFFSET = {"Lub d Philippines Makati": 8}


def _parse_master(content: bytes) -> dict:
    """Master tab as {label: value}, plus the report date and the timestamp
    MEWS generated the underlying Availability export."""
    wb = load_workbook(io.BytesIO(content), data_only=True)

    master = {}
    ws = wb["Master"]
    for row in range(1, ws.max_row + 1):
        label = str(ws.cell(row, 1).value or "").strip()
        if label:
            master[label] = ws.cell(row, 2).value
    if "Complimentary" not in master and "Complimentary Room" in master:
        master["Complimentary"] = master["Complimentary Room"]

    # Prefer the Availability parameters tab - several sheets also carry a
    # "Parameters Reservation"/"Parameters-Reservation" tab for the OTHER
    # export, whose Created time would be the wrong one to report.
    created = None
    params = next((wb[n] for n in ("Parameters", "Parameters-Availability") if n in wb.sheetnames), None)
    if params is None:
        params = next((wb[n] for n in wb.sheetnames
                       if n.startswith("Parameters") and "eservation" not in n), None)
    if params is not None:
        for row in range(1, 21):
            if str(params.cell(row, 1).value or "").strip() == "Created":
                created = params.cell(row, 2).value

    # The Arrivals/Departures tabs hold the SAME export broken down per space
    # category - the detail the Master tab's single total hides. Reading them
    # is what turns a morning's "Chinatown -2" into "CTS -1, TNK -1", which is
    # the difference between a number to investigate from scratch and a
    # difference somebody can act on. Layout is Service | Space category |
    # <date> | Total, with a trailing "Total" row that is skipped.
    per_category = {}
    for tab in ("Arrivals", "Departures"):
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        counts = {}
        for row in range(2, ws.max_row + 1):
            service = ws.cell(row, 1).value
            category = ws.cell(row, 2).value
            value = ws.cell(row, 3).value
            if not category or str(service).strip().lower() == "total":
                continue
            try:
                counts[str(category).strip()] = int(value or 0)
            except (TypeError, ValueError):
                continue
        per_category[tab.lower()] = counts

    report_date = master.get("Date")
    if isinstance(report_date, datetime):
        report_date = report_date.strftime("%Y-%m-%d")
    return {"master": master, "date": report_date, "created": created,
            "per_category": per_category}


async def _fetch_sheets() -> dict:
    """All 8 Master tabs, fetched concurrently over the workbooks' public
    export URL. A sheet that fails comes back as None rather than failing the
    whole run - one unreachable workbook should still leave seven comparable."""
    async def one(prop, sheet_id):
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        try:
            async with httpx.AsyncClient(follow_redirects=True, timeout=90) as client:
                r = await client.get(url)
                r.raise_for_status()
            return prop, _parse_master(r.content), None
        except Exception as e:
            logger.warning(f"ST compare: could not read {prop}'s sheet: {e}")
            return prop, None, str(e)

    results = await asyncio.gather(*(one(p, sid) for p, (_, sid) in SHEETS.items()))
    return {p: {"data": d, "error": e} for p, d, e in results}


def _as_int(v):
    """A blank Complimentary cell (Siam, Siem Reap) means zero, not missing."""
    if v is None or v == "":
        return 0
    return int(v) if isinstance(v, (int, float)) else v


def _local(ts: str, prop: str) -> str:
    if not ts:
        return ""
    ts = re.sub(r"\.(\d+)", lambda m: "." + m.group(1)[:6].ljust(6, "0"), ts)
    off = TZ_OFFSET.get(prop, 7)
    return datetime.fromisoformat(ts).astimezone(timezone(timedelta(hours=off))).strftime("%d %b %H:%M")


def _our_categories(property_name: str, date: str) -> dict:
    """Our stored Arrivals/Departures for one day, per space category, as
    {"arrivals": {cat: units}, "departures": {...}}.

    Reads the stored report blob directly: get_st_files_list returns the ten
    totals and nothing underneath them, and it is the breakdown underneath
    that says WHICH category a total is short on. Counts `spaces`, not rows,
    for the same reason the report's own arrivals_count does - one dorm
    reservation is several beds.
    """
    from app.services.encryption import encryption_service
    from app.services.sync_service import sync_service

    out = {"arrivals": {}, "departures": {}}
    try:
        res = sync_service.supabase.table("st_files_sync").select("data").eq(
            "property", property_name).eq("report_date", date).limit(1).execute()
        if not res.data:
            return out
        report = json.loads(encryption_service.decrypt(res.data[0]["data"]["blob"]))
    except Exception as e:
        logger.warning(f"ST compare: could not read {property_name}'s stored report: {e}")
        return out

    for key in ("arrivals", "departures"):
        counts = {}
        for row in report.get(key) or []:
            cat = (row.get("category") or "?").strip()
            counts[cat] = counts.get(cat, 0) + (row.get("spaces") or 0)
        out[key] = counts
    return out


def _category_note(ours: dict, sheet: dict) -> str:
    """"CTS -1, TNK -1" - the categories a total differs on, ours vs sheet."""
    bits = []
    for cat in sorted(set(ours) | set(sheet)):
        o, s = ours.get(cat, 0), sheet.get(cat, 0)
        if o != s:
            bits.append(f"{cat} {o - s:+d}")
    return ", ".join(bits)


async def build_comparison(want_date: str = None) -> dict:
    """The whole check, as data. `status` is one of:
      ok             - comparable, see columns/grid
      no_sheet_date  - the sheets disagree with each other about the date
      not_held       - a date was asked for that the sheets no longer hold
    """
    from app.services.sync_service import sync_service

    sheets = await _fetch_sheets()
    dates = {p: v["data"]["date"] for p, v in sheets.items() if v["data"]}
    distinct = sorted({d for d in dates.values() if d})

    if want_date and want_date not in distinct:
        # Hard stop rather than a warning: these workbooks hold one pasted
        # export each, so a day they don't hold cannot be reconstructed - and
        # a table comparing our 23rd against the sheets' 25th reads exactly as
        # authoritative as a real one while being nonsense.
        return {"status": "not_held", "want": want_date, "distinct": distinct, "dates": dates}
    if not want_date and len(distinct) != 1:
        return {"status": "no_sheet_date", "distinct": distinct, "dates": dates}

    date = want_date or distinct[0]

    ours = {}
    ours_by_category = {}
    for prop in SHEETS:
        try:
            rows = await sync_service.get_st_files_list(prop)
            ours[prop] = next((r for r in rows if r.get("date") == date), None)
        except Exception as e:
            logger.warning(f"ST compare: could not read st_files_sync for {prop}: {e}")
            ours[prop] = None
        # Our own Arrivals/Departures per space category, off the same stored
        # blob, so a total that disagrees can name the categories responsible.
        # Read separately from get_st_files_list because that one returns the
        # ten totals only.
        ours_by_category[prop] = _our_categories(prop, date)

    columns, grid, detail = [], {}, {}
    for label, key in METRICS:
        ok, notes = 0, []
        for prop, (short, _) in SHEETS.items():
            sh, ou = sheets[prop]["data"], ours.get(prop)
            if not sh or ou is None:
                notes.append(f"{short} ไม่มีข้อมูล")
                grid.setdefault(prop, {})[label] = (None, None)
                continue
            sv, ov = _as_int(sh["master"].get(label)), ou.get(key)
            grid.setdefault(prop, {})[label] = (ov, sv)
            if sv == ov:
                ok += 1
            else:
                note = f"{short} {ov - sv:+d}"
                # Arrivals is the one metric with a KNOWN, named source of
                # drift: the day-use night-tail rule (sync_service's
                # _ST_DAY_USE_NIGHT_END_HOUR) holds back same-day-checkout
                # stays that started before its cutoff, and MEWS's own
                # classification of those isn't purely hour-based - see that
                # rule's own comment for the contradictory evidence. Surfacing
                # the count here doesn't claim it explains the whole gap
                # (the exclusion count and the sheet gap can differ, e.g. a
                # separate bug on top), just saves whoever reads this mail
                # from re-deriving "is this the day-use thing again?" by hand
                # every single day.
                if key == "arrivals":
                    excluded = ou.get("day_use_arrivals_excluded", 0)
                    if excluded:
                        note += f" ({excluded} day-use excluded as night-tail)"
                # Which space categories the total is short (or long) on. The
                # sheet publishes Arrivals and Departures per category on their
                # own tabs; every other metric has only the one total, so this
                # can only be said for those two.
                if key in ("arrivals", "departures"):
                    cat_note = _category_note(
                        (ours_by_category.get(prop) or {}).get(key, {}),
                        (sh.get("per_category") or {}).get(key, {}))
                    if cat_note:
                        note += f" [{cat_note}]"
                notes.append(note)
                detail.setdefault(prop, []).append((label, ov, sv))
        columns.append({"label": label, "matched": ok, "total": len(SHEETS), "notes": notes})

    stamps = sorted(s for s in (_local((v or {}).get("synced_at", ""), p) for p, v in ours.items()) if s)
    total = len(METRICS) * len(SHEETS)
    mismatched = sum(len(v) for v in detail.values())
    return {
        "status": "ok",
        "date": date,
        "columns": columns,
        "grid": grid,
        "detail": detail,
        "total_cells": total,
        "matched_cells": total - mismatched,
        "window": (stamps[0], stamps[-1]) if stamps else None,
        "sheet_errors": {p: v["error"] for p, v in sheets.items() if v["error"]},
    }


def _title(result: dict) -> str:
    day = datetime.strptime(result["date"], "%Y-%m-%d")
    # "%-d" (unpadded day) is a glibc extension and raises ValueError on
    # Windows - see the same note in rr4_compare_service._title.
    return f"ST Files {day.day} {day.strftime('%b %Y')}"


def render_text(result: dict) -> str:
    """Plain-text form - the CLI output, and the email's text/plain part."""
    if result["status"] == "not_held":
        return (f"⛔ ชีตไม่ได้เก็บข้อมูลวันที่ {result['want']} — ตอนนี้ถือวันที่ "
                f"{', '.join(result['distinct'])}\n"
                "   ชีตแต่ละอันเก็บได้ครั้งละ 1 วันเท่านั้น (ทับของเดิมทุกครั้งที่วางข้อมูลใหม่)\n"
                "   จึงย้อนไปเทียบวันที่ผ่านมาไม่ได้")
    if result["status"] == "no_sheet_date":
        lines = ["⚠️  ชีตแต่ละอันถือคนละวัน - ยังเทียบไม่ได้:"]
        for p, d in result["dates"].items():
            lines.append(f"     {SHEETS[p][0]:<12} {d}")
        return "\n".join(lines)

    out = ["=" * 78, f"สรุปตามคอลัมน์ — {_title(result)}", "=" * 78]
    day = datetime.strptime(result["date"], "%Y-%m-%d")
    out.append(f"ชีตทั้ง {len(SHEETS)} ถือข้อมูลวันที่ {day.day} {day.strftime('%b %Y')} ตรงกัน")
    if result["window"]:
        out.append(f"เทียบ snapshot ของเรา (จับเวลา {result['window'][0]} – {result['window'][1]}) กับชีต")
    out += ["", f"{'คอลัมน์':<16}{'ตรง':<8}หมายเหตุ", "-" * 78]
    for c in result["columns"]:
        note = "✅" if c["matched"] == c["total"] else ", ".join(c["notes"])
        score = f"{c['matched']}/{c['total']}"
        out.append(f"{c['label']:<16}{score:<8}{note}")
    out.append("-" * 78)
    out.append(f"ตรงกัน {result['matched_cells']}/{result['total_cells']} ช่อง"
               + ("  — ทุกช่องตรงหมด ✅" if result["matched_cells"] == result["total_cells"] else ""))
    if result["detail"]:
        out.append("")
        out.append("รายละเอียดที่ต่าง (เรา / ชีต):")
        for prop, items in result["detail"].items():
            out.append(f"  {SHEETS[prop][0]:<12} " + ",  ".join(f"{l} {o}/{s}" for l, o, s in items))
    for prop, err in (result.get("sheet_errors") or {}).items():
        out.append(f"  !! {SHEETS[prop][0]}: อ่านชีตไม่ได้ - {err}")
    return "\n".join(out)


_TD = "padding:6px 10px;border:1px solid #e2e8f0;font-size:13px;"
_TH = "padding:6px 10px;border:1px solid #e2e8f0;font-size:11px;font-weight:700;background:#f8fafc;text-align:left;"


def render_summary_table(result: dict) -> str:
    """Per-column "ตรง X/8" summary - the <<SummaryTable>> token, and the
    table this whole mail was originally asked for."""
    if result["status"] != "ok":
        return f'<pre style="font-family:ui-monospace,monospace;font-size:13px">{render_text(result)}</pre>'

    h = [f'<table style="border-collapse:collapse;width:100%"><tr>'
         f'<th style="{_TH}">Column</th><th style="{_TH}">Matched</th><th style="{_TH}">Notes</th></tr>']
    for c in result["columns"]:
        good = c["matched"] == c["total"]
        note = "✅" if good else ", ".join(c["notes"])
        colour = "" if good else "color:#b45309;font-weight:700;"
        h.append(f'<tr><td style="{_TD}font-weight:600">{c["label"]}</td>'
                 f'<td style="{_TD}{colour}">{c["matched"]}/{c["total"]}</td>'
                 f'<td style="{_TD}{colour}">{note}</td></tr>')
    h.append("</table>")
    return "".join(h)


def render_grid_table(result: dict) -> str:
    """Every property x every metric, ours / sheet - the <<GridTable>> token.
    Wrapped in a horizontally scrollable box: ten columns of numbers is wider
    than a phone, and a table that overflows the body is unreadable."""
    if result["status"] != "ok":
        return ""

    h = ['<div style="overflow-x:auto">'
         f'<table style="border-collapse:collapse"><tr><th style="{_TH}">Property</th>']
    for label, _ in METRICS:
        h.append(f'<th style="{_TH}">{label}</th>')
    h.append("</tr>")
    for prop, (short, _) in SHEETS.items():
        h.append(f'<tr><td style="{_TD}font-weight:600;white-space:nowrap">{short}</td>')
        for label, _k in METRICS:
            ov, sv = result["grid"].get(prop, {}).get(label, (None, None))
            if ov is None and sv is None:
                cell, style = "\u2014", "color:#94a3b8;"
            elif ov == sv:
                cell, style = f"\u2713 {ov}", "color:#475569;"
            else:
                cell, style = f"<b>{ov} / {sv}</b>", "background:#fef3c7;color:#92400e;"
            h.append(f'<td style="{_TD}{style}">{cell}</td>')
        h.append("</tr>")
    h.append("</table></div>")
    h.append('<p style="font-size:11px;color:#94a3b8;margin:6px 0 0">Ours / Sheet — ✓ = match</p>')
    return "".join(h)


def render_tokens(result: dict) -> dict:
    """Everything the email template can substitute."""
    day = datetime.strptime(result["date"], "%Y-%m-%d") if result.get("date") else None
    window = result.get("window")
    return {
        "Date": day.strftime("%d/%m/%Y") if day else "\u2014",
        "PropertyCount": str(len(SHEETS)),
        "Matched": str(result.get("matched_cells", "\u2014")),
        "Total": str(result.get("total_cells", "\u2014")),
        "Window": f"{window[0]} \u2013 {window[1]}" if window else "\u2014",
        "SummaryTable": render_summary_table(result),
        "GridTable": render_grid_table(result),
    }


def render_html(result: dict) -> str:
    """Standalone body, used by the CLI's --email flag and as the fallback if
    the Admin template can't be read. The scheduled send builds its body from
    the template instead, substituting render_tokens above."""
    if result["status"] != "ok":
        return f'<pre style="font-family:ui-monospace,monospace;font-size:13px">{render_text(result)}</pre>'

    day = datetime.strptime(result["date"], "%Y-%m-%d")
    perfect = result["matched_cells"] == result["total_cells"]
    banner_bg, banner_fg = ("#dcfce7", "#166534") if perfect else ("#fef9c3", "#854d0e")

    h = [f'''<div style="font-family:-apple-system,Segoe UI,Helvetica,Arial,sans-serif;color:#0f172a;max-width:900px">
<h2 style="margin:0 0 4px">\u0e2a\u0e23\u0e38\u0e1b\u0e15\u0e32\u0e21\u0e04\u0e2d\u0e25\u0e31\u0e21\u0e19\u0e4c \u2014 {_title(result)}</h2>
<div style="background:{banner_bg};color:{banner_fg};padding:8px 12px;border-radius:6px;
 font-weight:700;font-size:14px;margin:10px 0">
 \u0e15\u0e23\u0e07\u0e01\u0e31\u0e19 {result['matched_cells']}/{result['total_cells']} \u0e0a\u0e48\u0e2d\u0e07</div>
<p style="font-size:13px;color:#475569;margin:6px 0 14px">
\u0e0a\u0e35\u0e15\u0e17\u0e31\u0e49\u0e07 {len(SHEETS)} \u0e16\u0e37\u0e2d\u0e02\u0e49\u0e2d\u0e21\u0e39\u0e25\u0e27\u0e31\u0e19\u0e17\u0e35\u0e48 {day.day} {day.strftime('%b %Y')}</p>''']
    h.append(render_summary_table(result))
    h.append('<h3 style="margin:22px 0 8px;font-size:15px">\u0e15\u0e32\u0e23\u0e32\u0e07\u0e40\u0e15\u0e47\u0e21 \u2014 \u0e23\u0e30\u0e1a\u0e1a\u0e40\u0e23\u0e32 / \u0e0a\u0e35\u0e15</h3>')
    h.append(render_grid_table(result))
    h.append("</div>")
    return "".join(h)
