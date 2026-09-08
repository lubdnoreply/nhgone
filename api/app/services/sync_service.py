from supabase import create_client, Client
from app.config import settings
import json
import logging
import re
import unicodedata
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
from typing import Optional
import asyncio
from collections import Counter
from app.services.mews_client import mews_client
from app.services.encryption import encryption_service
from app.services.email_service import (
    email_service, ST_FILES_DAILY_TEMPLATE_KEY, _escape_html,
    DEFAULT_ST_FILES_DAILY_PER_PROPERTY_SUBJECT, DEFAULT_ST_FILES_DAILY_PER_PROPERTY_TEMPLATE,
    RR4_TM30_DAILY_TEMPLATE_KEY,
    DEFAULT_RR4_TM30_DAILY_PER_PROPERTY_SUBJECT, DEFAULT_RR4_TM30_DAILY_PER_PROPERTY_TEMPLATE,
)
from app.services import ftp_service
from app.services.rr4_tm30_reference import RR4_NATIONALITY_CODE, TM30_NATIONALITY_CODE
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from io import BytesIO

logger = logging.getLogger(__name__)

# Characters Unicode's own NFKD decomposition leaves untouched because they
# aren't accented letters - they need an explicit ASCII stand-in. MEWS writes
# order-item descriptions as "{qty} × {name}", so U+00D7 is by far the most
# common one in practice; the dashes and quotes show up in free-text product
# names typed by staff.
_ASCII_SUBSTITUTES = {
    "×": "x",    # × multiplication sign (MEWS quantity separator)
    "–": "-", "—": "-",           # en/em dash
    "‘": "'", "’": "'",           # curly single quotes
    "“": '"', "”": '"',           # curly double quotes
    "…": "...",                        # ellipsis
    " ": " ",                          # non-breaking space
    "°": " deg", "€": "EUR", "£": "GBP", "¥": "JPY",
}


def _ascii_fold(text: str) -> str:
    """Reduce a string to plain 7-bit ASCII for the pipe-delimited SunSystems
    journal files.

    The interface reads these as a flat file in a single-byte encoding, so a
    UTF-8 multi-byte character does not survive the round trip - and because
    the import profile posts with "Post if no errors", one bad character
    rejects the entire day's journal rather than just its own line. MEWS's
    own reference export (PT_RV_20260813.csv, confirmed by `file` as plain
    ASCII "CSV text" where ours came out "UTF-8 text") never emits a
    non-ASCII byte, so matching that is the safe target.

    Accented letters fold through NFKD (é -> e); anything Unicode won't
    decompose gets an explicit substitute from _ASCII_SUBSTITUTES first, so
    that "-1 × Rounding adjustment" becomes "-1 x Rounding adjustment"
    rather than losing the separator entirely. Whatever is still non-ASCII
    after both passes is dropped - a garbled description is recoverable,
    a rejected journal is not.
    """
    if not text:
        return ""
    for ch, repl in _ASCII_SUBSTITUTES.items():
        if ch in text:
            text = text.replace(ch, repl)
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


_THAI_NUM = ["ศูนย์", "หนึ่ง", "สอง", "สาม", "สี่", "ห้า", "หก", "เจ็ด", "แปด", "เก้า"]
_THAI_UNIT = ["", "สิบ", "ร้อย", "พัน", "หมื่น", "แสน"]


def _read_thai_digits(num_str: str) -> str:
    num_str = num_str.lstrip("0")
    if not num_str:
        return ""
    n = len(num_str)
    result = ""
    for i, ch in enumerate(num_str):
        digit = int(ch)
        pos = n - i - 1
        if digit == 0:
            continue
        if pos == 0 and digit == 1 and n > 1:
            result += "เอ็ด"
        elif pos == 1 and digit == 2:
            result += "ยี่สิบ"
        elif pos == 1 and digit == 1:
            result += "สิบ"
        else:
            result += _THAI_NUM[digit] + _THAI_UNIT[pos]
    return result


def bahttext(amount) -> str:
    """Convert a numeric amount to Thai baht text, e.g. 1250.50 -> 'หนึ่งพันสองร้อยห้าสิบบาทห้าสิบสตางค์'."""
    try:
        amount = float(amount or 0)
    except (TypeError, ValueError):
        amount = 0.0
    amount = round(abs(amount) + 1e-9, 2)
    baht = int(amount)
    satang = int(round((amount - baht) * 100))

    if baht == 0:
        baht_words = "ศูนย์"
    else:
        groups = []
        s = str(baht)
        while s:
            groups.insert(0, s[-6:])
            s = s[:-6]
        pieces = []
        for idx, g in enumerate(groups):
            words = _read_thai_digits(g)
            if words:
                scale_ups = len(groups) - idx - 1
                pieces.append(words + "ล้าน" * scale_ups)
        baht_words = "".join(pieces)

    result = baht_words + "บาท"
    if satang == 0:
        result += "ถ้วน"
    else:
        result += _read_thai_digits(str(satang)) + "สตางค์"
    return result


# RR3 (ร.ร.๓ Thai Hotel Act lodger registration card) support - ported from the
# user's existing, proven Google Apps Script rather than re-derived from docs.
# Full ISO 3166-1 alpha-2 list (matches Mews's own NationalityCode/
# BirthCountryCode/Address.CountryCode fields, all documented as ISO 3166-1)
# rather than the small hand-picked subset this used to be - a guest whose
# code fell outside that subset used to show the raw 2-letter code on the
# printed RR3 form and Guest Profile instead of a real country name. Also
# backs the BCP Edit Guest dropdown's Nationality/Country of birth options
# on the frontend (GUEST_COUNTRY_OPTIONS in bcp/page.tsx) - keep both lists
# in sync if either changes.
_RR3_COUNTRY_MAP = {
    "AF": "Afghanistan", "AX": "Åland Islands", "AL": "Albania", "DZ": "Algeria",
    "AS": "American Samoa", "AD": "Andorra", "AO": "Angola", "AI": "Anguilla",
    "AQ": "Antarctica", "AG": "Antigua and Barbuda", "AR": "Argentina", "AM": "Armenia",
    "AW": "Aruba", "AU": "Australia", "AT": "Austria", "AZ": "Azerbaijan",
    "BS": "Bahamas", "BH": "Bahrain", "BD": "Bangladesh", "BB": "Barbados",
    "BY": "Belarus", "BE": "Belgium", "BZ": "Belize", "BJ": "Benin",
    "BM": "Bermuda", "BT": "Bhutan", "BO": "Bolivia", "BQ": "Bonaire, Sint Eustatius and Saba",
    "BA": "Bosnia and Herzegovina", "BW": "Botswana", "BV": "Bouvet Island", "BR": "Brazil",
    "IO": "British Indian Ocean Territory", "BN": "Brunei", "BG": "Bulgaria", "BF": "Burkina Faso",
    "BI": "Burundi", "CV": "Cabo Verde", "KH": "Cambodia", "CM": "Cameroon",
    "CA": "Canada", "KY": "Cayman Islands", "CF": "Central African Republic", "TD": "Chad",
    "CL": "Chile", "CN": "China", "CX": "Christmas Island", "CC": "Cocos (Keeling) Islands",
    "CO": "Colombia", "KM": "Comoros", "CD": "Congo (Democratic Republic)", "CG": "Congo (Republic)",
    "CK": "Cook Islands", "CR": "Costa Rica", "CI": "Côte d'Ivoire", "HR": "Croatia",
    "CU": "Cuba", "CW": "Curaçao", "CY": "Cyprus", "CZ": "Czechia",
    "DK": "Denmark", "DJ": "Djibouti", "DM": "Dominica", "DO": "Dominican Republic",
    "EC": "Ecuador", "EG": "Egypt", "SV": "El Salvador", "GQ": "Equatorial Guinea",
    "ER": "Eritrea", "EE": "Estonia", "SZ": "Eswatini", "ET": "Ethiopia",
    "FK": "Falkland Islands", "FO": "Faroe Islands", "FJ": "Fiji", "FI": "Finland",
    "FR": "France", "GF": "French Guiana", "PF": "French Polynesia", "TF": "French Southern Territories",
    "GA": "Gabon", "GM": "Gambia", "GE": "Georgia", "DE": "Germany",
    "GH": "Ghana", "GI": "Gibraltar", "GR": "Greece", "GL": "Greenland",
    "GD": "Grenada", "GP": "Guadeloupe", "GU": "Guam", "GT": "Guatemala",
    "GG": "Guernsey", "GN": "Guinea", "GW": "Guinea-Bissau", "GY": "Guyana",
    "HT": "Haiti", "HM": "Heard Island and McDonald Islands", "VA": "Holy See", "HN": "Honduras",
    "HK": "Hong Kong", "HU": "Hungary", "IS": "Iceland", "IN": "India",
    "ID": "Indonesia", "IR": "Iran", "IQ": "Iraq", "IE": "Ireland",
    "IM": "Isle of Man", "IL": "Israel", "IT": "Italy", "JM": "Jamaica",
    "JP": "Japan", "JE": "Jersey", "JO": "Jordan", "KZ": "Kazakhstan",
    "KE": "Kenya", "KI": "Kiribati", "KP": "North Korea", "KR": "South Korea",
    "KW": "Kuwait", "KG": "Kyrgyzstan", "LA": "Laos", "LV": "Latvia",
    "LB": "Lebanon", "LS": "Lesotho", "LR": "Liberia", "LY": "Libya",
    "LI": "Liechtenstein", "LT": "Lithuania", "LU": "Luxembourg", "MO": "Macao",
    "MG": "Madagascar", "MW": "Malawi", "MY": "Malaysia", "MV": "Maldives",
    "ML": "Mali", "MT": "Malta", "MH": "Marshall Islands", "MQ": "Martinique",
    "MR": "Mauritania", "MU": "Mauritius", "YT": "Mayotte", "MX": "Mexico",
    "FM": "Micronesia", "MD": "Moldova", "MC": "Monaco", "MN": "Mongolia",
    "ME": "Montenegro", "MS": "Montserrat", "MA": "Morocco", "MZ": "Mozambique",
    "MM": "Myanmar", "NA": "Namibia", "NR": "Nauru", "NP": "Nepal",
    "NL": "Netherlands", "NC": "New Caledonia", "NZ": "New Zealand", "NI": "Nicaragua",
    "NE": "Niger", "NG": "Nigeria", "NU": "Niue", "NF": "Norfolk Island",
    "MK": "North Macedonia", "MP": "Northern Mariana Islands", "NO": "Norway", "OM": "Oman",
    "PK": "Pakistan", "PW": "Palau", "PS": "Palestine", "PA": "Panama",
    "PG": "Papua New Guinea", "PY": "Paraguay", "PE": "Peru", "PH": "Philippines",
    "PN": "Pitcairn", "PL": "Poland", "PT": "Portugal", "PR": "Puerto Rico",
    "QA": "Qatar", "RO": "Romania", "RU": "Russia", "RW": "Rwanda",
    "BL": "Saint Barthélemy", "KN": "Saint Kitts and Nevis", "LC": "Saint Lucia", "MF": "Saint Martin",
    "PM": "Saint Pierre and Miquelon", "VC": "Saint Vincent and the Grenadines", "WS": "Samoa", "SM": "San Marino",
    "ST": "São Tomé and Príncipe", "SA": "Saudi Arabia", "SN": "Senegal", "RS": "Serbia",
    "SC": "Seychelles", "SL": "Sierra Leone", "SG": "Singapore", "SX": "Sint Maarten",
    "SK": "Slovakia", "SI": "Slovenia", "SB": "Solomon Islands", "SO": "Somalia",
    "ZA": "South Africa", "GS": "South Georgia and South Sandwich Islands", "SS": "South Sudan", "ES": "Spain",
    "LK": "Sri Lanka", "SD": "Sudan", "SR": "Suriname", "SJ": "Svalbard and Jan Mayen",
    "SE": "Sweden", "CH": "Switzerland", "SY": "Syria", "TW": "Taiwan",
    "TJ": "Tajikistan", "TZ": "Tanzania", "TH": "Thailand", "TL": "Timor-Leste",
    "TG": "Togo", "TK": "Tokelau", "TO": "Tonga", "TT": "Trinidad and Tobago",
    "TN": "Tunisia", "TR": "Turkey", "TM": "Turkmenistan", "TC": "Turks and Caicos Islands",
    "TV": "Tuvalu", "UG": "Uganda", "UA": "Ukraine", "AE": "United Arab Emirates",
    "GB": "United Kingdom", "US": "United States", "UM": "United States Minor Outlying Islands", "UY": "Uruguay",
    "UZ": "Uzbekistan", "VU": "Vanuatu", "VE": "Venezuela", "VN": "Vietnam",
    "VG": "Virgin Islands (British)", "VI": "Virgin Islands (U.S.)", "WF": "Wallis and Futuna", "EH": "Western Sahara",
    "YE": "Yemen", "ZM": "Zambia", "ZW": "Zimbabwe",
}

_RR3_PROPERTY_THAI_NAMES = {
    "Lub d Bangkok Chinatown": "หลับดี แบงค็อก เยาวราช",
    "Lub d Bangkok Siam": "หลับดี แบงค็อก สยาม",
    "Lub d Koh Samui Chaweng Beach": "หลับดี เกาะสมุย หาดเฉวง",
    "Lub d Koh Tao Tanote Bay": "หลับดี เกาะเต่า อ่าวโตนด",
    "Lub d Philippines Makati": "หลับดี มะนิลา มาคาติ",
    "Lub d Phuket Patong": "หลับดี ภูเก็ต ป่าตอง",
    "Lub d Siem Reap": "หลับดี เสียมเรียบ",
    "Marasca Samui": "มาราสก้า สมุย",
}

# Local hour before which a zero-night ("day use") stay is read as the tail of
# the PREVIOUS night rather than a daytime day room. See the block in
# get_st_files_report that uses it for the evidence on both sides.
#
# Bounded, not derived: a 02:00 start is a night tail (Siem Reap #149736/#149737,
# 26-Aug-2026) so this must be > 2, and a 04:30 start is a day room (Patong
# #190439/#190440, 30-Aug-2026 - both counted as arrivals by MEWS's own export)
# so it must be <= 4. That leaves 3 or 4; 4 is chosen as the wider reading of
# "still night". The 30-Aug pair is what pulled this down from 6, which had
# been inferred against a much looser upper bound of 10:00 and was costing
# Patong two arrivals a day.
_ST_DAY_USE_NIGHT_END_HOUR = 4

# ...but the hour alone was never the whole rule, and a WALK-IN overrides it.
#
# A guest who walks up to the desk at 02:00 and takes a room for the day is an
# arrival however early it is - MEWS counts them - while a booking that merely
# ENDS in those hours is the tail of the night before. The rate name is what
# separates the two, and it separates them better than any hour can: scored
# against all 14 day-use stays whose true answer is known from MEWS's own
# per-category Arrivals export (26-Aug, 30-Aug and 04-Sep-2026), the hour rule
# alone gets 10 and "walk-in rate OR hour >= 4" gets 13 - and the three it adds
# are strictly extra, its misses being a subset of the hour rule's.
#
# 04-Sep-2026 is what forced this: Siam filed 24 arrivals to our 22 and Siem
# Reap 117 to our 116, and the sheets' own Arrivals tabs put every unit of both
# gaps on one category each (Siam UPG 2/0, Siem Reap DKR 21/20). Those were
# exactly three 'Walk In Room Only' day rooms starting 02:00, 02:00 and 00:36 -
# while the OTA and Advance Purchase stays alongside them, starting 01:28,
# 00:10 and 00:18, were correctly left out by BOTH sides.
#
# The one case still unexplained is the documented irreducible pair: Chinatown
# #96148 and #96160, 30-Aug-2026, both 00:00, both "Static Room Only B2B",
# neither a walk-in, and MEWS counted the first and not the second. No field on
# the two reservations separates them, so no rule expressible here can.
#
# Matched on the rate NAME because that is what the property actually types;
# the three live variants are "Walk In Room Only", "Walk In Room with
# Breakfast" and "Walk In with Breakfast", and nothing else in any property's
# rate list contains the word. Deliberately NOT the business segment: "Direct
# Host" covers walk-ins but also covers Extend Stay and B2B rows that MEWS
# leaves out (Siem Reap #149737, Chinatown #96160), and scored worse.
#
# Word-bounded so a rate that merely CONTAINS the letters can't trip it - an
# unanchored search matched "Sidewalk Inn". Nothing like that is in any
# property's rate list today, but a rate name is free text somebody types.
_ST_WALK_IN_RATE_RE = re.compile(r"\bwalk[\s-]*in\b", re.IGNORECASE)

# Offline test seam for the ST arrivals rule - see where it is read in
# get_st_files_report. Production leaves this None and nothing in the app ever
# assigns it; scripts/st_arrival_sweep.py sets it to score candidate rules
# against the real sheets using the report's own category/space-unit
# resolution instead of a re-implementation.
_ST_ARRIVAL_RULE = None

# The RR4 (ร.ร.๔) hotel-register filing needs the property's Thai REGISTERED
# name, which is not always the same string as _RR3_PROPERTY_THAI_NAMES above
# (that one backs guest-facing RR3 cards, a different, less formal use) -
# confirmed distinct for Chinatown ("หลับดี บางกอก ไชน่าทาว์น" here vs
# "หลับดี แบงค็อก เยาวราช" there) against the real RR4 file it already
# produces. The other 7 properties fall back to the RR3 name below until
# their own exact registered name is supplied and added here.
_RR4_PROPERTY_THAI_NAMES = {
    "Lub d Bangkok Chinatown": "หลับดี บางกอก ไชน่าทาว์น",
    "Lub d Bangkok Siam": "โรงแรมหลับดี สยาม",
}


def _rr3_country_name(code: str) -> str:
    return _RR3_COUNTRY_MAP.get(code, code or "")


# RR4 only: MEWS's own "Customer profiles" export writes the full formal
# country name in the address / come-from columns, where _RR3_COUNTRY_MAP
# (built for the ร.ร.๓ card, and still used verbatim there and on the Guest
# Profile page) carries the short colloquial one.
#
# Derived from the real MEWS exports pasted into all six generator sheets for
# 19-Aug-2026 (Chinatown, Siam, Koh Samui, Patong, Koh Tao, Marasca Samui).
# Those pastes between them carry 66 distinct nationality strings; 62 already
# came out of _RR3_COUNTRY_MAP verbatim and the rest are listed here, each one
# additionally agreed on unanimously by all six sheets' own RR4-Nationality
# lookup tables.
#
# Deliberately an override of specific codes, NOT a blanket switch to ISO
# 3166-1 formal names: MEWS does not follow ISO consistently (it writes plain
# "Taiwan", not ISO's "Taiwan, Province of China", and appends the old name in
# "Türkiye (Turkey)"), so applying ISO wholesale would break countries that are
# correct today. MEWS exposes no countries endpoint to derive the rest from
# (countries/getAll 401s on the property tokens), so anything unlisted keeps
# its _RR3_COUNTRY_MAP name.
#
# Per-property TM30 arrival-day start, as each generator sheet's own
# Parameter-ImportCP tab declares it. Only a stopgap default: the DB columns
# tm30_day_start_hour/_minute (api/sql/tm30_day_window.sql) override this, and
# are what should actually be edited once that migration has been run - these
# windows change on the sheet side without warning, so a hardcoded value will
# go stale. See _resolve_tm30_day_start.
#
# Chinatown is the only property whose sheet exports TM30 over anything but
# plain midnight; the other five declare 00:00 and are left at the default.
# NOTE the consequence, measured against the real 29-Aug-2026 sheet: a 12:15
# window yields 56 of that sheet's 67 guests, where plain midnight yields all
# 67 plus 9 more. The 9 extra are real arrivals MEWS timestamps StartUtc
# 00:15 while their ActualStartUtc is 00:34-01:41; the 11 lost are guests
# checking in between midnight and 12:15 who ARE on the filed sheet. This is
# configured to match the sheet by explicit instruction - matching the filed
# document is the standing rule for this module - but it does mean TM30 files
# fewer foreign arrivals than MEWS reports for the day, which is worth
# re-testing whenever the sheet's window changes.
#
# SETTLED 05-Sep-2026, after the same question came up a third time. The five
# non-Chinatown properties had at some point been configured to the ~02:00
# times their sheets DECLARE in Master!B2, on the reasoning that matching the
# declared window would match the numbers. Measured across all six sheets for
# 04-Sep-2026 it does the opposite: the declared windows reconciled on 2 of 6
# properties, plain midnight on 4 of 6, and midnight made no property worse.
# The sheets declare a window they do not actually filter by - Chinatown's
# said 12:15 while holding 58 arrivals to our 53, Patong's said 02:05 while
# agreeing with us either way. So the five are back on midnight and only
# Chinatown keeps a shifted window, again by explicit instruction.
#
# Two guests are what this cost per day: Siam's PEI FANG LEE (StartUtc
# 2026-09-03T18:28Z = 01:28 local) and Samui's Pulkit Chaudhary
# (2026-09-03T17:06Z = 00:06 local), both on their sheets, both dropped by an
# 02:0x lower bound and both recovered by midnight.
#
# RE-MEASURED AND REVERSED 08-Sep-2026: the five are back on 02:00, because
# the same measurement run against 07-Sep-2026 came out the other way -
# 02:00 reconciled 5 of 5 comparable properties exactly (Siam 16/16, Samui
# 87/87, Koh Tao 13/13, Patong 53/53, Marasca 26/26) where midnight managed
# 3 of 5.
#
# Both results are real, and the reason they disagree is that the window
# start moves BOTH ends. A guest arriving 00:00-02:00 on day D lands in D
# under midnight and in D-1 under 02:00; a guest arriving 00:00-02:00 on day
# D+1 lands in D+1 under midnight and in D under 02:00. 04-Sep's boundary
# guests were the first kind and 07-Sep's were the second (Siam's two at
# 08-Sep 02:00 and Patong's Anton Kulikov at 08-Sep 00:17, all three on the
# sheet's 07-Sep file and all three missed by a midnight upper bound).
#
# So this WILL look wrong again on a day whose boundary guests fall the other
# way. Re-measure before moving it, on more than one day if possible, and
# expect neither value to be right every morning - the sheets' own ~02:0x
# declared start is what the current setting is tracking.
_TM30_DAY_START_FALLBACK = {
    "Lub d Bangkok Chinatown": (12, 15),
}

# Three entries in those lookup tables were deliberately NOT copied here: they
# label Georgia "Georgian", Guinea "Guine" and Somalia "Somali" (demonyms and a
# typo, none of which MEWS has ever actually emitted), and they give RR4 code
# 104 - North Korea - the name "Korea (Republic of)", which is South Korea.
# Add a code here only with a real MEWS export to confirm it against.
_RR4_MEWS_COUNTRY_NAMES = {
    # Confirmed 2026-08-26 from Siam's own ImportInhouse paste (the MEWS
    # export itself), which writes "Brunei Darussalam" in both Nationality
    # and Country where _RR3_COUNTRY_MAP carries the short "Brunei". Missed
    # by the original 19-Aug sample only because no Brunei guest was in it.
    "BN": "Brunei Darussalam",
    "BO": "Bolivia (Plurinational State of)",
    "CZ": "Czech Republic",
    "GB": "United Kingdom of Great Britain and Northern Ireland",
    "IR": "Iran (Islamic Republic of)",
    "KR": "Korea (Republic of)",
    "LA": "Lao People's Democratic Republic",
    "MD": "Moldova (Republic of)",
    "MK": "Republic of North Macedonia",
    "PS": "Palestine (State of)",
    "RU": "Russian Federation",
    "SY": "Syrian Arab Republic",
    "TR": "Türkiye (Turkey)",
    "US": "United States of America",
}


def _rr4_country_name(code: str) -> str:
    return _RR4_MEWS_COUNTRY_NAMES.get(code) or _rr3_country_name(code)


# MEWS's Customer.Title is an enum ("Mister", "Missis", ...) - MEWS's own
# Profile screen displays these as the short form ("Mr.", "Mrs.", ...), so
# the Guest Profile page matches that instead of showing the raw enum value.
# Falls back to the raw value for anything not in this list, so an unmapped
# title still shows something rather than silently disappearing.
_MEWS_TITLE_DISPLAY = {
    "Mister": "Mr.",
    "Miss": "Miss",
    "Missis": "Mrs.",
    "Ms": "Ms.",
    "Doctor": "Dr.",
}


def _mews_title_display(title) -> str:
    if not title:
        return ""
    return _MEWS_TITLE_DISPLAY.get(title, title)


_RR3_MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _rr3_format_thai_date(utc_str: str) -> str:
    if not utc_str:
        return ""
    try:
        dt = datetime.fromisoformat(utc_str.replace("Z", "+00:00")).astimezone(ZoneInfo("Asia/Bangkok"))
        # dd/mmm/yyyy (e.g. 28/Jul/2026) - a fixed abbreviation list instead of
        # strftime("%b") avoids the month name depending on the server's locale.
        return f"{dt.day:02d}/{_RR3_MONTH_ABBR[dt.month - 1]}/{dt.year}"
    except Exception:
        return ""


def _rr3_format_thai_time(utc_str: str) -> str:
    if not utc_str:
        return ""
    try:
        dt = datetime.fromisoformat(utc_str.replace("Z", "+00:00")).astimezone(ZoneInfo("Asia/Bangkok"))
        return dt.strftime("%H:%M")
    except Exception:
        return ""


class SyncService:
    def __init__(self):
        # We need the SERVICE_ROLE_KEY to bypass RLS for sync operations
        # If not provided, we fall back to the anon key (might fail RLS)
        self.url = settings.SUPABASE_URL
        self.key = settings.SUPABASE_SERVICE_ROLE_KEY or settings.SUPABASE_ANON_KEY
        if self.url and self.key:
            self.supabase: Client = create_client(self.url, self.key)
        else:
            self.supabase = None
            logger.warning("Supabase credentials missing. Sync will not work.")

    _DEFAULT_RETRY_SETTINGS = {"retry_count": 2, "retry_interval_minutes": 60}

    async def get_sync_retry_settings(self) -> dict:
        """
        Global policy for main.py's retry_scheduled_syncs: how many times,
        and how many minutes apart, a property's Data Mart sync (Reservations/
        Customers/Payments/Bills/Resources) is auto-retried after its own
        scheduled run if a table is still missing or errored that day - e.g.
        the default (2 retries, 60 min apart) fires at sync_time+1h and
        sync_time+2h. Editable at Admin > Sync (runs on its own 5-minute
        cron - see retry_scheduled_syncs' docstring on why minutes needed a
        dedicated cron instead of piggybacking on the hourly one). Falls
        back to the default if no row has been saved yet, or the table
        doesn't exist (migration not run) - retries keep working with the
        old hardcoded behavior either way.
        """
        if not self.supabase:
            return dict(self._DEFAULT_RETRY_SETTINGS)
        try:
            res = self.supabase.table("sync_retry_settings").select(
                "retry_count, retry_interval_minutes").limit(1).execute()
        except Exception as e:
            logger.warning(f"sync_retry_settings lookup failed, using defaults: {e}")
            return dict(self._DEFAULT_RETRY_SETTINGS)
        if not res.data:
            return dict(self._DEFAULT_RETRY_SETTINGS)
        row = res.data[0]
        return {
            "retry_count": row.get("retry_count") or self._DEFAULT_RETRY_SETTINGS["retry_count"],
            "retry_interval_minutes": row.get("retry_interval_minutes") or self._DEFAULT_RETRY_SETTINGS["retry_interval_minutes"],
        }

    async def sync_reservation(self, data: dict):
        if not self.supabase:
            raise Exception("Supabase client not initialized")
        
        # Upsert pattern preserving NHGOne notes
        # We use a POST request via supabase client
        try:
            res = self.supabase.table("reservations").upsert(
                data,
                on_conflict="mews_id"
            ).execute()
            return res.data
        except Exception as e:
            logger.error(f"Sync error: {e}")
            raise e

    async def get_mapped_reservations(self, property_name: str, start_date: str = None, end_date: str = None, cursor: str = None, chunk_limit: int = None):
        """
        Fetch live reservations and map them to the 58 columns Mews Reservation Report.
        Shared between API router and background sync job.
        """
        try:
            if not start_date or not end_date:
                # Default to Yesterday 00:01:00 to 23:59:59 (Asia/Bangkok time), exported as UTC for Mews API
                bkk_tz = ZoneInfo("Asia/Bangkok")
                now_bkk = datetime.now(bkk_tz)
                yesterday_bkk = now_bkk - timedelta(days=1)
                
                if not start_date:
                    start_dt = yesterday_bkk.replace(hour=0, minute=0, second=0, microsecond=0)
                    start_date = start_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
                if not end_date:
                    end_dt = yesterday_bkk.replace(hour=23, minute=59, second=59, microsecond=999999)
                    end_date = end_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

            chunk_size = 500
            
            all_reservations = []
            current_cursor = cursor
            chunks_fetched = 0

            while True:
                payload = {
                    "CollidingUtc": {"StartUtc": start_date, "EndUtc": end_date},
                    "States": ["Canceled", "Started", "Processed", "Confirmed", "Inquired", "Optional"],
                    "Limitation": {"Count": chunk_size}
                }
                if current_cursor:
                    payload["Limitation"]["Cursor"] = current_cursor

                # 1. Fetch Reservations
                response_data = await mews_client.post("/api/connector/v1/reservations/getAll/2023-06-06", payload, property_name=property_name)
                reservations_chunk = response_data.get("Reservations", [])
                current_cursor = response_data.get("Cursor")
                
                all_reservations.extend(reservations_chunk)
                chunks_fetched += 1
                
                if not current_cursor or not reservations_chunk or (chunk_limit and chunks_fetched >= chunk_limit):
                    break

            reservations = all_reservations
            return_cursor = current_cursor

            if not reservations:
                return {"data": [], "cursor": None}

            # 2. Collect unique IDs for relations
            account_ids = {r.get("AccountId") for r in reservations if r.get("AccountId")}
            booker_ids = {r.get("BookerId") for r in reservations if r.get("BookerId")}
            customer_ids = list(account_ids.union(booker_ids))
            
            company_ids = {r.get("CompanyId") for r in reservations if r.get("CompanyId")}
            ta_ids = {r.get("TravelAgencyId") for r in reservations if r.get("TravelAgencyId")}
            all_company_ids = list(company_ids.union(ta_ids))
            
            resource_ids = list({r.get("AssignedResourceId") for r in reservations if r.get("AssignedResourceId")})
            category_ids = list({r.get("RequestedCategoryId") for r in reservations if r.get("RequestedCategoryId")})
            rate_ids = list({r.get("RateId") for r in reservations if r.get("RateId")})
            group_ids = list({r.get("ReservationGroupId") for r in reservations if r.get("ReservationGroupId")})

            # 3. Fetch Relations Concurrently
            async def fetch_entity(endpoint, payload_key, ids, response_key):
                if not ids: return {}
                try:
                    res = await mews_client.post(endpoint, {payload_key: ids[:200]}, property_name=property_name)
                    return {item["Id"]: item for item in res.get(response_key, [])}
                except Exception as e:
                    logger.error(f"Failed to fetch {endpoint}: {e}")
                    return {}

            customers_dict, companies_dict, resources_dict, categories_dict, rates_dict, groups_dict = await asyncio.gather(
                fetch_entity("/api/connector/v1/customers/getAll", "CustomerIds", customer_ids, "Customers"),
                # companies/getAll filters by "Ids", not "CompanyIds" - the
                # latter is silently accepted but ignored, so it was
                # returning an arbitrary/unfiltered batch of companies
                # instead of the ones actually requested (confirmed live:
                # asking for one specific TravelAgencyId came back with 10
                # unrelated companies, none matching).
                fetch_entity("/api/connector/v1/companies/getAll", "Ids", all_company_ids, "Companies"),
                fetch_entity("/api/connector/v1/resources/getAll", "ResourceIds", resource_ids, "Resources"),
                fetch_entity("/api/connector/v1/resourceCategories/getAll", "ResourceCategoryIds", category_ids, "ResourceCategories"),
                fetch_entity("/api/connector/v1/rates/getAll", "RateIds", rate_ids, "Rates"),
                fetch_entity("/api/connector/v1/reservationGroups/getAll", "ReservationGroupIds", group_ids, "ReservationGroups")
            )

            mapped_data = []

            def get_date(utc_str):
                if not utc_str: return ""
                return utc_str.replace("T", " ")[:19]

            for res_item in reservations:
                c = customers_dict.get(res_item.get("AccountId"), {})
                b = customers_dict.get(res_item.get("BookerId"), {})
                res = resources_dict.get(res_item.get("AssignedResourceId"), {})
                cat = categories_dict.get(res_item.get("RequestedCategoryId"), {})
                comp = companies_dict.get(res_item.get("CompanyId"), {})
                ta = companies_dict.get(res_item.get("TravelAgencyId"), {})
                grp = groups_dict.get(res_item.get("ReservationGroupId"), {})
                rate = rates_dict.get(res_item.get("RateId"), {})

                start = res_item.get("StartUtc")
                end = res_item.get("EndUtc")
                nights = ""
                if start and end:
                    try:
                        s_date = datetime.fromisoformat(start.replace("Z", "+00:00"))
                        e_date = datetime.fromisoformat(end.replace("Z", "+00:00"))
                        nights = (e_date - s_date).days
                    except:
                        pass
                        
                p_counts = res_item.get("PersonCounts") or []
                total_persons = sum(pc.get("Count", 0) for pc in p_counts)

                row = {
                    "Number": res_item.get("Number", ""),
                    "Status": {"Started": "Checked in", "Processed": "Checked out"}.get(res_item.get("State"), res_item.get("State", "")),
                    "Arrival": get_date(res_item.get("StartUtc")),
                    "Departure": get_date(res_item.get("EndUtc")),
                    "Last name": c.get("LastName", ""),
                    "First name": c.get("FirstName", ""),
                    "Email": c.get("Email", ""),
                    "Telephone": c.get("Phone", ""),
                    "Group name": grp.get("Name", ""),
                    "Address": c.get("Address", {}).get("Line1", "") if isinstance(c.get("Address"), dict) else "",
                    "Customer nationality": c.get("NationalityCode", ""),
                    "Send marketing emails": "", 
                    "Booker": f'{b.get("FirstName", "")} {b.get("LastName", "")}'.strip(),
                    "Creator": res_item.get("CreatorProfileId", ""),
                    "Created": get_date(res_item.get("CreatedUtc")),
                    "Release": get_date(res_item.get("ReleasedUtc")),
                    "Confirmed": get_date(res_item.get("UpdatedUtc")), 
                    "Canceled": get_date(res_item.get("CancelledUtc")),
                    "Count (nights)": nights,
                    "Person count": total_persons,
                    "Count (bed, nightly)": "",
                    "Requested category": cat.get("Name", ""),
                    "Space category": "", 
                    "Space number": res.get("Name", ""),
                    "Origin": res_item.get("Origin", ""),
                    "Channel manager ID": res_item.get("ChannelManagerId", ""),
                    "Group channel manager ID": "",
                    "Group channel confirmation number": "",
                    "Travel agency confirmation number": "",
                    "Segment": res_item.get("BusinessSegmentId", ""),
                    "Rate": rate.get("Name", ""),
                    "Voucher": res_item.get("VoucherId", ""),
                    "Products": "",
                    "Company": comp.get("Name", ""),
                    "Travel agency": ta.get("Name", ""),
                    "Average rate (nightly)": "",
                    "Total amount": res_item.get("RequestedPaymentAmount", {}).get("Value", "") if isinstance(res_item.get("RequestedPaymentAmount"), dict) else "",
                    "Canceled cost": "",
                    "Commission": "",
                    "Customer cost": "",
                    "Balance of companions": "",
                    "Payment card type": "",
                    "Payment card number": "",
                    "Expiration": "",
                    "Automatic payment": "",
                    "Bills": "",
                    "Cancellation reason": res_item.get("CancellationReason", ""),
                    "Notes": res_item.get("Notes", ""),
                    "Customer notes": "", 
                    "Customer classifications": "",
                    "Pricing classification": "",
                    "Booking purpose": "",
                    "Reservation source": res_item.get("Origin", ""),
                    "Identifier": res_item.get("Id", ""),
                    "Company Identifier": res_item.get("CompanyId", ""),
                    "Travel agency Identifier": res_item.get("TravelAgencyId", ""),
                    "Reservation origin details": res_item.get("OriginDetails", ""),
                    "Restoration reason": ""
                }
                mapped_data.append(row)

            return {
                "data": mapped_data,
                "cursor": return_cursor
            }
        except Exception as e:
            logger.error(f"Error mapping reservations for {property_name}: {str(e)}")
            raise e

    async def get_mapped_members(self, property_name: str, start_date: str = None, end_date: str = None):
        """
        Fetch members (customers) directly from MEWS customers/getAll API
        using UpdatedUtc date range filter.
        """
        try:
            if not start_date or not end_date:
                # Default to Yesterday 00:01:00 to 23:59:59 (Asia/Bangkok time), exported as UTC for Mews API
                bkk_tz = ZoneInfo("Asia/Bangkok")
                now_bkk = datetime.now(bkk_tz)
                yesterday_bkk = now_bkk - timedelta(days=1)
                
                if not start_date:
                    start_dt = yesterday_bkk.replace(hour=0, minute=0, second=0, microsecond=0)
                    start_date = start_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
                if not end_date:
                    end_dt = yesterday_bkk.replace(hour=23, minute=59, second=59, microsecond=999999)
                    end_date = end_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

            # Fetch customers directly using UpdatedUtc date filter with pagination
            all_customers = []
            current_cursor = None
            chunk_size = 500

            while True:
                cust_payload = {
                    "UpdatedUtc": {
                        "StartUtc": start_date,
                        "EndUtc": end_date
                    },
                    "Extent": {
                        "Customers": True,
                        "Documents": False,
                        "Addresses": False
                    },
                    "ActivityStates": ["Active"],
                    "Limitation": {"Count": chunk_size}
                }
                if current_cursor:
                    cust_payload["Limitation"]["Cursor"] = current_cursor

                cust_res = await mews_client.post(
                    "/api/connector/v1/customers/getAll",
                    cust_payload,
                    property_name=property_name
                )
                chunk = cust_res.get("Customers", [])
                current_cursor = cust_res.get("Cursor")
                all_customers.extend(chunk)

                if not current_cursor or not chunk:
                    break

            mapped_members = []
            for cust in all_customers:
                mapped_members.append({
                    "Number": cust.get("Number", ""),
                    "Title": cust.get("Title", ""),
                    "Last Name": cust.get("LastName", ""),
                    "First Name": cust.get("FirstName", ""),
                    "Second Last Name": cust.get("SecondLastName", ""),
                    "Nationality": cust.get("NationalityCode", ""),
                    "Preferred Language": cust.get("PreferredLanguageCode", ""),
                    "Language": cust.get("LanguageCode", ""),
                    "Birth Date": cust.get("BirthDate", ""),
                    "Birth Place": cust.get("BirthPlace", ""),
                    "Occupation": cust.get("Occupation", ""),
                    "Email": cust.get("Email", ""),
                    "Phone": cust.get("Phone", ""),
                    "Tax ID": cust.get("TaxIdentificationNumber", ""),
                    "Loyalty Code": cust.get("LoyaltyCode", ""),
                    "Accounting Code": cust.get("AccountingCode", ""),
                    "Billing Code": cust.get("BillingCode", ""),
                    "Car Registration": cust.get("CarRegistrationNumber", ""),
                    "Dietary": cust.get("DietaryRequirements", ""),
                    "Notes": cust.get("Notes", ""),
                    "Created": cust.get("CreatedUtc", ""),
                    "Updated": cust.get("UpdatedUtc", ""),
                    "Active": cust.get("IsActive", True),
                    "Classifications": ", ".join(cust.get("Classifications", [])) if cust.get("Classifications") else "",
                    "Options": ", ".join(cust.get("Options", [])) if cust.get("Options") else "",
                    "Identifier": cust.get("Id", ""),
                    "mews_id": cust.get("Id", "")  # Keep for compatibility
                })

            return mapped_members
        except Exception as e:
            logger.error(f"Error mapping members for {property_name}: {str(e)}")
            raise e

    async def get_mapped_resources(self, property_name: str, start_date: str = None, end_date: str = None):
        """
        Fetch resources (rooms/spaces) directly from MEWS resources/getAll API
        using UpdatedUtc date range filter, same pattern as get_mapped_members.
        """
        try:
            if not start_date or not end_date:
                bkk_tz = ZoneInfo("Asia/Bangkok")
                now_bkk = datetime.now(bkk_tz)
                yesterday_bkk = now_bkk - timedelta(days=1)

                if not start_date:
                    start_dt = yesterday_bkk.replace(hour=0, minute=0, second=0, microsecond=0)
                    start_date = start_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
                if not end_date:
                    end_dt = yesterday_bkk.replace(hour=23, minute=59, second=59, microsecond=999999)
                    end_date = end_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

            all_resources = []
            current_cursor = None
            chunk_size = 500

            while True:
                res_payload = {
                    "UpdatedUtc": {
                        "StartUtc": start_date,
                        "EndUtc": end_date
                    },
                    "Extent": {"Resources": True},
                    "Limitation": {"Count": chunk_size}
                }
                if current_cursor:
                    res_payload["Limitation"]["Cursor"] = current_cursor

                res_res = await mews_client.post(
                    "/api/connector/v1/resources/getAll",
                    res_payload,
                    property_name=property_name
                )
                chunk = res_res.get("Resources", [])
                current_cursor = res_res.get("Cursor")
                all_resources.extend(chunk)

                if not current_cursor or not chunk:
                    break

            mapped_resources = []
            for res in all_resources:
                data_obj = res.get("Data") or {}
                mapped_resources.append({
                    "Name": res.get("Name", ""),
                    "State": res.get("State", ""),
                    "Active": res.get("IsActive", True),
                    "Parent Resource Id": res.get("ParentResourceId", ""),
                    "Floor Number": data_obj.get("FloorNumber", ""),
                    "Location Notes": data_obj.get("LocationNotes", ""),
                    "Created": res.get("CreatedUtc", ""),
                    "Updated": res.get("UpdatedUtc", ""),
                    "Identifier": res.get("Id", ""),
                    "mews_id": res.get("Id", "")
                })

            return mapped_resources
        except Exception as e:
            logger.error(f"Error mapping resources for {property_name}: {str(e)}")
            raise e

    async def get_mapped_payments(self, property_name: str, start_date: str = None, end_date: str = None):
        """
        Fetch payments (with Cursor pagination) and map all columns.
        """
        try:
            if not start_date or not end_date:
                bkk_tz = ZoneInfo("Asia/Bangkok")
                now_bkk = datetime.now(bkk_tz)
                yesterday_bkk = now_bkk - timedelta(days=1)

                if not start_date:
                    start_dt = yesterday_bkk.replace(hour=0, minute=0, second=0, microsecond=0)
                    start_date = start_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
                if not end_date:
                    end_dt = yesterday_bkk.replace(hour=23, minute=59, second=59, microsecond=999999)
                    end_date = end_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

            chunk_size = 1000
            all_payments = []
            current_cursor = None

            while True:
                payload = {
                    "CreatedUtc": {
                        "StartUtc": start_date,
                        "EndUtc": end_date
                    },
                    "Limitation": {"Count": chunk_size}
                }
                if current_cursor:
                    payload["Limitation"]["Cursor"] = current_cursor

                response = await mews_client.post("/api/connector/v1/payments/getAll", payload, property_name=property_name)
                chunk = response.get("Payments", [])
                current_cursor = response.get("Cursor")
                all_payments.extend(chunk)

                if not current_cursor or not chunk:
                    break

            mapped_payments = []
            for pay in all_payments:
                amount = pay.get("Amount") or {}
                orig = pay.get("OriginalAmount") or {}
                mapped_payments.append({
                    "mews_id": pay["Id"],
                    "Amount": amount.get("GrossValue", amount.get("NetValue")),
                    "Currency": amount.get("Currency"),
                    "Original Amount": orig.get("GrossValue", orig.get("NetValue")),
                    "Status": pay.get("State"),
                    "Type": pay.get("Type"),
                    "Kind": pay.get("Kind"),
                    "Number": pay.get("Number"),
                    "Processed At": pay.get("CreatedUtc"),
                    "Charged At": pay.get("ChargedUtc"),
                    "Identifier": pay.get("Identifier") or pay.get("Id"),
                    "Receipt Identifier": pay.get("ReceiptIdentifier"),
                    "Bill Id": pay.get("BillId"),
                    "Account Id": pay.get("AccountId"),
                    "Notes": pay.get("Notes"),
                })
            return mapped_payments
        except Exception as e:
            logger.error(f"Error mapping payments for {property_name}: {str(e)}")
            raise e

    @staticmethod
    def _split_date_windows(start_date: str, end_date: str, max_days: int = 89):
        """
        MEWS date-range filters (IssuedUtc, CreatedUtc, etc.) are capped at ~3 months per call.
        Split a wider range into consecutive windows so callers can loop over them.
        """
        start_dt = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
        end_dt = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
        windows = []
        cur = start_dt
        while cur < end_dt:
            window_end = min(cur + timedelta(days=max_days), end_dt)
            windows.append((
                cur.strftime("%Y-%m-%dT%H:%M:%SZ"),
                window_end.strftime("%Y-%m-%dT%H:%M:%SZ"),
            ))
            cur = window_end
        return windows or [(start_date, end_date)]

    @staticmethod
    def _extract_owner_name(owner_data: dict) -> str:
        wrapper = owner_data or {}
        owner = wrapper.get("Value") or {}
        if wrapper.get("Discriminator") == "BillCompanyData":
            return owner.get("Name") or ""
        return " ".join(filter(None, [
            owner.get("TitlePrefix"), owner.get("FirstName"),
            owner.get("SecondLastName"), owner.get("LastName")
        ]))

    @staticmethod
    def _extract_owner_address(owner_data: dict) -> dict:
        """
        Shared by get_mapped_bills_with_items (archived for the Data Mart) and
        get_bill_invoice (live print path) so a cached bill has everything
        needed to print without re-fetching from MEWS.
        """
        wrapper = owner_data or {}
        owner = wrapper.get("Value") or {}
        address = owner.get("Address") or {}
        legal = owner.get("LegalIdentifiers") or {}
        tax_id = legal.get("TaxIdentifier") or owner.get("TaxIdentifier") or ""
        address_lines = [l for l in [
            address.get("Line1"),
            address.get("Line2"),
            ", ".join(filter(None, [address.get("City"), address.get("SubdivisionCode")])),
            address.get("CountryCode"),
        ] if l]
        return {
            "address_lines": address_lines,
            "post_code": address.get("PostalCode") or "",
            "tax_id": tax_id,
        }

    async def get_mapped_bills_with_items(self, property_name: str, start_date: str = None, end_date: str = None):
        """
        Full Bill + Order Item archive for the Data Mart (unlike get_mapped_bills,
        which is header-only for the fast list view). For each bill in the date
        range, fetches its order items in bulk via BillIds (chunked to <=1000 ids
        per MEWS's limit) rather than one-by-one, to avoid N+1 calls across what
        can be thousands of bills for a wide date range.
        """
        try:
            if not start_date or not end_date:
                bkk_tz = ZoneInfo("Asia/Bangkok")
                now_bkk = datetime.now(bkk_tz)
                yesterday_bkk = now_bkk - timedelta(days=1)
                if not start_date:
                    start_dt = yesterday_bkk.replace(hour=0, minute=0, second=0, microsecond=0)
                    start_date = start_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
                if not end_date:
                    end_dt = yesterday_bkk.replace(hour=23, minute=59, second=59, microsecond=999999)
                    end_date = end_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

            all_bills = []
            for window_start, window_end in self._split_date_windows(start_date, end_date):
                current_cursor = None
                while True:
                    payload = {
                        "IssuedUtc": {"StartUtc": window_start, "EndUtc": window_end},
                        "Limitation": {"Count": 1000}
                    }
                    if current_cursor:
                        payload["Limitation"]["Cursor"] = current_cursor

                    response = await mews_client.post("/api/connector/v1/bills/getAll", payload, property_name=property_name)
                    chunk = response.get("Bills", [])
                    current_cursor = response.get("Cursor")
                    all_bills.extend(chunk)

                    if not current_cursor or not chunk:
                        break

            order_items_by_bill = {}
            bill_ids = [b.get("Id") for b in all_bills if b.get("Id")]
            for i in range(0, len(bill_ids), 1000):
                id_batch = bill_ids[i:i + 1000]
                current_cursor = None
                while True:
                    item_payload = {"BillIds": id_batch, "Limitation": {"Count": 1000}}
                    if current_cursor:
                        item_payload["Limitation"]["Cursor"] = current_cursor

                    item_res = await mews_client.post("/api/connector/v1/orderItems/getAll", item_payload, property_name=property_name)
                    item_chunk = item_res.get("OrderItems", [])
                    current_cursor = item_res.get("Cursor")

                    for item in item_chunk:
                        bill_id_ref = item.get("BillId")
                        if not bill_id_ref:
                            continue
                        amt = item.get("Amount") or {}
                        order_items_by_bill.setdefault(bill_id_ref, []).append({
                            "Name": item.get("BillingName") or item.get("Type") or "",
                            "Type": item.get("Type"),
                            "Amount": amt.get("GrossValue"),
                            "Net Amount": amt.get("NetValue"),
                            "Currency": amt.get("Currency"),
                            # Per-tax-rate breakdown ([{Code, Value}]) so invoices can
                            # split VAT 7% from Provincial Tax 1% instead of lumping
                            # them into one blended rate.
                            "Tax Values": amt.get("TaxValues") or [],
                        })

                    if not current_cursor or not item_chunk:
                        break

            mapped = []
            for b in all_bills:
                bill_id = b.get("Id")
                owner_address = self._extract_owner_address(b.get("OwnerData"))
                mapped.append({
                    "mews_id": bill_id,
                    "Number": b.get("Number"),
                    "Type": b.get("Type"),
                    "State": b.get("State"),
                    "Owner Name": self._extract_owner_name(b.get("OwnerData")),
                    "Address Lines": owner_address["address_lines"],
                    "Post Code": owner_address["post_code"],
                    "Tax Id": owner_address["tax_id"],
                    "Issued At": b.get("IssuedUtc"),
                    "Due At": b.get("DueUtc"),
                    "Paid At": b.get("PaidUtc"),
                    "Notes": b.get("Notes"),
                    "Order Items": order_items_by_bill.get(bill_id, []),
                })
            return mapped
        except Exception as e:
            logger.error(f"Error mapping bills with items for {property_name}: {str(e)}")
            raise e

    async def get_bill_invoice(self, property_name: str, bill_id: str):
        """
        Build the full itemized invoice payload for a single bill: header, guest/company
        address, line items (order items), computed VAT/subtotal/total, Thai baht-text,
        and a best-effort payment-method match — shaped to populate the
        "FullTAXInvoice LubdKohTao-TP" template placeholders.

        Checks bills_sync (Data Mart) first: a backfilled bill already has the
        header/owner/order-items, so this skips the two heaviest live MEWS calls
        (bills/getAll, orderItems/getAll) for anything already synced - only
        payment method is always fetched live, since the payments Data Mart table
        doesn't store a queryable Bill Id. Falls back to a full live fetch if the
        bill isn't cached, or the cached row predates the Address/Tax Id fields.
        """
        try:
            cached = None
            if self.supabase:
                try:
                    cache_res = self.supabase.table("bills_sync").select("data").eq("mews_id", bill_id).limit(1).execute()
                    if cache_res.data:
                        cached = encryption_service.decrypt_data(cache_res.data[0]["data"])
                except Exception as cache_err:
                    logger.warning(f"bills_sync cache lookup failed for {bill_id}: {cache_err}")

            rows = []
            if cached and "Address Lines" in cached and "Order Items" in cached:
                owner_name = cached.get("Owner Name") or ""
                address_lines = cached.get("Address Lines") or []
                post_code = cached.get("Post Code") or ""
                tax_id = cached.get("Tax Id") or ""
                number = cached.get("Number")
                bill_type = cached.get("Type")
                state = cached.get("State")
                issued_at = cached.get("Issued At")
                due_at = cached.get("Due At")
                for it in cached.get("Order Items") or []:
                    gross = it.get("Amount") or 0
                    net = it.get("Net Amount") or 0
                    rows.append({"name": it.get("Name") or "Item", "gross": gross, "net": net, "tax": gross - net,
                                 "tax_values": it.get("Tax Values") or []})
            else:
                bill_res = await mews_client.post(
                    "/api/connector/v1/bills/getAll",
                    {"BillIds": [bill_id], "Limitation": {"Count": 1}},
                    property_name=property_name
                )
                bills = bill_res.get("Bills", [])
                if not bills:
                    raise Exception("Bill not found")
                bill = bills[0]

                items = []
                current_cursor = None
                while True:
                    payload = {"BillIds": [bill_id], "Limitation": {"Count": 1000}}
                    if current_cursor:
                        payload["Limitation"]["Cursor"] = current_cursor
                    item_res = await mews_client.post("/api/connector/v1/orderItems/getAll", payload, property_name=property_name)
                    chunk = item_res.get("OrderItems", [])
                    current_cursor = item_res.get("Cursor")
                    items.extend(chunk)
                    if not current_cursor or not chunk:
                        break

                owner_wrapper = bill.get("OwnerData") or {}
                owner_name = self._extract_owner_name(owner_wrapper)
                owner_address = self._extract_owner_address(owner_wrapper)
                address_lines = owner_address["address_lines"]
                post_code = owner_address["post_code"]
                tax_id = owner_address["tax_id"]
                number = bill.get("Number")
                bill_type = bill.get("Type")
                state = bill.get("State")
                issued_at = bill.get("IssuedUtc")
                due_at = bill.get("DueUtc")

                for it in items:
                    amt = it.get("Amount") or {}
                    gross = amt.get("GrossValue") or 0
                    net = amt.get("NetValue") or 0
                    rows.append({
                        "name": it.get("BillingName") or it.get("Type") or "Item",
                        "gross": gross,
                        "net": net,
                        "tax": gross - net,
                        "tax_values": amt.get("TaxValues") or [],
                    })

            pay_res = await mews_client.post(
                "/api/connector/v1/payments/getAll",
                {"BillIds": [bill_id], "Limitation": {"Count": 100}},
                property_name=property_name
            )
            payments = pay_res.get("Payments", [])

            computed = self._compute_invoice_amounts(rows, payments)

            return {
                "mews_id": bill_id,
                "number": number,
                "type": bill_type,
                "state": state,
                "issued_at": issued_at,
                "due_at": due_at,
                "owner_name": owner_name,
                "address_lines": address_lines,
                "post_code": post_code,
                "tax_id": tax_id,
                **computed,
            }
        except Exception as e:
            logger.error(f"Error building bill invoice for {bill_id}: {str(e)}")
            raise e

    @staticmethod
    def _split_item_taxes(row: dict) -> tuple:
        """
        Returns (vat, other_tax) for one order-item row. Thai invoices must show
        VAT at its statutory 7% - MEWS items on room revenue also carry a 1%
        Provincial Tax, and lumping both into "VAT" yields a bogus blended rate
        (e.g. 7.64%). Prefers the item's MEWS TaxValues breakdown; older
        bills_sync rows predate "Tax Values", so those fall back to assuming the
        portion above 7% of net is the provincial tax.
        """
        net = row.get("net") or 0
        tax = row.get("tax") or 0
        tax_values = row.get("tax_values") or []
        if tax_values and net:
            # A tax entry well below the 7% band (e.g. 1% provincial) is "other";
            # vat is the remainder so the pair always reconciles with gross - net.
            other = sum((tv.get("Value") or 0) for tv in tax_values
                        if ((tv.get("Value") or 0) / net) < 0.055)
            return tax - other, other
        if net and (tax / net) > 0.075:
            vat = net * 0.07
            return vat, tax - vat
        return tax, 0.0

    @staticmethod
    def _compute_invoice_amounts(rows: list, payments: list) -> dict:
        """
        Shared by get_bill_invoice (single) and get_bill_invoices_batch (batch):
        turns raw order-item rows + payment records into the invoice's computed
        fields (5-row line items with >5 bundled into "Other charges", VAT 7% /
        Provincial Tax 1% breakdown, Thai baht-text, and a best-effort
        payment-method match).
        """
        rows = list(rows)
        for r in rows:
            r["vat_part"], r["other_part"] = SyncService._split_item_taxes(r)
        if len(rows) > 5:
            head, tail = rows[:4], rows[4:]
            rows = head + [{
                "name": f"Other charges ({len(tail)} items)",
                "gross": sum(r["gross"] for r in tail),
                "net": sum(r["net"] for r in tail),
                "tax": sum(r["tax"] for r in tail),
                "vat_part": sum(r["vat_part"] for r in tail),
                "other_part": sum(r["other_part"] for r in tail),
            }]

        sub_total = round(sum(r["net"] for r in rows), 2)
        total_tax = sum(r["tax"] for r in rows)
        provincial_tax = round(sum(r["other_part"] for r in rows), 2)
        # Total first, then VAT as the remainder, so rounding can never leave
        # SubTotal + VAT + Provincial != Total on the printed invoice.
        net_amount = round(sub_total + total_tax, 2)
        vat_total = round(net_amount - sub_total - provincial_tax, 2)
        # Thai statutory VAT - fixed label, not back-computed from the amounts.
        vat_rate_pct = 7

        line_items = []
        for i in range(5):
            if i < len(rows):
                # Mirror MEWS's per-item tax sub-line: items carrying provincial
                # tax say so in the description (amount column stays gross).
                desc = rows[i]["name"]
                if rows[i]["other_part"] > 0.005:
                    desc = f"{desc} (Provincial Tax 1%: {rows[i]['other_part']:,.2f})"
                line_items.append({"no": i + 1, "description": desc, "amount": round(rows[i]["gross"], 2)})
            else:
                line_items.append({"no": "", "description": "", "amount": ""})

        method = {"cash": False, "card": False, "bank_transfer": False, "cheque": False}
        bank_transfer_ref = ""
        bank_transfer_date = ""
        cheque = {"bank_name": "", "branch": "", "number": "", "date": ""}
        for p in payments:
            ptype = (p.get("Type") or "").lower()
            pkind = (p.get("Kind") or "").lower()
            if "cash" in ptype or "cash" in pkind:
                method["cash"] = True
            elif "card" in ptype or "card" in pkind:
                method["card"] = True
            elif "transfer" in ptype or "transfer" in pkind:
                method["bank_transfer"] = True
                bank_transfer_ref = p.get("Identifier") or ""
                bank_transfer_date = p.get("ChargedUtc") or p.get("CreatedUtc") or ""
            elif "cheque" in ptype or "check" in ptype or "cheque" in pkind:
                method["cheque"] = True
                cheque["number"] = p.get("Number") or ""
                cheque["date"] = p.get("ChargedUtc") or ""

        return {
            "line_items": line_items,
            "sub_total": sub_total,
            "vat_rate_pct": vat_rate_pct,
            "vat": vat_total,
            "provincial_tax_rate_pct": 1,
            "provincial_tax": provincial_tax,
            "net_amount": net_amount,
            "baht_text": bahttext(net_amount),
            "payment_method": method,
            "bank_transfer_ref": bank_transfer_ref,
            "bank_transfer_date": bank_transfer_date,
            "cheque": cheque,
        }

    @staticmethod
    def compute_bill_totals(order_items: list) -> dict:
        """
        Net/VAT/Total for the Bill Generator list view (a lightweight summary,
        not the full itemized invoice) - reuses _compute_invoice_amounts' same
        VAT-7%/Provincial-Tax-1% split so these numbers always agree with what
        print-bill would show for the same bill. "Net Amount" is the pre-tax
        subtotal and "Total Amount" is the grand total, matching the labels
        _compute_invoice_amounts itself uses for sub_total/net_amount (a naming
        holdover from the Thai invoice template, where "Net Amount" labels the
        subtotal row and the token <<NetAmount>> is actually the grand total).
        """
        rows = []
        for it in order_items or []:
            gross = it.get("Amount") or 0
            net = it.get("Net Amount") or 0
            rows.append({
                "name": it.get("Name") or "Item", "gross": gross, "net": net, "tax": gross - net,
                "tax_values": it.get("Tax Values") or [],
            })
        computed = SyncService._compute_invoice_amounts(rows, [])
        return {
            "Net Amount": computed["sub_total"],
            "VAT": computed["vat"],
            "Total Amount": computed["net_amount"],
        }

    async def get_bill_invoices_batch(self, property_name: str, bill_ids: list) -> dict:
        """
        Same invoice-building logic as get_bill_invoice, but for many bills at
        once: one bills_sync cache lookup (IN query) instead of N, one live
        bills/getAll + orderItems/getAll pair for whatever isn't cached (BillIds
        batch, matching get_mapped_bills_with_items's pattern) instead of N live
        pairs, and one payments/getAll call for the whole batch instead of N -
        this is what makes multi-print meaningfully faster than printing each
        bill individually through get_bill_invoice in a loop.

        Returns {bill_id: invoice_dict} for every bill that was found; ids that
        don't exist or fail to resolve are simply omitted from the result.
        """
        try:
            seen = set()
            bill_ids = [b for b in bill_ids if b and not (b in seen or seen.add(b))]
            if not bill_ids:
                return {}

            headers = {}

            if self.supabase:
                try:
                    cache_res = self.supabase.table("bills_sync").select("mews_id, data").in_("mews_id", bill_ids).execute()
                    for r in cache_res.data or []:
                        decrypted = encryption_service.decrypt_data(r["data"])
                        if "Address Lines" not in decrypted or "Order Items" not in decrypted:
                            continue
                        rows = []
                        for it in decrypted.get("Order Items") or []:
                            gross = it.get("Amount") or 0
                            net = it.get("Net Amount") or 0
                            rows.append({"name": it.get("Name") or "Item", "gross": gross, "net": net, "tax": gross - net,
                                         "tax_values": it.get("Tax Values") or []})
                        headers[r["mews_id"]] = {
                            "owner_name": decrypted.get("Owner Name") or "",
                            "address_lines": decrypted.get("Address Lines") or [],
                            "post_code": decrypted.get("Post Code") or "",
                            "tax_id": decrypted.get("Tax Id") or "",
                            "number": decrypted.get("Number"),
                            "type": decrypted.get("Type"),
                            "state": decrypted.get("State"),
                            "issued_at": decrypted.get("Issued At"),
                            "due_at": decrypted.get("Due At"),
                            "rows": rows,
                        }
                except Exception as cache_err:
                    logger.warning(f"bills_sync batch cache lookup failed: {cache_err}")

            missing_ids = [b for b in bill_ids if b not in headers]
            if missing_ids:
                bills_by_id = {}
                for i in range(0, len(missing_ids), 1000):
                    id_batch = missing_ids[i:i + 1000]
                    current_cursor = None
                    while True:
                        payload = {"BillIds": id_batch, "Limitation": {"Count": 1000}}
                        if current_cursor:
                            payload["Limitation"]["Cursor"] = current_cursor
                        res = await mews_client.post("/api/connector/v1/bills/getAll", payload, property_name=property_name)
                        chunk = res.get("Bills", [])
                        current_cursor = res.get("Cursor")
                        for b in chunk:
                            bills_by_id[b.get("Id")] = b
                        if not current_cursor or not chunk:
                            break

                items_by_bill = {}
                for i in range(0, len(missing_ids), 1000):
                    id_batch = missing_ids[i:i + 1000]
                    current_cursor = None
                    while True:
                        payload = {"BillIds": id_batch, "Limitation": {"Count": 1000}}
                        if current_cursor:
                            payload["Limitation"]["Cursor"] = current_cursor
                        res = await mews_client.post("/api/connector/v1/orderItems/getAll", payload, property_name=property_name)
                        chunk = res.get("OrderItems", [])
                        current_cursor = res.get("Cursor")
                        for item in chunk:
                            bref = item.get("BillId")
                            if not bref:
                                continue
                            amt = item.get("Amount") or {}
                            gross = amt.get("GrossValue") or 0
                            net = amt.get("NetValue") or 0
                            items_by_bill.setdefault(bref, []).append({
                                "name": item.get("BillingName") or item.get("Type") or "Item",
                                "gross": gross,
                                "net": net,
                                "tax": gross - net,
                                "tax_values": amt.get("TaxValues") or [],
                            })
                        if not current_cursor or not chunk:
                            break

                for bid, bill in bills_by_id.items():
                    owner_wrapper = bill.get("OwnerData") or {}
                    owner_address = self._extract_owner_address(owner_wrapper)
                    headers[bid] = {
                        "owner_name": self._extract_owner_name(owner_wrapper),
                        "address_lines": owner_address["address_lines"],
                        "post_code": owner_address["post_code"],
                        "tax_id": owner_address["tax_id"],
                        "number": bill.get("Number"),
                        "type": bill.get("Type"),
                        "state": bill.get("State"),
                        "issued_at": bill.get("IssuedUtc"),
                        "due_at": bill.get("DueUtc"),
                        "rows": items_by_bill.get(bid, []),
                    }

            found_ids = [b for b in bill_ids if b in headers]
            if not found_ids:
                return {}

            payments_by_bill = {}
            for i in range(0, len(found_ids), 1000):
                id_batch = found_ids[i:i + 1000]
                current_cursor = None
                while True:
                    payload = {"BillIds": id_batch, "Limitation": {"Count": 1000}}
                    if current_cursor:
                        payload["Limitation"]["Cursor"] = current_cursor
                    res = await mews_client.post("/api/connector/v1/payments/getAll", payload, property_name=property_name)
                    chunk = res.get("Payments", [])
                    current_cursor = res.get("Cursor")
                    for p in chunk:
                        bref = p.get("BillId")
                        if bref:
                            payments_by_bill.setdefault(bref, []).append(p)
                    if not current_cursor or not chunk:
                        break

            results = {}
            for bid in found_ids:
                h = headers[bid]
                computed = self._compute_invoice_amounts(h["rows"], payments_by_bill.get(bid, []))
                results[bid] = {
                    "mews_id": bid,
                    "number": h["number"],
                    "type": h["type"],
                    "state": h["state"],
                    "issued_at": h["issued_at"],
                    "due_at": h["due_at"],
                    "owner_name": h["owner_name"],
                    "address_lines": h["address_lines"],
                    "post_code": h["post_code"],
                    "tax_id": h["tax_id"],
                    **computed,
                }
            return results
        except Exception as e:
            logger.error(f"Error building batch bill invoices for {property_name}: {str(e)}")
            raise e

    async def get_bill_pdf(self, property_name: str, bill_id: str, pdf_template: str = None,
                            print_reason: str = None, bill_print_event_id: str = None):
        """
        Calls MEWS bills/getPdf to get MEWS's own generated PDF for a bill (as an
        alternative to our custom Thai tax invoice template). Per MEWS docs, the
        response is either the ready PDF (Discriminator "BillPdfFile") or a pending
        event (Discriminator "BillPrintEvent") that should be retried with the same
        BillPrintEventId to avoid consuming additional print-event quota. We retry
        briefly server-side; if still pending after that, we hand the event id back
        to the caller so a later request can resume it.
        """
        try:
            payload: dict = {"BillId": bill_id}
            if pdf_template:
                payload["PdfTemplate"] = pdf_template
            if print_reason:
                payload["PrintReason"] = print_reason

            max_attempts = 4
            wait_seconds = 2
            event_id = bill_print_event_id

            for attempt in range(max_attempts):
                if event_id:
                    payload["BillPrintEventId"] = event_id

                response = await mews_client.post("/api/connector/v1/bills/getPdf", payload, property_name=property_name)
                result = response.get("Result") or {}
                discriminator = result.get("Discriminator")
                value = result.get("Value") or {}

                if discriminator == "BillPdfFile":
                    return {"ready": True, "base64": value.get("Base64Data")}

                event_id = value.get("BillPrintEventId") or event_id
                if attempt < max_attempts - 1:
                    await asyncio.sleep(wait_seconds)

            return {"ready": False, "event_id": event_id}
        except Exception as e:
            logger.error(f"Error fetching bill PDF for {bill_id}: {str(e)}")
            raise e

    async def get_rr3_cards(self, property_name: str, start_date: str = None, end_date: str = None):
        """
        Builds Thai Hotel Act RR3 (ร.ร.๓) lodger registration cards by joining
        Reservations + Customers + Resources for a date range - a direct port of
        the user's existing, proven Google Apps Script: same request shape (the
        older, un-versioned reservations/getAll with top-level StartUtc/EndUtc +
        Extent, CustomerId/CompanionIds fields - deliberately NOT the newer
        /getAll/2023-06-06 + AccountId shape get_mapped_reservations uses, since
        that variant doesn't support Extent-embedded Customers/Resources), same
        client-side re-filter by StartUtc, and same field fallback logic.
        """
        try:
            if not start_date or not end_date:
                now_utc = datetime.now(timezone.utc)
                yesterday_utc = now_utc - timedelta(days=1)
                if not start_date:
                    start_date = yesterday_utc.strftime("%Y-%m-%dT00:00:00Z")
                if not end_date:
                    end_date = now_utc.strftime("%Y-%m-%dT23:59:59Z")

            payload = {
                "StartUtc": start_date,
                "EndUtc": end_date,
                "Extent": {"Reservations": True, "Customers": True, "Resources": True}
            }
            response = await mews_client.post("/api/connector/v1/reservations/getAll", payload, property_name=property_name)

            wanted_start = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
            wanted_end = datetime.fromisoformat(end_date.replace("Z", "+00:00"))

            def in_range(res):
                start_utc = res.get("StartUtc")
                if not start_utc:
                    return False
                try:
                    t = datetime.fromisoformat(start_utc.replace("Z", "+00:00"))
                except Exception:
                    return False
                return wanted_start <= t <= wanted_end

            reservations = [r for r in response.get("Reservations", []) if in_range(r)]
            customers_map = {c["Id"]: c for c in response.get("Customers", []) if c.get("Id")}
            resources_map = {r["Id"]: r for r in response.get("Resources", []) if r.get("Id")}

            # StartUtc/EndUtc above are the legacy (2017-04-12) endpoint's
            # DEPRECATED fields - they hold the reservation's originally
            # SCHEDULED arrival/departure (e.g. the hotel's standard 14:00
            # check-in), not the real front-desk timestamp, and never get
            # updated once the guest actually checks in. The real timestamps
            # (ActualStartUtc/ActualEndUtc) only exist on the newer
            # 2023-06-06 endpoint, which in turn can't embed Customers/
            # Resources the way the legacy one does (confirmed against MEWS's
            # own docs - neither endpoint version has both) - hence this
            # second, narrow call by exact ReservationIds (max 1000/request)
            # to backfill just the two fields the legacy call can't provide.
            actual_times = {}
            reservation_ids = [r["Id"] for r in reservations if r.get("Id")]
            for i in range(0, len(reservation_ids), 1000):
                id_batch = reservation_ids[i:i + 1000]
                try:
                    actual_res = await mews_client.post(
                        "/api/connector/v1/reservations/getAll/2023-06-06",
                        {"ReservationIds": id_batch, "Limitation": {"Count": len(id_batch)}},
                        property_name=property_name,
                    )
                    for r in actual_res.get("Reservations", []):
                        if r.get("Id"):
                            actual_times[r["Id"]] = (r.get("ActualStartUtc"), r.get("ActualEndUtc"))
                except Exception as e:
                    # Falls back to the legacy (scheduled) times below rather
                    # than failing the whole card list over this.
                    logger.warning(f"RR3: failed to fetch ActualStartUtc/ActualEndUtc for {property_name}: {e}")

            hotel_name = _RR3_PROPERTY_THAI_NAMES.get(property_name, property_name)

            cards = []
            for reservation in reservations:
                actual_start, actual_end = actual_times.get(reservation.get("Id"), (None, None))
                # Not yet checked in/out (shouldn't normally happen - this
                # page is defined as "generated from MEWS check-ins" - but
                # falls back safely to the scheduled time rather than a blank).
                check_in_utc = actual_start or reservation.get("StartUtc")
                check_out_utc = actual_end or reservation.get("EndUtc")

                companion_ids = reservation.get("CompanionIds") or []
                guest_ids = companion_ids if companion_ids else (
                    [reservation["CustomerId"]] if reservation.get("CustomerId") else []
                )

                for guest_id in guest_ids:
                    customer = customers_map.get(guest_id)
                    if not customer:
                        continue

                    room_number = ""
                    assigned_resource_id = reservation.get("AssignedResourceId")
                    if assigned_resource_id and assigned_resource_id in resources_map:
                        room_number = resources_map[assigned_resource_id].get("Name", "")

                    address = customer.get("Address") or {}
                    line1 = (address.get("Line1") or "").strip()
                    line2 = (address.get("Line2") or "").strip()
                    if line1 or line2:
                        parts = [address.get("Line1"), address.get("Line2"), address.get("City"),
                                 address.get("PostalCode"), address.get("CountryCode")]
                        address_details = " ".join(p for p in parts if p)
                    elif (customer.get("BirthPlace") or "").strip():
                        address_details = customer.get("BirthPlace")
                    else:
                        address_details = _rr3_country_name(customer.get("NationalityCode"))

                    identity_card_value = ""
                    identity_card = customer.get("IdentityCard")
                    identity_cards = customer.get("IdentityCards")
                    if isinstance(identity_card, dict):
                        identity_card_value = identity_card.get("Number", "")
                    elif isinstance(identity_cards, list) and identity_cards:
                        identity_card_value = identity_cards[0].get("Number", "")

                    passport = customer.get("Passport") or {}
                    occupation = customer.get("Occupation") or "นักธุรกิจ"
                    first_name = customer.get("FirstName", "")
                    last_name = customer.get("LastName", "")

                    cards.append({
                        "CardId": f"{reservation.get('Number', '')}::{guest_id}",
                        "ReservationsNumber": reservation.get("Number", ""),
                        "HotelName": hotel_name,
                        "FirstName": first_name,
                        "LastName": last_name,
                        "RoomNumber": room_number,
                        "CheckIn": _rr3_format_thai_date(check_in_utc),
                        "CheckInTime": _rr3_format_thai_time(check_in_utc),
                        "CheckOut": _rr3_format_thai_date(check_out_utc),
                        "CheckOutTime": _rr3_format_thai_time(check_out_utc),
                        "PassportNumber": passport.get("Number", ""),
                        "IdentityCardNumber": identity_card_value,
                        "NationalityCode": customer.get("NationalityCode", ""),
                        "NationalityName": _rr3_country_name(customer.get("NationalityCode")),
                        "AddressDetails": address_details,
                        "Telephone": customer.get("Phone", ""),
                        "Email": customer.get("Email", ""),
                        "Occupation": occupation,
                        "AlienBook": customer.get("IdentityDocumentSupportNumber", ""),
                        "GuestSign": f"{first_name} {last_name}".strip(),
                        "Departure": "",
                        "Destination": "",
                    })

            return cards
        except Exception as e:
            logger.error(f"Error building RR3 cards for {property_name}: {str(e)}")
            raise e

    # Fallback when a property has no st_space_types configured. Each
    # property's MEWS export schedule carries its OWN "Space types" filter, so
    # there is no single correct list: Chinatown/Siam/Samui/Makati/Patong/Siem
    # Reap are Room+Bed (verified - Chinatown sums to exactly 176, and Siem
    # Reap's 6-space Suite category is deliberately NOT in its 222), while Koh
    # Tao and Marasca Samui also count Suite (51+3=54 and 57+5=62, matching
    # their exports exactly). Set st_space_types per property to override.
    _ST_FILES_SPACE_TYPES = ("Room", "Bed")

    async def _resolve_st_space_types(self, property_name: str) -> tuple:
        """Per-property "Space types" filter, comma-separated in
        property_api_settings.st_space_types (Admin > API Settings). Falls
        back to Room+Bed when unset - or when the column doesn't exist yet, so
        the report keeps working before that migration is run."""
        if not self.supabase:
            return self._ST_FILES_SPACE_TYPES
        try:
            res = self.supabase.table("property_api_settings").select(
                "st_space_types").eq("property_name", property_name).limit(1).execute()
        except Exception as e:
            logger.warning(f"ST Files space types lookup failed for {property_name}: {e}")
            return self._ST_FILES_SPACE_TYPES
        raw = (res.data[0].get("st_space_types") if res.data else None) or ""
        return tuple(t.strip() for t in raw.split(",") if t.strip()) or self._ST_FILES_SPACE_TYPES

    @staticmethod
    def _resolve_stay_service(services: list) -> Optional[dict]:
        """Picks the property's real accommodation ("Stay") service out of
        services/getAll's response. Filtering on Discriminator=="Bookable"
        alone isn't enough - a property can have OTHER, inactive Bookable
        services too, and MEWS doesn't order the array with the real one
        first. Confirmed live on Lub d Bangkok Siam: two inactive
        "Co-Working Space" Bookable services sorted before the actual
        active "Stay" one, so the old `next(... Discriminator=="Bookable")`
        picked an inactive service and every call downstream
        (getAvailability, resourceCategories) failed with MEWS's "Invalid
        ServiceId." Filters to active ones, then prefers one literally
        named "Stay" (MEWS's own near-universal convention for the
        accommodation service) in case a property has more than one
        active Bookable service - falls back to the first active one so a
        property naming it something else doesn't silently return nothing.
        """
        active_bookable = [
            s for s in services
            if (s.get("Data") or {}).get("Discriminator") == "Bookable" and s.get("IsActive", True)
        ]
        if not active_bookable:
            return None
        return next((s for s in active_bookable if s.get("Name") == "Stay"), active_bookable[0])

    async def _resolve_property_timezone(self, property_name: str) -> ZoneInfo:
        """Nearly every property here is Asia/Bangkok, but not all - MEWS's
        own Enterprise.TimeZoneIdentifier for "Lub d Philippines Makati" is
        actually "Asia/Singapore" (UTC+8, an hour ahead of Bangkok).
        get_st_files_report used to hardcode Asia/Bangkok for every
        property's day-boundary math, which MEWS's own getAvailability
        endpoint then rejected for Makati specifically with "Invalid
        FirstTimeUnitStartUtc: not start of TimeUnit" - MEWS computes that
        endpoint's TimeUnit boundaries from the property's own configured
        timezone, not Bangkok's, so the request has to match. Falls back to
        Bangkok (the correct value for every other property today) if
        configuration/get fails or the field's ever missing/malformed,
        rather than raising and breaking the whole report over it.
        """
        try:
            res = await mews_client.post(
                "/api/connector/v1/configuration/get",
                {},
                property_name=property_name,
            )
            tz_name = (res.get("Enterprise") or {}).get("TimeZoneIdentifier")
            if tz_name:
                return ZoneInfo(tz_name)
        except Exception as e:
            logger.warning(f"Could not resolve MEWS timezone for {property_name}, defaulting to Asia/Bangkok: {e}")
        return ZoneInfo("Asia/Bangkok")

    # ST Files List's "Complimentary" column. The Master tab of all 8
    # "<Name>-ST" sheets (the ground truth for what gets filed) computes it as:
    #
    #   Status = "Checked in"  AND  ( Rate = "Complimentary Room"
    #                                 OR Segment = "Complimentary" )
    #
    # BOTH arms are load-bearing, and the Segment arm is the reliable one -
    # checked against the four real complimentary rooms on 25-Aug-2026:
    #
    #   Chinatown #87434  Rate "Complimentary Room"                 Segment Complimentary
    #   Marasca   #44501  Rate "Complimentary Room"                 Segment Complimentary
    #   Koh Tao   #40971  Rate "Complimentary Room with Breakfast"  Segment Complimentary
    #   Patong    #156361 Rate "Group Leisure Series - Room Only"   Segment Complimentary
    #
    # Segment is right on 4 of 4, Rate on only 2. Patong is the case that
    # settles it: its rate is an ordinary commercial rate with nothing
    # complimentary about the name, and only the Segment marks the room -
    # so no amount of rate-name matching (prefix, fuzzy, or otherwise) could
    # ever have found it. A Rate-only rule under-counted Koh Tao and Patong
    # by one each.
    #
    # State == "Started" is MEWS's own "Checked in", so the Status arm needs
    # no translation. Voucher is deliberately NOT checked, unlike an older
    # version of the sheet formula: live testing found VoucherId essentially
    # never populated (0 of 381 sampled) and several properties' Connector
    # tokens don't have vouchers/getAll enabled at all (401).
    #
    # Lub d Siem Reap's own sheet uses a narrower formula of its own
    # (Rate only, plus "not cancelled", no Segment and no Status). It is
    # deliberately NOT reproduced: it would miss a Segment-only complimentary
    # room exactly the way Patong's would have been missed. Both rules give
    # Siem Reap 0 today, so nothing changes for it now.
    _COMPLIMENTARY_RATE_NAMES = {"Complimentary", "Complimentary Room"}
    _COMPLIMENTARY_SEGMENT_NAMES = {"Complimentary"}

    def _is_complimentary(self, res: dict, rates_by_id: dict, segments_by_id: dict) -> bool:
        """The rule above, in one place - the ST file's headline Complimentary
        count and the per-reservation audit flag both call this, so the two
        can't drift into disagreeing about the same guest."""
        if res.get("State") != "Started":
            return False
        rate_name = (rates_by_id.get(res.get("RateId"), {}).get("Name") or "").strip()
        segment_name = (segments_by_id.get(res.get("BusinessSegmentId")) or "").strip()
        return (rate_name in self._COMPLIMENTARY_RATE_NAMES
                or segment_name in self._COMPLIMENTARY_SEGMENT_NAMES)

    # ST Files List's "Download" file - the legacy pipe-delimited PMSST/RMSST
    # statistics export. Field layout confirmed against the source Google
    # Sheet formula itself (see get_st_report_export's docstring) - field 5
    # is a report-month code, not a property code; the real per-property
    # value (field 17) comes from property_api_settings.st_property_code.

    # (metric_code, field-8 label, totals-dict key, record_type). "key" is
    # None for "No. of Day", whose value is always the literal 1.
    _ST_REPORT_METRICS = [
        ("90107", "Spaces", "spaces", "RMSST"),
        ("90100", "Occupied", "occupied", "PMSST"),
        ("90104", "House uses", "house_use", "PMSST"),
        ("90103", "Out of order", "out_of_order", "PMSST"),
        ("90102", "Availability", "availability", "PMSST"),
        ("90105", "Customers", "customers", "PMSST"),
        ("90114", "Arrivals", "arrivals", "PMSST"),
        ("90115", "Departures", "departures", "PMSST"),
        ("90101", "Complimentary Room", "complimentary", "PMSST"),
        ("90108", "No. of Day", None, "PMSST"),
    ]

    # Field 9 (currency) isn't THB everywhere - confirmed against real
    # ST files for every property on 09-Aug-2026: Makati (Philippines) files
    # in PHP and Siem Reap (Cambodia) in USD, every other property in THB.
    _ST_REPORT_CURRENCY = {
        "Lub d Philippines Makati": "PHP",
        "Lub d Siem Reap": "USD",
    }

    # Field 7's date suffix isn't uniformly yyyymmdd - confirmed against real
    # ST files for all 8 properties on 09-Aug-2026: the "No. of Day" row
    # always uses ddmmyyyy ("ST09082026") regardless of property, while the
    # other 9 metric rows use ddmmyyyy for these four properties and
    # yyyymmdd ("ST20260809") for every other property. Single-day sample -
    # revisit if a second day's files contradict the grouping.
    _ST_REPORT_DDMMYYYY_REF_PROPERTIES = {
        "Lub d Bangkok Chinatown",
        "Lub d Koh Samui Chaweng Beach",
        "Lub d Koh Tao Tanote Bay",
        "Lub d Philippines Makati",
    }

    # The statuses the generator sheets ask MEWS for. Chinatown's
    # Parameter-ImportInhouse tab - the only one that records the report
    # config - read "Confirmed, Checked in, Checked out" on 20-Aug-2026 and
    # "Confirmed, Checked in, Checked out, Optional" on 21-Aug, so whoever
    # generates the file added Optional in between. Those are MEWS's
    # report-side names for Confirmed / Started / Processed / Optional.
    #
    # Optional costs nothing to carry: across all six Thai properties for
    # both 19- and 20-Aug-2026 every reservation MEWS returned was either
    # Started or Processed, so it changes no current figure. Keeping it
    # matches the documented config, and an extra status can only ever add a
    # guest to a lodger register, never drop one.
    _RR4_TM30_ACTIVE_STATES = {"Confirmed", "Started", "Processed", "Optional"}

    async def _rr4_tm30_fetch_day(self, property_name: str, date: str):
        """
        Shared fetch for RR4/TM30: resolves the property's own calendar-day
        window and pulls every active-state reservation colliding with it -
        the same un-versioned Extent-join reservations/getAll call
        get_rr3_cards and get_st_files_report both use. Returns (day,
        day_start_utc, day_end_utc, reservations, customers_map,
        resources_map); `day` is the tz-aware property-local midnight for
        the report date, reused by callers to format the header date.

        day_start_utc/day_end_utc are set from property_api_settings.
        rr4_tm30_day_start_hour/_minute and rr4_tm30_day_end_hour/_minute
        (all default 0, matching MEWS's own BusinessDayClosingOffset -
        confirmed 0 for every property checked via configuration/get) - a
        per-property override for hotels whose own front-desk/finance team
        files arrivals/departures under a different manual window than
        MEWS's official midnight-to-midnight one. Minute-level precision
        exists specifically so this can mirror MEWS's own native
        "Customer profiles" report window exactly (e.g. 12:15 to 12:15 the
        next day - confirmed against a real MEWS export for Lub d Bangkok
        Chinatown), not just whole hours. End wraps to the next day
        whenever its time-of-day isn't strictly after start's (matching
        how people describe an overnight window, e.g. "22:00 to 06:00") -
        so the default 0:00/0:00 still resolves to the full current-day
        window, and an explicit 14:00/12:00 resolves to today 14:00
        through tomorrow 12:00 (a 22-hour window, deliberately not
        required to be 24h). This is independently editable per request -
        a prior version forced it to always be exactly 24h after a stale
        14/12 value left on Chinatown silently undercounted its RR4
        register by 76 guests (160 rows instead of the correct 241,
        confirmed against a real MEWS export); the safety constraint was
        explicitly reverted, so a non-default end time again carries that
        same risk if left stale - double-check any non-zero value here
        after setting it. `day` itself stays plain local midnight - only
        the window used for the actual reservation query and in_window()
        checks shifts, so header/display dates aren't affected.
        """
        property_tz = await self._resolve_property_timezone(property_name)
        day_start_hour = day_start_minute = day_end_hour = day_end_minute = 0
        if self.supabase:
            try:
                prop_res = self.supabase.table("property_api_settings").select(
                    "rr4_tm30_day_start_hour, rr4_tm30_day_start_minute, "
                    "rr4_tm30_day_end_hour, rr4_tm30_day_end_minute").eq(
                    "property_name", property_name).limit(1).execute()
                row = prop_res.data[0] if prop_res.data else {}
                day_start_hour = row.get("rr4_tm30_day_start_hour") or 0
                day_start_minute = row.get("rr4_tm30_day_start_minute") or 0
                day_end_hour = row.get("rr4_tm30_day_end_hour") or 0
                day_end_minute = row.get("rr4_tm30_day_end_minute") or 0
            except Exception:
                pass
        day = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=property_tz)
        start_of_day = day_start_hour * 60 + day_start_minute
        end_of_day = day_end_hour * 60 + day_end_minute
        end_days = 1 if end_of_day <= start_of_day else 0
        day_start_utc = (day + timedelta(hours=day_start_hour, minutes=day_start_minute)).astimezone(timezone.utc)
        day_end_utc = (day + timedelta(days=end_days, hours=day_end_hour, minutes=day_end_minute)).astimezone(timezone.utc)
        start_iso = day_start_utc.strftime("%Y-%m-%dT%H:%M:%SZ")
        end_iso = day_end_utc.strftime("%Y-%m-%dT%H:%M:%SZ")

        resv_res = await mews_client.post(
            "/api/connector/v1/reservations/getAll",
            {"StartUtc": start_iso, "EndUtc": end_iso,
             "Extent": {"Reservations": True, "Customers": True, "Resources": True}},
            property_name=property_name,
        )
        reservations = [r for r in resv_res.get("Reservations", [])
                        if r.get("State") in self._RR4_TM30_ACTIVE_STATES]
        customers_map = {c["Id"]: c for c in resv_res.get("Customers", []) if c.get("Id")}
        resources_map = {r["Id"]: r for r in resv_res.get("Resources", []) if r.get("Id")}
        return day, day_start_utc, day_end_utc, reservations, customers_map, resources_map

    @staticmethod
    def _rr4_tm30_guest_ids(res: dict) -> list:
        """Who actually SLEEPS in this space, deduped and order-preserved.

        CompanionIds, whenever it's populated, is MEWS's definitive occupant
        list for that one space - its length matches the reservation's own
        AdultCount+ChildCount exactly - so it is used alone. CustomerId is
        only appended when CompanionIds is empty, because CustomerId means
        "who owns/pays for this booking", which is NOT the same person as
        "who sleeps in this room" once a booking covers several rooms.

        An earlier version unioned CustomerId with CompanionIds on the
        assumption that CompanionIds always already contained the owner. It
        does - but only for the single room the owner personally occupies;
        for a family/group that books several rooms at once, MEWS repeats
        that same owner as CustomerId on EVERY room's reservation while
        listing only the real occupants in each one's CompanionIds. Unioning
        therefore filed the booker into every room they paid for: one guest
        at Lub d Koh Tao Tanote Bay (16-Aug-2026) appeared in 3 different
        rooms at once, and the register ran 12 rows over MEWS's own figure
        on that day alone. Confirmed against real MEWS "Customer profiles In
        house" exports for all 5 properties with reference data - this rule
        removes every over-count without dropping a single guest MEWS lists,
        and brings each reservation's row count back in line with the
        headcount it was actually booked for.
        """
        ids = [c for c in (res.get("CompanionIds") or []) if c] or \
              ([res["CustomerId"]] if res.get("CustomerId") else [])
        seen, out = set(), []
        for cid in ids:
            if cid not in seen:
                seen.add(cid)
                out.append(cid)
        return out

    async def _resolve_rr4_property_thai_name(self, property_name: str) -> str:
        """Admin > API's per-property "Property Thai Name" field
        (property_api_settings.rr4_property_thai_name) - the property's
        real registered name for RR4/TM30 filings, editable without a code
        change. Falls back to the hardcoded _RR4_PROPERTY_THAI_NAMES /
        _RR3_PROPERTY_THAI_NAMES chain (then the raw property_name) for any
        property that hasn't set it yet, so nothing regresses for
        Chinatown/Siam - the two already confirmed against real reference
        filings - until someone fills in the field for the rest."""
        if self.supabase:
            try:
                res = self.supabase.table("property_api_settings").select(
                    "rr4_property_thai_name").eq("property_name", property_name).limit(1).execute()
                name = res.data[0].get("rr4_property_thai_name") if res.data else None
                if name:
                    return name
            except Exception:
                pass
        return _RR4_PROPERTY_THAI_NAMES.get(
            property_name, _RR3_PROPERTY_THAI_NAMES.get(property_name, property_name))

    async def _resolve_rr4_nationality_codes(self) -> dict:
        """Admin > RR4-Nationality's editable Thai Hotel Act numeric code
        table (rr4_nationality_codes), letting staff correct or add
        nationality codes without a deploy. Keyed by country NAME
        (mews_nationality) rather than alpha-2 - matching the original
        manual spreadsheet's own VLOOKUP-by-name convention, and letting
        non-technical staff add a nationality without knowing its ISO code
        - so it's joined here via _RR3_COUNTRY_MAP (alpha-2 -> the exact
        same English name this table's rows were seeded from) rather than
        a raw alpha-2 match. Returns an alpha-2-keyed dict (same shape
        RR4_NATIONALITY_CODE uses) so the caller doesn't need to know about
        this indirection. Starts from the hardcoded RR4_NATIONALITY_CODE
        dict and layers the DB rows on top per-entry, so a table that's
        empty/unreachable never blanks a nationality that already worked -
        same graceful-degradation pattern role_permissions uses when its
        own table is empty."""
        codes = dict(RR4_NATIONALITY_CODE)
        if self.supabase:
            try:
                res = self.supabase.table("rr4_nationality_codes").select(
                    "mews_nationality, rr4_code").execute()
                by_name = {}
                for row in res.data or []:
                    name = (row.get("mews_nationality") or "").strip().lower()
                    code = (row.get("rr4_code") or "").strip()
                    if name and code:
                        by_name[name] = code
                for alpha2, country_name in _RR3_COUNTRY_MAP.items():
                    code = by_name.get(country_name.strip().lower())
                    if code:
                        codes[alpha2] = code
            except Exception:
                pass
        return codes

    async def _resolve_tm30_day_start(self, property_name: str) -> tuple:
        """(hour, minute) that TM30's own arrival day opens at, per property.

        SEPARATE from RR4's rr4_tm30_day_start_* on purpose: the generator
        sheets declare the two windows independently, and on five of the six
        Thai properties they genuinely differ (Siam's sheet exports RR4 over
        02:15 but TM30 over plain midnight, and so on). Tying TM30 to RR4's
        window would move those five off midnight and break registers that
        currently match their sheet exactly.

        Reads tm30_day_start_hour/_minute from property_api_settings when
        those columns exist (api/sql/tm30_day_window.sql), falling back to
        _TM30_DAY_START_FALLBACK and then to midnight - same degrade-to-
        default shape as get_ftp_settings, so this keeps working before
        anyone has run the migration.
        """
        h, m = _TM30_DAY_START_FALLBACK.get(property_name, (0, 0))
        if self.supabase:
            try:
                res = self.supabase.table("property_api_settings").select(
                    "tm30_day_start_hour, tm30_day_start_minute").eq(
                    "property_name", property_name).limit(1).execute()
                row = res.data[0] if res.data else {}
                # `is not None` rather than `or`: an explicit 0 configured in
                # Admin means midnight and must beat the fallback, not read as
                # "unset" and hand the property back its hardcoded 12:15.
                if row.get("tm30_day_start_hour") is not None:
                    h = row["tm30_day_start_hour"]
                    m = row.get("tm30_day_start_minute") or 0
            except Exception:
                pass
        return h, m

    async def _resolve_tm30_nationality_codes(self) -> dict:
        """Same shape/rationale as _resolve_rr4_nationality_codes, for TM30's
        own admin-editable alpha-3 code table (tm30_nationality_codes, Admin
        > TM30-Nationality). Joined by country name via _RR3_COUNTRY_MAP,
        layered over the hardcoded TM30_NATIONALITY_CODE fallback per-entry."""
        codes = dict(TM30_NATIONALITY_CODE)
        if self.supabase:
            try:
                res = self.supabase.table("tm30_nationality_codes").select(
                    "mews_nationality, tm30_code").execute()
                by_name = {}
                for row in res.data or []:
                    name = (row.get("mews_nationality") or "").strip().lower()
                    code = (row.get("tm30_code") or "").strip()
                    if name and code:
                        by_name[name] = code
                for alpha2, country_name in _RR3_COUNTRY_MAP.items():
                    code = by_name.get(country_name.strip().lower())
                    if code:
                        codes[alpha2] = code
            except Exception:
                pass
        return codes

    @staticmethod
    def _rr4_tm30_identity_card(c: dict) -> str:
        identity_card = c.get("IdentityCard")
        identity_cards = c.get("IdentityCards")
        if isinstance(identity_card, dict):
            return identity_card.get("Number", "")
        if isinstance(identity_cards, list) and identity_cards:
            return identity_cards[0].get("Number", "")
        return ""

    # ---- Manual edit overrides (RR4 & TM30 Files > Edit) ------------------
    #
    # A generated register is not always the one that can be filed: a guest's
    # MEWS profile can be missing a passport number, carry a mis-typed name,
    # or have no Thai name at all, and front office has to correct that
    # before the file goes to the district office.
    #
    # Those corrections live in their own rr4_tm30_overrides table rather
    # than being written back into the rr4_tm30_sync blob, because
    # "Re-Generate Files" and the nightly auto-import both overwrite that
    # blob wholesale - baking an edit into it means one Re-Generate silently
    # un-does a correction on a document that has already been filed. Keeping
    # the two apart also preserves the distinction between what MEWS said and
    # what staff changed, which is exactly what you want to be able to show
    # when a filing is queried.
    #
    # Rows are Fernet-encrypted individually, for the same reason the report
    # blobs are: an override carries precisely the PII the register does
    # (names, passport/ID numbers, addresses).
    OVERRIDES_TABLE = "rr4_tm30_overrides"

    def _rr4_tm30_overrides_map(self, property_name: str, date_str: str, kind: str) -> dict:
        """{row_key: {field: value}} for one (property, date, kind), or {} if
        nothing has been edited - and also {} when the table doesn't exist
        yet (migration not run), the same graceful-degrade shape
        ftp_service.get_ftp_settings uses: an un-migrated database must not
        stop a register being generated."""
        if not self.supabase:
            return {}
        try:
            res = self.supabase.table(self.OVERRIDES_TABLE).select("row_key, data").eq(
                "property", property_name).eq("report_date", date_str).eq("kind", kind).execute()
        except Exception as e:
            logger.info(f"RR4/TM30: no overrides for {property_name} {date_str} {kind} ({e})")
            return {}
        out = {}
        for row in res.data or []:
            try:
                fields = json.loads(encryption_service.decrypt((row.get("data") or {}).get("blob", "")))
            except Exception:
                # A row that won't decrypt (e.g. written under a rotated key)
                # is skipped rather than failing the whole register.
                continue
            if isinstance(fields, dict) and fields:
                out[row["row_key"]] = fields
        return out

    def apply_rr4_tm30_overrides(self, report: dict, property_name: str, date_str: str, kind: str) -> dict:
        """Patches a freshly-built or freshly-decrypted report in place with
        whatever has been edited for that day, stamping each touched row with
        `_edited` (the field names that are no longer MEWS's own answer) so
        the Edit page can mark those cells and offer to reset them.

        Assigns rather than only replacing existing keys: an override may
        legitimately introduce a column the generator never sets - the RR4
        export's four Thai-name columns come back blank from MEWS on every
        guest, and filling one in is a correction, not a stray key."""
        overrides = self._rr4_tm30_overrides_map(property_name, date_str, kind)
        if not overrides:
            return report
        for row in report.get("rows") or []:
            fields = overrides.get(row.get("_key") or "")
            if not fields:
                continue
            for field, value in fields.items():
                row[field] = value
            row["_edited"] = sorted(fields.keys())
        return report

    async def get_rr4_report(self, property_name: str, date: str, apply_overrides: bool = True) -> dict:
        """
        Builds the daily RR4 (ร.ร.๔) hotel guest register: every guest whose
        stay overlaps the given day (Thai and foreign alike), in the Thai
        Hotel Act's column layout - a direct port of the "RR4-TM30-
        Chinatown-Gen" Google Sheet's RR4 tab (fed from MEWS's own "Customer
        profiles In house" report), reverse-engineered from its formulas.
        Nationality codes come from _resolve_rr4_nationality_codes (Admin >
        RR4-Nationality, editable) layered over the hardcoded fallback in
        rr4_tm30_reference.py.

        occupation/willGo/willGoCountry/timeCheckOut are fixed constants in
        the source sheet, not derived per guest - kept identical here rather
        than "improved", since that's what the real filed form already does.

        "Stay overlaps the day" is a standard interval-overlap test
        (StartUtc < window end AND EndUtc > window start) rather than
        BCP's narrower State=="Started" test, so a same-day arrival or
        departure still appears on the register even before front desk has
        checked them in. NOT get_st_files_report's Customers-tab rule
        (stays_the_night: EndUtc > window end, OR day_use: both ends
        inside the window) - that pair works by coincidence when the
        window ends at midnight (every real checkout is "after midnight"),
        but breaks once the window is shifted near typical checkout time
        (e.g. the 12:15-to-12:15 window this page's own Admin > Sync page
        can configure, matching MEWS's native report window exactly) -
        confirmed against a real MEWS export for Chinatown, 15-Aug-2026:
        the old rule silently excluded every guest checking out between
        the window's start-of-day time and its end time, 61 guests short
        with a 12:15 window.

        date_check_in/time_check_in use ActualStartUtc where available, not
        the legacy StartUtc field the day-window/reservation fetch itself
        runs on - same fix as get_rr3_cards for the same reason (StartUtc
        is the deprecated SCHEDULED time, e.g. the hotel's standard 14:00
        check-in, and never updates once the guest actually walks in;
        confirmed against a real reference RR4 for Lub d Bangkok Siam,
        15-Aug-2026: reservation #70105 showed 14:00 here vs the reference's
        correct 20:49, MEWS's own ActualStartUtc). Second, narrow call by
        exact ReservationIds (chunked at 1000/request, MEWS's own limit).
        """
        # The configured window still drives which reservations get FETCHED;
        # the inclusion test below is anchored to the report date's own
        # midnight instead, so neither bound is needed again here.
        day, _day_start_utc, _day_end_utc, reservations, customers_map, resources_map = \
            await self._rr4_tm30_fetch_day(property_name, date)

        actual_times = {}
        reservation_ids = [r["Id"] for r in reservations if r.get("Id")]
        for i in range(0, len(reservation_ids), 1000):
            id_batch = reservation_ids[i:i + 1000]
            try:
                actual_res = await mews_client.post(
                    "/api/connector/v1/reservations/getAll/2023-06-06",
                    {"ReservationIds": id_batch, "Limitation": {"Count": len(id_batch)}},
                    property_name=property_name,
                )
                for r in actual_res.get("Reservations", []):
                    if r.get("Id"):
                        actual_times[r["Id"]] = r.get("ActualStartUtc")
            except Exception as e:
                logger.warning(f"RR4: failed to fetch ActualStartUtc for {property_name}: {e}")

        def parse_utc(ts):
            if not ts:
                return None
            try:
                return datetime.fromisoformat(ts.replace("Z", "+00:00"))
            except Exception:
                return None

        def buddhist_date(ts):
            t = parse_utc(ts)
            if not t:
                return ""
            local = t.astimezone(day.tzinfo)
            return f"{local.day:02d}/{local.month:02d}/{local.year + 543}"

        def time_hhmm(ts):
            t = parse_utc(ts)
            if not t:
                return ""
            local = t.astimezone(day.tzinfo)
            return f"{local.hour:02d}.{local.minute:02d}"

        nationality_codes = await self._resolve_rr4_nationality_codes()

        # The report date's OWN midnight-at-the-end, i.e. 00:00 of the next
        # calendar day, property-local. This - not the configured window's
        # midpoint - is what decides whether a guest "stayed the night" of
        # the report date. See the inclusion test below.
        next_midnight_utc = (day + timedelta(days=1)).astimezone(timezone.utc)
        # ...and the one after it, which decides whether a guest who only
        # arrived after next_midnight_utc belongs to this register or the
        # next day's. See the inclusion test below.
        midnight_after_utc = (day + timedelta(days=2)).astimezone(timezone.utc)

        rows = []
        for res in reservations:
            start_utc = parse_utc(res.get("StartUtc"))
            end_utc = parse_utc(res.get("EndUtc"))
            if not start_utc or not end_utc:
                continue
            # Reproduces MEWS's "Customer profiles / In house" report - the
            # report the ร.ร.๔ file that actually gets filed is generated
            # from (RR4-TM30-<property>-Gen sheet: its RR4 tab is a straight
            # 1-row-per-row transform of an ImportInhouse paste of that
            # export, so matching the export IS matching the filed file).
            #
            # 1. Stayed the night: still resident AT the report date's own
            #    midnight. Note the >=, not > - a guest whose checkout is
            #    booked for exactly 00:00 slept there and MEWS files them
            #    (the five Wiegratz family rows in Lub d Koh Tao 109/110,
            #    20-Aug-2026, every one departing 21-Aug 00:00:00 on the
            #    dot). This clause also replaces an older "past the window's
            #    midpoint" heuristic, which only ever looked right because
            #    Chinatown's then-12:15 window put its midpoint at 00:15;
            #    on a 02:00-ish window the midpoint lands mid-afternoon and
            #    lets in day-use guests who never stayed a night at all
            #    (Koh Samui 1107, 19-Aug: in 13:53, out 18:11).
            #
            # 2. ...unless the whole stay belongs to the NEXT day's
            #    register: arrived after that midnight AND gone again before
            #    the following one, i.e. it covers no night of this date.
            #    A next-day arrival who stays on past the following midnight
            #    does appear here. Both halves are load-bearing, and the
            #    boundary is the second midnight rather than checkout state:
            #    on Chinatown's 20-Aug export, room 333 (in 21-Aug 00:15,
            #    out 21-Aug 12:00) is absent while room 233 (in 21-Aug
            #    00:43 - later - but out 22-Aug 12:00) is present.
            #
            # Verified against the real generator sheets for BOTH 19-Aug and
            # 20-Aug-2026 across six properties and four window settings
            # (Chinatown, Siam 02:15, Koh Samui 02:03, Patong 02:05, Koh Tao
            # 02:05, Marasca 02:03): 10 of those 12 property-days reproduce
            # the export's room set exactly, row counts included. The two
            # that don't are both live-data drift, not rule failures - see
            # the commit that introduced this. An earlier version keyed
            # clause 2 off State=="Processed" instead and matched only 7 of
            # 12; dropping the State test also makes regenerating an old
            # date deterministic, where the State version would quietly lose
            # next-day arrivals once they eventually checked out.
            stayed_the_night = end_utc >= next_midnight_utc
            belongs_to_the_next_register = (
                start_utc >= next_midnight_utc and end_utc <= midnight_after_utc
            )
            if not stayed_the_night or belongs_to_the_next_register:
                continue
            room = resources_map.get(res.get("AssignedResourceId"), {})
            check_in_utc = actual_times.get(res.get("Id")) or res.get("StartUtc")
            attached_count = 0
            headcount = (res.get("AdultCount") or 0) + (res.get("ChildCount") or 0)
            for guest_id in self._rr4_tm30_guest_ids(res):
                c = customers_map.get(guest_id)
                if not c:
                    continue
                attached_count += 1
                nationality_code = c.get("NationalityCode", "")
                rr4_code = nationality_codes.get(nationality_code, "")
                # MEWS's formal country name, not the ร.ร.๓ short one - see
                # _RR4_MEWS_COUNTRY_NAMES. Feeds address / come_from below.
                nationality_name = _rr4_country_name(nationality_code)
                passport = c.get("Passport") or {}
                rows.append({
                    # Stable identity for the manual-edit overrides layer
                    # (Admin-side "Edit" on the RR4 & TM30 Files table) -
                    # reservation id + guest id, so a correction survives a
                    # Re-Generate/re-import even though row_no below is
                    # positional and shifts whenever MEWS data moves.
                    "_key": f"{res.get('Id') or ''}:{guest_id}",
                    "date_check_in": buddhist_date(check_in_utc),
                    "time_check_in": time_hhmm(check_in_utc),
                    "room_no": room.get("Name", ""),
                    "title_en": "MS." if c.get("Sex") == "Female" else "MR.",
                    "name_en": c.get("FirstName", ""),
                    "middle_name_en": c.get("SecondLastName", ""),
                    "surname_en": c.get("LastName", ""),
                    "nationality": rr4_code,
                    "pid": self._rr4_tm30_identity_card(c),
                    "passport": passport.get("Number", ""),
                    "issued_by": rr4_code,
                    "address": nationality_name,
                    "address_country": rr4_code,
                    "occupation": "16",
                    "come_from": nationality_name,
                    "come_from_country": rr4_code,
                    "will_go": "Thailand",
                    "will_go_country": "99",
                    "date_check_out": buddhist_date(res.get("EndUtc")),
                    "time_check_out": "12.00",
                    "data_status": 1,
                    "_sort_start": check_in_utc or "",
                })

            # MEWS "Customer profiles" export creates a placeholder row for
            # every booked guest slot (headcount) even if the companion's
            # profile hasn't been attached yet.
            #
            # Only the four columns the generator sheet keys off the ROOM
            # cell carry a value here (rowNo/dateCheckIn/timeCheckIn/roomNo);
            # every other column in that sheet is wrapped in
            # `if(<nameEn cell>="","",...)` and therefore renders blank on a
            # nameless row - including the ones that are fixed constants on a
            # normal row (occupation "16", willGo "Thailand", willGoCountry
            # "99", timeCheckOut "12.00", dataStatus 1) and dateCheckOut.
            # Filling those in here put 6 spurious values on every placeholder
            # row: 48 cells across the 8 placeholder rows of Lub d Bangkok
            # Siam's 19-Aug-2026 file alone.
            unattached = max(0, headcount - attached_count)
            for slot in range(unattached):
                rows.append({
                    # No guest id to key on - the slot's position within its
                    # own reservation is the only stable handle these have.
                    "_key": f"{res.get('Id') or ''}:placeholder:{slot}",
                    "date_check_in": buddhist_date(check_in_utc),
                    "time_check_in": time_hhmm(check_in_utc),
                    "room_no": room.get("Name", ""),
                    "title_en": "",
                    "name_en": "",
                    "middle_name_en": "",
                    "surname_en": "",
                    "nationality": "",
                    "pid": "",
                    "passport": "",
                    "issued_by": "",
                    "address": "",
                    "address_country": "",
                    "occupation": "",
                    "come_from": "",
                    "come_from_country": "",
                    "will_go": "",
                    "will_go_country": "",
                    "date_check_out": "",
                    "time_check_out": "",
                    "data_status": "",
                    "_sort_start": check_in_utc or "",
                })
        rows.sort(key=lambda r: (r["_sort_start"], r["room_no"]))
        for i, r in enumerate(rows, start=1):
            r["row_no"] = i
            del r["_sort_start"]

        report = {
            "property": property_name,
            "property_thai_name": await self._resolve_rr4_property_thai_name(property_name),
            "date": date,
            "date_buddhist": f"{day.day:02d}/{day.month:02d}/{day.year + 543}",
            "rows": rows,
        }
        # Default on, so every consumer (export, preview, download, on-screen
        # MEWS mode) shows the corrected register. The ONE caller that opts
        # out is the import that writes rr4_tm30_sync - see
        # routers/rr4.py:sync_rr4_tm30_day - which deliberately stores MEWS's
        # raw answer so a reset can fall back to it without a re-fetch.
        if apply_overrides:
            self.apply_rr4_tm30_overrides(report, property_name, date, "rr4")
        return report

    # Verbatim from the reference RR4 sheet's row 1 (bold+italic there, not
    # literal asterisks) - lists which fields are mandatory vs either/or
    # (Thai name OR English name, national ID OR passport). Sits to the
    # left of the title/property-name/ร.ร.๔ cells on the same row.
    _RR4_DISCLAIMER = (
        "ข้อมูลสำคัญที่จำเป็นต้องระบุได้แก่ เลขลำดับ, วันที่เข้าพัก, เวลาที่เข้าพัก, "
        "ห้องพักเลขที่, ชื่อ-นามสกุล ภาษาไทยหรือภาษาอังกฤษ (เลือกระบุอย่างใดอย่างหนึ่ง), "
        "เลขประจำตัวประชาชน หรือ เลขหนังสือเดินทาง (เลือกระบุอย่างใดอย่างหนึ่ง), "
        "ที่อยู่ปัจจุบัน อยู่ที่ ตำบลอำเภอ จังหวัด หรือประเทศ, ประเทศที่อยู่ปัจจุบัน, อาชีพ, "
        "มาจากตำบล อำเภอ จังหวัด หรือประเทศ, มาจากประเทศ, จะไปที่ อำเภอ จังหวัด หรือประเทศใด, "
        "จะไปประเทศ, วันที่จะออก, เวลาที่จะออก"
    )

    # (internal key, Thai header label, English field-key label) - the
    # reference sheet carries BOTH header rows (Thai label row, then a
    # second row of plain English field-key names like the ones already
    # used as this report's own dict keys) rather than just one; the pid/
    # passport/address labels also carry a second instruction line the
    # first pass here had dropped (verified against the reference RR4 for
    # Lub d Bangkok Siam cell-by-cell).
    _RR4_COLUMNS = [
        ("row_no", "เลข\nลำดับ", "rowNo"),
        ("date_check_in", "วันที่เข้าพัก", "dateCheckIn"),
        ("time_check_in", "เวลาที่เข้าพัก", "timeCheckIn"),
        ("room_no", "ห้องพักเลขที่", "roomNo"),
        ("title_th", "คำนำหน้าชื่อ", "titleNameTh"),
        ("name_th", "ชื่อ", "nameTh"),
        ("middle_name_th", "ชื่อกลาง", "middleNameTh"),
        ("surname_th", "นามสกุล", "surnameTh"),
        ("title_en", "คำนำหน้าชื่อ(ภาษาอังกฤษ)", "titleNameEn"),
        ("name_en", "ชื่อ(ภาษาอังกฤษ)", "nameEn"),
        ("middle_name_en", "ชื่อกลาง(ภาษาอังกฤษ)", "middleNameEn"),
        ("surname_en", "นามสกุล(ภาษาอังกฤษ)", "surnameEn"),
        ("nationality", "สัญชาติ", "nationality"),
        ("pid", "เลขประจำตัวประชาชน\nเลขที่..... ออกให้โดย......", "pid"),
        ("passport", "ใบสำคัญประจำตัวคนต่างด้าวหรือหนังสือเดินทาง\nเลขที่..... ออกให้โดย......", "passport"),
        ("issued_by", "หนังสือเดิททางออกให้โดย", "issuedBy"),
        ("address", "ที่อยู่ปัจจุบัน\nอยู่ที่ ตำบลอำเภอ จังหวัด หรือประเทศ", "address"),
        ("address_country", "ประเทศที่อยู่ปัจจุบัน", "addressCountry"),
        ("occupation", "อาชีพ", "occupation"),
        ("come_from", "มาจากตำบล อำเภอจังหวัด หรือประเทศ", "comeFrom"),
        ("come_from_country", "มาจากประเทศ", "comeFromCountry"),
        ("will_go", "จะไปที่ อำเภอจังหวัด หรือประเทศใด", "willGo"),
        ("will_go_country", "จะไปประเทศ", "willGoCountry"),
        ("date_check_out", "วันที่จะออก", "dateCheckOut"),
        ("time_check_out", "เวลาที่จะออก", "timeCheckOut"),
        ("data_status", "สถานะข้อมูล", "dataStatus"),
        ("remarks", "หมายเหตุ", "remarks"),
    ]

    # Columns E-K of the filed RR4, i.e. every name column - the four Thai
    # ones, then the English title/first/middle. Sliced out of _RR4_COLUMNS
    # by position rather than hardcoded, so inserting a column can't quietly
    # shift which cells this reads.
    _RR4_NAME_COLUMN_KEYS = [k for k, _l, _f in _RR4_COLUMNS[4:11]]

    def _rr4_filed_rows(self, rows: list) -> list:
        """
        Drops every row with nothing in columns E-K and renumbers what's
        left, for the .xlsx that actually gets filed.

        Those rows are MEWS's own unnamed occupant slots: "Customer profiles"
        emits one per booked headcount even when the companion's profile was
        never attached, so they carry a check-in, a room and nothing else.
        The generator sheet keeps them (every column past the room is wrapped
        in `if(<nameEn>="","",...)`, which blanks them rather than removing
        the row); the filed file should not, since a row with no name is not
        a lodger the Hotel Act asks us to register.

        Applied at RENDER time, not in get_rr4_report, on purpose: the report
        keeps the slots so Admin > RR4 & TM30 Files > Edit can still show one
        and let someone type the missing guest in. Once named, the row stops
        being blank and appears in the file on its own.

        Runs after the overrides layer for the same reason - a correction
        that supplies the name must be able to save the row, not be
        discarded before it is read.
        """
        kept = [dict(r) for r in rows
                if any(str(r.get(k) or "").strip() for k in self._RR4_NAME_COLUMN_KEYS)]
        # row_no is positional and has to close the gaps the drop leaves, or
        # the filed register counts 1, 2, 3, 5, 6. Renumbered on copies: the
        # caller's report is shared (the email path renders it and then counts
        # it) and must keep the numbering the on-screen table shows.
        for i, row in enumerate(kept, start=1):
            row["row_no"] = i
        return kept

    async def get_rr4_export(self, property_name: str, date: str) -> tuple:
        """Renders get_rr4_report to the .xlsx layout the reference sheet
        uses: row 1 has the required-fields disclaimer, title, property
        name, and "ร.ร.๔" as four separate cells side by side (not the
        property name merged into the date line - confirmed against a real
        reference RR4 for Lub d Bangkok Siam, whose row 1 has the
        disclaimer / "ทะเบียนผู้เข้าพักในโรงแรม" / "โรงแรมหลับดี สยาม" /
        "ร.ร.๔" in that order), row 2 is "ประจำวันที่"/date as the same
        L/N unmerged-cell pair - no merged cells anywhere in this file -
        then the Thai column headers, then one row per guest. Returns
        (bytes, filename)."""
        report = await self.get_rr4_report(property_name, date)
        property_code = property_name
        if self.supabase:
            try:
                prop_res = self.supabase.table("property_api_settings").select("st_property_code").eq(
                    "property_name", property_name).limit(1).execute()
                property_code = (prop_res.data[0].get("st_property_code") if prop_res.data else None) or property_name
            except Exception:
                pass
        return self._render_rr4_export(report, date, property_code)

    def _render_rr4_export(self, report: dict, date: str, property_code: str) -> tuple:
        """The actual .xlsx rendering behind get_rr4_export, split out so the
        RR4/TM30 daily email (sync_service.send_rr4_tm30_property_email/
        send_rr4_tm30_bundled_digest) can render straight from an
        already-fetched report - either the live one get_rr4_export just
        built, or one read back out of rr4_tm30_sync (the storage the daily
        import job already populates) so the email never re-hits MEWS on
        top of that day's import. Same reasoning as get_st_report_export
        reading st_files_sync instead of re-fetching live."""
        n_cols = len(self._RR4_COLUMNS)

        wb = Workbook()
        ws = wb.active
        ws.title = "RR4"

        # Single, unmerged cells - matches the reference sheet's own layout
        # exactly (confirmed cell-by-cell by the user against the real
        # sheet: column A = disclaimer, L = title, N = property name,
        # AA/27 = ร.ร.๔ - not one merged block each). Text simply overflows
        # into the empty cells to its right, same as the reference does, so
        # no wrap_text/row-height override is needed.
        ws.cell(1, 1, self._RR4_DISCLAIMER).font = Font(bold=True, italic=True, size=9)
        ws.cell(1, 1).alignment = Alignment(horizontal="left")
        ws.cell(1, 12, "ทะเบียนผู้เข้าพักในโรงแรม").font = Font(bold=True, size=13)
        ws.cell(1, 12).alignment = Alignment(horizontal="left")
        ws.cell(1, 14, report["property_thai_name"]).font = Font(bold=True, size=13)
        ws.cell(1, 14).alignment = Alignment(horizontal="left")
        ws.cell(1, n_cols, "ร.ร.๔").font = Font(bold=True, size=13)
        ws.cell(1, n_cols).alignment = Alignment(horizontal="left")

        # Same two-cell pattern as row 1's title/property-name pair (L/N,
        # single unmerged cells) - not one merged block across the row.
        # No slashes in the date value - confirmed against reference sheets
        # for both Siam and Chinatown ("15082569", not "15/08/2569").
        ws.cell(2, 12, "ประจำวันที่").font = Font(bold=True, size=13)
        ws.cell(2, 12).alignment = Alignment(horizontal="left")
        ws.cell(2, 14, report["date_buddhist"].replace("/", "")).font = Font(bold=True, size=13)
        ws.cell(2, 14).alignment = Alignment(horizontal="left")

        header_row = 3
        for col, (key, label, _field_key) in enumerate(self._RR4_COLUMNS, start=1):
            cell = ws.cell(header_row, col, label)
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = 14

        # Second header row of plain English field-key names (rowNo,
        # dateCheckIn, ...) - present in the reference sheet as its own row,
        # separate from the Thai labels above.
        field_key_row = header_row + 1
        for col, (key, _label, field_key) in enumerate(self._RR4_COLUMNS, start=1):
            cell = ws.cell(field_key_row, col, field_key)
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # NO literal apostrophe on any column. The importer's own official
        # example file (a plain CSV, not a formula-driven sheet) settles
        # this: every data row's dateCheckIn/dateCheckOut/timeCheckIn/
        # timeCheckOut is a bare value - "05/01/2568", "08.21", "08/01/2568",
        # "12.00" - with no "'" character anywhere. The apostrophe shown once
        # in that file's own instructions ("...Text... ' (...) ...'01/02/2568")
        # is describing the Excel KEYSTROKE convention for a human typing a
        # value into a General-formatted cell (typing ' first forces Excel's
        # live quote-prefix mode, which is a display-only flag - the stored
        # value never gains an actual apostrophe character). A formula like
        # `"'" & TEXT(...)`, which several of the generator sheets use, does
        # NOT reproduce that: string concatenation bakes a literal apostrophe
        # character into the computed value, so it fails the importer's
        # `^\d\d/\d\d/\d{4}$` check outright. A prior version of this comment
        # concluded the opposite from reading those formulas directly - wrong;
        # verified 2026-08-25 against a real "ต้องระบุด้วยรูปแบบ วว/ดด/ปปปป
        # เท่านั้น" rejection on every row after shipping that literal
        # apostrophe, confirming the official example's plain values are the
        # authority, not the generator sheets' formulas. Text number_format
        # ("@") is what actually keeps Excel from reinterpreting the value on
        # open/re-save - no apostrophe, literal or quote-prefixed, is needed.
        for i, row in enumerate(self._rr4_filed_rows(report["rows"]), start=1):
            for col, (key, _label, _field_key) in enumerate(self._RR4_COLUMNS, start=1):
                cell = ws.cell(field_key_row + i, col, row.get(key, ""))
                cell.number_format = "@"

        buf = BytesIO()
        wb.save(buf)

        yyyymmdd = date.replace("-", "")
        filename = f"{property_code}_RR4_{yyyymmdd}.xlsx"
        return buf.getvalue(), filename

    async def get_tm30_report(self, property_name: str, date: str, apply_overrides: bool = True) -> dict:
        """
        Builds the daily TM30 foreign-national arrival notification: guests
        ARRIVING (checking in) on the given day, filtered to non-Thai
        nationals only - a direct port of the reference sheet's TM30-Gen
        tab (fed from MEWS's "Customer profiles Arrival" report) followed by
        its TM30 tab's own filter (`Nationality <> 'THA' AND name is not
        null`). Alpha-3 nationality codes come from
        _resolve_tm30_nationality_codes (Admin > TM30-Nationality, editable)
        layered over the hardcoded fallback in rr4_tm30_reference.py.

        "Arriving" runs over TM30's OWN per-property window
        (_resolve_tm30_day_start, start to start+24h), which is independent
        of both get_st_files_report's Arrivals tab and RR4's
        rr4_tm30_day_start_*. It defaults to local midnight, which is what
        five of the six Thai generator sheets declare; Chinatown's declares
        12:15 and is configured to match it.

        The midnight default is not arbitrary. Confirmed against real MEWS
        "Customer profiles Arrival" exports for Chinatown, 16 & 17-Aug-2026:
        a guest whose single, organic check-in happens in the FIRST HALF of
        their own calendar day still needs to be filed under THAT day, not
        shifted back to the day before - a 12:15 lower bound missed 9 guests
        on the 17th and 3 more on the 16th, e.g. one checking in 09:58am.
        Re-measured on 29-Aug-2026 the same effect costs 11 of that sheet's
        own 67 guests. Chinatown is set to 12:15 anyway, by explicit
        instruction, because matching the filed sheet is this module's
        standing rule - but that is a deliberate trade, not a free one, and
        _TM30_DAY_START_FALLBACK records what it costs.

        Moving the lower bound EARLIER cannot double-file anyone: MEWS's own
        reservations/getAll StartUtc/EndUtc filter already scopes each
        reservation to exactly one day's candidate pool by some mechanism of
        its own (confirmed against 12 guests directly - both the
        newly-included ones and the ones already correctly filed near the
        window's far edge, checking in just after midnight on the NEXT
        calendar day, were each retrievable from only one specific day's raw
        fetch, never both its own day's and the adjacent day's).
        """
        # day_start_utc (RR4's cutoff-hour window start) is intentionally
        # unused below - TM30 resolves its own independent window instead,
        # see _resolve_tm30_day_start and in_window.
        day, _day_start_utc, day_end_utc, reservations, customers_map, resources_map = \
            await self._rr4_tm30_fetch_day(property_name, date)
        tm30_start_hour, tm30_start_minute = await self._resolve_tm30_day_start(property_name)
        nationality_codes = await self._resolve_tm30_nationality_codes()

        def parse_utc(ts):
            if not ts:
                return None
            try:
                return datetime.fromisoformat(ts.replace("Z", "+00:00"))
            except Exception:
                return None

        # The TM30 day is plain local midnight to local midnight, with no
        # relation to the RR4 cutoff-hour window. Confirmed from the
        # generator sheets' own report config: Chinatown's
        # Parameter-ImportCP tab records Start 20-Aug 00:00 / End 21-Aug
        # 00:00 for the very same file whose Parameter-ImportInhouse tab
        # records 02:05 to 02:05, and every other property's Master tab
        # shows the same split (ImportCP always on midnight).
        #
        # A later Chinatown export (23-Aug-2026) showed Parameter-ImportCP's
        # own "Start" field had drifted to read 12:15 too, matching
        # ImportInhouse exactly - which looked at first like the window
        # itself had changed (2 guests, Nam Ann Chia/Yunmi Kang, seemed to
        # confirm it). It hadn't: re-testing a 12:15-anchored window against
        # that day's FULL 71-guest reference list (not just those 2)
        # excluded 13 real guests who check in most days between local
        # midnight and 12:15 and are correctly on the reference sheet
        # (Thomas Herd 12:09, Emma Wolfe 12:03, Laure Cautillo 11:12
        # Bangkok, among others) - proving the Parameter tab's "Start"
        # value isn't authoritative for which guests the Arrival-mode
        # export actually includes, unlike ImportInhouse's In-house-mode
        # report where it is. Plain midnight-to-midnight is correct.
        #
        # The upper bound used to be day_end_utc, the RR4 window's close,
        # which let in guests who arrived in the small hours of the NEXT
        # day: 7 rows on 20-Aug-2026 alone (Chinatown's room 333 pair plus
        # Palangdao, Siam's Ainul and the two Pujols, Patong's Waters), and
        # Patong's Wu Zheng on 19-Aug. Each of them belongs to the next
        # day's notification, and MEWS's own Arrival export excludes them.
        #
        # The whole window now shifts by the property's own configured TM30
        # start (_resolve_tm30_day_start), which is 00:00 for five of the six
        # Thai properties and so leaves this exactly as described above for
        # them. Chinatown's sheet declares 12:15 and is set to match it - see
        # _TM30_DAY_START_FALLBACK for what that costs and why it was chosen.
        # The end is always start + 24h rather than RR4's separately
        # configurable rr4_tm30_day_end_*: an arrival register covers one
        # day, and RR4's end exists to support deliberately non-24h stay
        # windows, which has no meaning here.
        window_start_utc = (day + timedelta(hours=tm30_start_hour,
                                            minutes=tm30_start_minute)).astimezone(timezone.utc)
        window_end_utc = window_start_utc + timedelta(days=1)

        # Note which end is inclusive: a StartUtc landing EXACTLY on the
        # window tick is filed under the day that is ENDING, not the one
        # beginning - so the window is (start, start+24h]. Same principle
        # get_rr4_report already encodes on its own boundary (a checkout
        # booked for exactly 00:00 means the guest slept the previous night -
        # the Wiegratz family, Koh Tao 109/110), which is why both halves
        # flip together rather than this being an arbitrary off-by-one.
        # Confirmed against the 22-Aug-2026 reference sheets: Chinatown's
        # only two midnight-tick guests sit one on each end and BOTH were
        # filed the other way by a [00:00, next 00:00) window - Nam Ann Chia
        # (StartUtc exactly 22-Aug 00:00, on the 21st's sheet, we wrongly
        # added her to the 22nd) and Yunmi Kang (StartUtc exactly 23-Aug
        # 00:00, on the 22nd's sheet, we wrongly dropped her). This form
        # matches all 6 properties exactly, 289 guests, zero diffs; the other
        # five have no midnight-tick guests that day, so it is also a no-op
        # for every one of them.
        def in_window(ts):
            t = parse_utc(ts)
            return t is not None and window_start_utc < t <= window_end_utc

        def gregorian_date(ts):
            t = parse_utc(ts)
            if not t:
                return ""
            # Two different kinds of value reach this helper: check_out_date is
            # a real UTC instant (EndUtc, carries a "Z") and must be shifted
            # into the property's local day, while birth_date is a plain
            # calendar date MEWS returns with no timezone at all. Converting
            # the naive one makes Python read it as the SERVER's local time -
            # which raises OSError 22 on Windows for any pre-1970 birth date
            # (one such guest was enough to fail the whole property's TM30
            # report) and, on any server not behind the property's timezone,
            # silently files the guest one day off. A date with no timezone is
            # already the date to print.
            local = t.astimezone(day.tzinfo) if t.tzinfo else t
            return f"{local.day:02d}/{local.month:02d}/{local.year}"

        rows = []
        for res in reservations:
            if not in_window(res.get("StartUtc")):
                continue
            for guest_id in self._rr4_tm30_guest_ids(res):
                c = customers_map.get(guest_id)
                if not c or not (c.get("FirstName") or c.get("LastName")):
                    continue
                nationality_code = c.get("NationalityCode", "")
                # Mirrors the generator sheet exactly. Its TM30-Gen tab does
                # `iferror(vlookup(<MEWS nationality>, 'TM30-Nationality'!A:D, 4, 0), "Not found")`
                # and its TM30 tab then keeps `where F <> 'THA'` - so an
                # unresolvable nationality yields the literal string "Not found"
                # and the guest STAYS on the filing; only Thai nationals drop out.
                #
                # This used to `continue` on an unresolvable code, which silently
                # dropped the guest from the notification altogether. Found on
                # Marasca Samui, 19-Aug-2026: four guests (Razon Mantelmacher,
                # Mantelmacher, Lahav, Avraham) have no NationalityCode at all on
                # their MEWS profile, so we filed 31 arrivals where the real
                # document has 35. Omitting a foreign national from a TM30 is the
                # worse failure, and "Not found" is visible to whoever checks the
                # file - the actual fix is filling the nationality in on the MEWS
                # profile, which this makes obvious instead of hiding.
                tm30_code = nationality_codes.get(nationality_code, "") or "Not found"
                if tm30_code == "THA":
                    continue  # Thai nationals are out of scope for TM30
                # The column is "เลขหนังสือเดินทาง / Passport No. *", so the
                # passport wins when a guest carries both documents; the
                # identity card is the fallback for the guests who have no
                # passport recorded (27 of 370 arrivals checked). The
                # generator sheet instead does CONCATENATE(identity, passport)
                # and, on the first guest to actually have both - Lub d Koh
                # Tao's Julia Pola Kotowska, 20-Aug-2026 - emits the two
                # jammed together as "04042947201594852404FK2749027", which is
                # not a document number anyone can act on.
                identity_card = self._rr4_tm30_identity_card(c)
                passport = (c.get("Passport") or {}).get("Number", "")
                rows.append({
                    # Same identity scheme as RR4's rows - see its own note.
                    "_key": f"{res.get('Id') or ''}:{guest_id}",
                    "first_name": c.get("FirstName", ""),
                    "middle_name": c.get("SecondLastName", ""),
                    "last_name": c.get("LastName", ""),
                    # Matches the reference sheet's own formula exactly -
                    # defaults to F for anything that isn't literally "Male"
                    # (opposite default from RR4's title, which defaults
                    # unknown/blank Sex to MR. - the two source tabs disagree
                    # with each other, not a bug introduced here).
                    "gender": "M" if c.get("Sex") == "Male" else "F",
                    "passport_no": passport or identity_card,
                    "nationality": tm30_code,
                    "birth_date": gregorian_date(c.get("BirthDate")),
                    "check_out_date": gregorian_date(res.get("EndUtc")),
                    "phone": c.get("Phone", ""),
                })

        report = {
            "property": property_name,
            "property_thai_name": await self._resolve_rr4_property_thai_name(property_name),
            "date": date,
            "rows": rows,
        }
        # See get_rr4_report's note for why this defaults on and who opts out.
        if apply_overrides:
            self.apply_rr4_tm30_overrides(report, property_name, date, "tm30")
        return report

    _TM30_COLUMNS = [
        ("first_name", "ชื่อ\nFirst Name *"), ("middle_name", "ชื่อกลาง\nMiddle Name"),
        ("last_name", "นามสกุล\nLast Name"), ("gender", "เพศ\nGender *"),
        ("passport_no", "เลขหนังสือเดินทาง\nPassport No. *"), ("nationality", "สัญชาติ\nNationality *"),
        # Both date columns carry the government form's own full instruction
        # text, era marker and worked examples included - not an abbreviated
        # "DD/MM/YYYY". The examples are load-bearing on the real form: they
        # are what tell the filer the year is A.D. (ค.ศ.) rather than the
        # Buddhist era used on RR4, and that an unknown day or month is
        # written as 00 rather than left blank. Confirmed character-for-
        # character against the TM30 tab of all six properties' reference
        # sheets, 22-Aug-2026.
        ("birth_date", "วัน เดือน ปี เกิด\nBirth Date\nDD/MM/YYYY(ค.ศ. / A.D.) \nเช่น 17/06/1985 หรือ 10/00/1985 หรือ 00/00/1985"),
        ("check_out_date", "วันที่แจ้งออกจากที่พัก\nCheck-out Date\nDD/MM/YYYY(ค.ศ. / A.D.) \nเช่น 14/06/2023"),
        ("phone", "เบอร์โทรศัพท์\nPhone No."),
    ]

    async def get_tm30_export(self, property_name: str, date: str) -> tuple:
        """Renders get_tm30_report to the .xlsx layout the government
        upload template uses (bilingual header row, one row per foreign
        arrival). Returns (bytes, filename)."""
        report = await self.get_tm30_report(property_name, date)
        property_code = property_name
        if self.supabase:
            try:
                prop_res = self.supabase.table("property_api_settings").select("st_property_code").eq(
                    "property_name", property_name).limit(1).execute()
                property_code = (prop_res.data[0].get("st_property_code") if prop_res.data else None) or property_name
            except Exception:
                pass
        return self._render_tm30_export(report, date, property_code)

    def _render_tm30_export(self, report: dict, date: str, property_code: str) -> tuple:
        """The actual .xlsx rendering behind get_tm30_export - split out for
        the same reason as _render_rr4_export above, same caller."""
        wb = Workbook()
        ws = wb.active
        ws.title = "TM30"

        for col, (key, label) in enumerate(self._TM30_COLUMNS, start=1):
            cell = ws.cell(1, col, label)
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = 18

        for i, row in enumerate(report["rows"], start=1):
            for col, (key, _label) in enumerate(self._TM30_COLUMNS, start=1):
                ws.cell(1 + i, col, row.get(key, ""))

        buf = BytesIO()
        wb.save(buf)

        yyyymmdd = date.replace("-", "")
        filename = f"{property_code}_TM30_{yyyymmdd}.xlsx"
        return buf.getvalue(), filename

    @staticmethod
    def _st_report_counts(report: dict) -> dict:
        """
        Customers/Arrivals/Departures totals for a stored report. Reports
        imported before these were counted separately from their row lists
        fall back to the list lengths, which undercount dorms.
        """
        return {
            key: report.get(f"{key}_count", len(report.get(key, [])))
            for key in ("customers", "arrivals", "departures")
        }

    async def get_st_files_report(self, property_name: str, date: str):
        """
        Builds the daily "ST Files" occupancy/availability report for one
        property + one calendar date in that property's own MEWS-configured
        timezone (see _resolve_property_timezone - Bangkok for every
        property except Makati), replicating the user's manual Google Sheet
        (tabs: Spaces / Occupied / House uses / Out of order / Availability
        / Customers / Arrivals / Departures).

        `date` is YYYY-MM-DD interpreted as that property's own calendar day.
        All MEWS aggregate numbers come per resource category; the category
        list itself requires the Resource Categories permission on the
        property's Connector token (403s cleanly if MEWS hasn't enabled it).
        """
        property_tz = await self._resolve_property_timezone(property_name)
        space_types = await self._resolve_st_space_types(property_name)
        day = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=property_tz)
        day_start_utc = day.astimezone(timezone.utc)
        day_end_utc = (day + timedelta(days=1)).astimezone(timezone.utc)
        start_iso = day_start_utc.strftime("%Y-%m-%dT%H:%M:%SZ")
        end_iso = day_end_utc.strftime("%Y-%m-%dT%H:%M:%SZ")

        # 1. Resolve the bookable (accommodation) service
        services_res = await mews_client.post(
            "/api/connector/v1/services/getAll",
            {"Limitation": {"Count": 100}},
            property_name=property_name,
        )
        stay = self._resolve_stay_service(services_res.get("Services", []))
        if not stay:
            raise Exception(f"No bookable (accommodation) service found for {property_name}")
        service_id = stay["Id"]

        # 2. Categories (names/short codes/types) - requires ServiceIds filter
        cats_res = await mews_client.post(
            "/api/connector/v1/resourceCategories/getAll",
            {"ServiceIds": [service_id], "Limitation": {"Count": 200}},
            property_name=property_name,
        )
        categories = {}
        for c in cats_res.get("ResourceCategories", []):
            if not c.get("IsActive", True):
                continue
            names = c.get("Names") or {}
            shorts = c.get("ShortNames") or {}
            categories[c["Id"]] = {
                "category_id": c["Id"],
                "short_name": shorts.get("en-US") or shorts.get("en-GB") or next(iter(shorts.values()), ""),
                "name": names.get("en-US") or names.get("en-GB") or next(iter(names.values()), ""),
                "type": c.get("Type", ""),
                "capacity": c.get("Capacity"),
                "ordering": c.get("Ordering", 0),
                "in_report": c.get("Type") in space_types,
            }

        def category_rows(count_by_cat_id):
            rows = []
            for cat_id, cat in categories.items():
                if not cat["in_report"]:
                    continue
                rows.append({
                    "short_name": cat["short_name"],
                    "name": cat["name"],
                    "type": cat["type"],
                    "count": count_by_cat_id.get(cat_id, 0),
                    "_ordering": cat["ordering"],
                })
            rows.sort(key=lambda r: r.pop("_ordering"))
            return rows

        # 3. Availability metrics for the day (2024-01-22 version): occupied /
        #    house use / out of order / active per category
        avail_res = await mews_client.post(
            "/api/connector/v1/services/getAvailability/2024-01-22",
            {
                "ServiceId": service_id,
                "FirstTimeUnitStartUtc": start_iso,
                "LastTimeUnitStartUtc": start_iso,
                "Metrics": ["Occupied", "HouseUse", "OutOfOrderBlocks", "ActiveResources"],
            },
            property_name=property_name,
        )
        metric = {"Occupied": {}, "HouseUse": {}, "OutOfOrderBlocks": {}, "ActiveResources": {}}
        for entry in avail_res.get("ResourceCategoryAvailabilities", []):
            for m in metric:
                values = (entry.get("Metrics") or {}).get(m) or [0]
                metric[m][entry["ResourceCategoryId"]] = values[0]

        # 4. Availability (free-to-sell) from the legacy endpoint - MEWS's own
        #    precomputed number, safer than deriving it from raw metrics
        legacy_res = await mews_client.post(
            "/api/connector/v1/services/getAvailability",
            {
                "ServiceId": service_id,
                "FirstTimeUnitStartUtc": start_iso,
                "LastTimeUnitStartUtc": start_iso,
            },
            property_name=property_name,
        )
        availability_by_cat = {}
        for entry in legacy_res.get("CategoryAvailabilities", []):
            values = entry.get("Availabilities") or [0]
            availability_by_cat[entry.get("CategoryId")] = values[0]

        # 5. Resource blocks colliding with the day (named OOO/House-use rows)
        blocks_res = await mews_client.post(
            "/api/connector/v1/resourceBlocks/getAll",
            {"CollidingUtc": {"StartUtc": start_iso, "EndUtc": end_iso}, "Limitation": {"Count": 500}},
            property_name=property_name,
        )
        blocks = [b for b in blocks_res.get("ResourceBlocks", []) if b.get("IsActive", True)]

        # 6. Reservations colliding with the day + their customers/resources
        #    (same un-versioned Extent-join call get_rr3_cards uses)
        resv_res = await mews_client.post(
            "/api/connector/v1/reservations/getAll",
            {
                "StartUtc": start_iso,
                "EndUtc": end_iso,
                "Extent": {"Reservations": True, "Customers": True, "Resources": True},
            },
            property_name=property_name,
        )
        reservations = resv_res.get("Reservations", [])
        customers_map = {c["Id"]: c for c in resv_res.get("Customers", []) if c.get("Id")}
        resources_map = {r["Id"]: r for r in resv_res.get("Resources", []) if r.get("Id")}

        # Real check-in times. StartUtc here is the legacy endpoint's
        # DEPRECATED *scheduled* arrival and never moves once the guest turns
        # up, the same trap CLAUDE.md records as a real bug in RR3 and RR4.
        # Chinatown #97197 on 06-Sep-2026 is the case: scheduled 05-Sep 23:45,
        # actually checked in 06-Sep 00:01, checked out 06-Sep 11:31 - an
        # ordinary one-night stay that crossed midnight on the way in, which
        # we filed on the 5th and MEWS's own export counted on the 6th.
        #
        # Used to fold in as a UNION (scheduled OR actual, whichever matched
        # THIS day) rather than replace StartUtc outright, on the reasoning
        # that a union can only ever ADD a day a reservation qualifies for and
        # so could never cost an already-correct match. That reasoning holds
        # for any single day in isolation and fails across two: Makati #227062
        # on 06/07-Sep-2026 was scheduled 06-Sep 23:45 and actually checked in
        # 07-Sep 03:45. 06-Sep's own report (built that day, matching
        # scheduled) filed it as a 06-Sep arrival; 07-Sep's report (built the
        # next day, matching actual under the union) filed it AGAIN - one
        # guest on two consecutive days' submitted ST files, because each
        # day's report is built independently and a union never revokes a
        # match once the real check-in time moves it elsewhere.
        #
        # Now exactly one canonical instant per reservation - actual when
        # MEWS has recorded it, scheduled otherwise - so a reservation
        # belongs to exactly one day, full stop. This DOES mean a reservation
        # whose actual check-in lands outside the day its own scheduled time
        # would have placed it in is filed under the actual day only, even
        # when that moves it away from a sheet total measured before this
        # existed; a duplicate across two real submitted files is a worse
        # defect than a debatable single-day total, and is not one this
        # report is in a position to accept for the sake of matching a
        # spreadsheet snapshot. ActualStartUtc lives only on the 2023-06-06
        # endpoint, which cannot embed Customers/Resources - hence the second
        # narrow call by ReservationIds, the same one get_rr3_cards makes.
        # Fetched ONLY for the offline rule sweep - production reads
        # StartUtc and never needs this, so the extra round trip is not paid
        # on the nightly import. scripts/st_arrival_sweep.py sets the seam,
        # which is what turns this on.
        actual_start = {}
        if _ST_ARRIVAL_RULE is not None:
            _ids = [r["Id"] for r in reservations if r.get("Id")]
            for _i in range(0, len(_ids), 1000):
                _batch = _ids[_i:_i + 1000]
                try:
                    _ar = await mews_client.post(
                        "/api/connector/v1/reservations/getAll/2023-06-06",
                        {"ReservationIds": _batch, "Limitation": {"Count": len(_batch)}},
                        property_name=property_name,
                    )
                    for _r in _ar.get("Reservations", []):
                        if _r.get("Id") and _r.get("ActualStartUtc"):
                            actual_start[_r["Id"]] = _r["ActualStartUtc"]
                except Exception as e:
                    logger.warning(
                        f"ST Files: could not read ActualStartUtc for {property_name}: {e}")

        # 6b. Complimentary count - Rate AND Segment, see _is_complimentary
        # above for the sheet formula both of them come from.
        rate_ids = list({r.get("RateId") for r in reservations if r.get("RateId")})
        rates_by_id = {}
        if rate_ids:
            try:
                rates_res = await mews_client.post(
                    "/api/connector/v1/rates/getAll",
                    {"RateIds": rate_ids, "Limitation": {"Count": 200}},
                    property_name=property_name,
                )
                rates_by_id = {r["Id"]: r for r in rates_res.get("Rates", []) if r.get("Id")}
            except Exception as e:
                logger.warning(f"ST Files complimentary check: rates lookup failed for {property_name}: {e}")

        # Unfiltered, like _rv_market_segments does it - a property has only a
        # handful of segments, so fetching the lot is cheaper than assembling
        # an id list. Degrades to {} on failure the same way the rates lookup
        # above does: a segment outage must not fail a whole day's report, it
        # just falls back to the Rate arm alone.
        segments_by_id = {}
        try:
            segments_res = await mews_client.post(
                "/api/connector/v1/businessSegments/getAll",
                {"Limitation": {"Count": 200}},
                property_name=property_name,
            )
            # Stripped: MEWS stores the segment name as typed, and several are
            # saved with a trailing space (Koh Tao's are literally "Travel
            # Agent " and "Group Business Series "). An exact lookup silently
            # missed those in the RV report once already - see
            # _rv_market_segments.
            segments_by_id = {s.get("Id"): (s.get("Name") or "").strip()
                              for s in segments_res.get("BusinessSegments", [])}
        except Exception as e:
            logger.warning(f"ST Files complimentary check: business segments lookup failed for {property_name}: {e}")

        complimentary_count = sum(
            1 for r in reservations if self._is_complimentary(r, rates_by_id, segments_by_id)
        )

        # Full resource list + their category assignments. Needed for two
        # things: room names on block rows (the reservation Extent only knows
        # rooms that had a reservation that day), and mapping each reservation
        # to the space category MEWS files it under below.
        all_res = await mews_client.post(
            "/api/connector/v1/resources/getAll",
            {"Extent": {"Resources": True, "ResourceCategoryAssignments": True}, "Limitation": {"Count": 1000}},
            property_name=property_name,
        )
        for r in all_res.get("Resources", []):
            if r.get("Id"):
                resources_map.setdefault(r["Id"], r)
        resource_category = {
            a["ResourceId"]: a["CategoryId"]
            for a in all_res.get("ResourceCategoryAssignments", [])
            if a.get("ResourceId") and a.get("CategoryId")
        }

        # Some spaces are sold both whole and in parts: a dorm ("210") whose
        # children are its individually bookable beds, or a two-bedroom suite
        # ("412/414") whose children are the two rooms. Booking the whole thing
        # assigns the PARENT, whose own category is Type=Dorm/Suite and so is
        # absent from the Room/Bed tables - MEWS still counts it, once per child
        # space, under the children's category. Map parent -> child ids so those
        # reservations can be expanded instead of silently dropped.
        parent_children = {}
        for r in all_res.get("Resources", []):
            parent = r.get("ParentResourceId")
            if parent and r.get("Id"):
                parent_children.setdefault(parent, []).append(r["Id"])

        def block_rows(block_type):
            rows = []
            for b in blocks:
                if b.get("Type") != block_type:
                    continue
                room = resources_map.get(b.get("AssignedResourceId"), {})
                rows.append({
                    "room": room.get("Name", ""),
                    "name": b.get("Name", ""),
                    "notes": b.get("Notes") or "",
                    "start_utc": b.get("StartUtc", ""),
                    "end_utc": b.get("EndUtc", ""),
                })
            rows.sort(key=lambda r: r["room"])
            return rows

        def parse_utc(ts):
            if not ts:
                return None
            try:
                return datetime.fromisoformat(ts.replace("Z", "+00:00"))
            except Exception:
                return None

        def in_window(ts):
            t = parse_utc(ts)
            return t is not None and day_start_utc <= t < day_end_utc

        active_states = {"Confirmed", "Started", "Processed", "Optional"}

        def reservation_row(res):
            customer = customers_map.get(res.get("CustomerId"), {})
            room = resources_map.get(res.get("AssignedResourceId"), {})
            # The room actually ASSIGNED, not RequestedCategoryId - a guest
            # who requested one category and was moved to (or upgraded into)
            # another shows up here under the one they're actually sleeping
            # in, matching the category space_categories() above counts them
            # under. Confirmed against Lub d Siem Reap, 31-Aug-2026: three
            # DKR-requested guests assigned into DTR rooms were filed as DKR
            # arrivals under the old RequestedCategoryId lookup while the
            # aggregate arrivals_count (built from space_categories(), which
            # already resolves off the assignment) correctly counted them as
            # DTR - so the per-category numbers in the exported ST file
            # disagreed with the total next to them. Same resolution order
            # as space_categories() - assigned resource's own category first,
            # then (for a whole-unit booking like a connecting suite whose
            # own category isn't one of this report's space types) the first
            # in-report child, then RequestedCategoryId as a last resort for
            # a reservation with no assignment on record at all. Reservation
            # #150326 that same day (booked on the whole "402/404" connecting
            # suite, itself category DXCN/Suite - not in this report - whose
            # two child rooms are both DCR) is what the child fallback is for.
            assigned_id = res.get("AssignedResourceId")
            assigned_cat_id = resource_category.get(assigned_id)
            if not categories.get(assigned_cat_id, {}).get("in_report"):
                child_cat_id = next(
                    (cc for cc in (resource_category.get(k) for k in parent_children.get(assigned_id, []))
                     if categories.get(cc, {}).get("in_report")),
                    None,
                )
                if child_cat_id:
                    assigned_cat_id = child_cat_id
            cat = categories.get(assigned_cat_id) or categories.get(res.get("RequestedCategoryId"), {})
            return {
                "number": res.get("Number", ""),
                "guest": f"{customer.get('FirstName', '')} {customer.get('LastName', '')}".strip(),
                "nationality": customer.get("NationalityCode", ""),
                "room": room.get("Name", ""),
                "category": cat.get("short_name") or cat.get("name", ""),
                "check_in": res.get("StartUtc", ""),
                "check_out": res.get("EndUtc", ""),
                "state": res.get("State", ""),
                "adults": res.get("AdultCount", 0),
                "children": res.get("ChildCount", 0),
            }

        # How MEWS's own Availability report counts these three, verified
        # category-by-category against its exports (see _ST_REPORT_METRICS):
        # Arrivals/Departures count *spaces*, while Customers counts *people*
        # staying the night, filed under the assigned space's own category - so
        # a whole-dorm or whole-suite booking adds 0 there, its parent category
        # being Type=Dorm/Suite rather than one of the report's own.
        def headcount(res):
            return (res.get("AdultCount") or 0) + (res.get("ChildCount") or 0)

        def space_categories(res):
            """One entry per space the reservation occupies. Booking a parent
            (whole dorm, whole two-bedroom suite) counts once per CHILD space,
            not once per guest - Siem Reap 2026-08-06 is the proof: its two
            arriving suites are 4 DCR rooms, not their 3+4 occupants, and its
            two arriving dorms are 10+10 beds, not 7+8."""
            resource_id = res.get("AssignedResourceId")
            cat_id = resource_category.get(resource_id)
            if categories.get(cat_id, {}).get("in_report"):
                return [cat_id]
            child_cats = [
                resource_category.get(child)
                for child in parent_children.get(resource_id, [])
            ]
            child_cats = [c for c in child_cats if categories.get(c, {}).get("in_report")]
            if child_cats:
                return child_cats
            requested = res.get("RequestedCategoryId")
            if categories.get(requested, {}).get("in_report"):
                return [requested]
            return []

        def requested_is_space_type(res):
            """True if the guest actually booked a per-space (Room/Bed)
            product - as opposed to a whole-unit private-hire product like
            Chinatown's "TRIBE HIDEOUT - ALL YOURS!" (Type=Dorm, its own
            category, distinct from the per-bed "1 BED IN..." product) that
            happens to land on the same parent resource. Res#92258 on
            10-Aug-2026 (7 guests booked into the whole dorm as one MDD-type
            unit) proved MEWS's own Customers figure excludes these, while a
            genuine per-bed guest who merely got assigned the parent resource
            for capacity reasons (the Makati case below) still counts."""
            return categories.get(res.get("RequestedCategoryId"), {}).get("in_report", False)

        arrivals, departures = [], []
        arrivals_count = departures_count = customers_count = 0
        # Units excluded from arrivals by the night-tail half of the day-use
        # rule below - surfaced in the report so a mismatch against the
        # sheet can be told apart from an unrelated bug at a glance (see
        # st_compare_service's use of it) rather than needing this file's
        # own comments re-read every time the mail shows a difference.
        day_use_arrivals_excluded = 0
        night_guest_ids = set()
        for res in reservations:
            if res.get("State") not in active_states:
                continue
            # Arrivals/Departures count spaces (beds/rooms) moving in and out,
            # matching MEWS's own Availability/ST export (e.g. Lub d Siem Reap:
            # 10-bed dorm bookings count as 10 bed arrivals/departures).
            units = len(space_categories(res))
            # The arrival day is the SCHEDULED StartUtc's day, full stop.
            #
            # Two attempts to improve on that with the real check-in time have
            # now been measured and both were wrong. A union (scheduled OR
            # actual) double-counted Makati #227062 across 06 and 07-Sep,
            # putting one guest on two consecutive days' submitted files.
            # Replacing it (actual preferred, scheduled as fallback) fixed
            # that but moved #227062 and #227099 - both scheduled 06-Sep 23:45,
            # both actually checked in in the small hours of the 7th - onto
            # the 7th, where the sheet does not have them, leaving Makati +2
            # on BLDD and DTR.
            #
            # Scored the honest way, cached import against the sheet on
            # 07-Sep-2026 (both frozen within the hour, so drift cannot
            # flatter either side), plain scheduled reproduces all eight
            # properties exactly. The earlier Chinatown case that argued for
            # the real check-in (#97197, 06-Sep) was measured against LIVE
            # MEWS, and Makati has since been shown to move by six
            # reservations in twelve hours - so that measurement is no longer
            # trustworthy and its sheet has rolled over, which means it cannot
            # be re-checked. If it resurfaces, re-measure against a cached
            # import and a sheet of the same date, never against live.
            arrives = in_window(res.get("StartUtc"))
            sched_arrives = arrives
            departs = in_window(res.get("EndUtc"))
            day_use = arrives and departs
            # The Customers tab keeps the SCHEDULED-only reading. Letting the
            # union reach it counted Chinatown 194 against the sheet's 192 on
            # 06-Sep-2026: a guest who checks in at 00:01 and out at 11:31 is
            # an arrival that day (MEWS counts them) but never slept there, and
            # MEWS's Customers figure leaves them out. Arrivals and Customers
            # genuinely answer different questions here.
            sched_day_use = sched_arrives and departs
            # A zero-night ("day use") stay ALWAYS counts as a departure; what
            # varies is whether it also counts as an arrival, and that turns on
            # the hour it started. Measured across three days:
            #  - 26-Aug-2026: four stays that began 00:00-02:00 and ended that
            #    same morning (Chinatown #95528, Siam #71156, Siem Reap #149736
            #    and #149737) were counted as DEPARTURES only. They are the tail
            #    of the night before - the guest walked in after midnight and
            #    left the same morning - so their arrival was already filed
            #    against the previous day.
            #  - 27-Aug-2026: Samui #184016, a 10:00-17:00 day room, was counted
            #    as an ARRIVAL only. Confirmed twice over: that sheet's own MEWS
            #    "Reservation" tab lists 11 DKR arrivals against the Availability
            #    tab's 12, and the one it omits is exactly this stay (the tab
            #    covers in-house guests, and a day room has already left).
            #  - 30-Aug-2026: Patong #190439 (04:40-07:33) and #190440
            #    (04:38-11:33) were counted on BOTH sides - that sheet files 63
            #    arrivals and 80 departures, and no one-side-only rule reaches
            #    both numbers. This is what retired the previous "exactly one
            #    side" model and pulled the cutoff down from 6, which had been
            #    inferred against a much looser upper bound.
            # Multi-night stays are untouched - a 00:00-02:00 check-in that DOES
            # stay the night is an ordinary arrival (Siem Reap #149665,
            # Chinatown #95502).
            #
            # Two of the nine measured stays still come out wrong, and both are
            # on the departure side or unexplainable by any hour: Samui #184016
            # (counted as an arrival but NOT a departure) and Chinatown #96148
            # below. 16 of 18 arrival/departure decisions is the best any tested
            # rule managed; the earlier split forms scored 15 and "count on both
            # sides always" scored 12.
            #
            # Known unexplained, and part of that remaining 2 of 18:
            # Chinatown #96148 and #96160 both started 00:00 on 30-Aug (actual
            # check-ins 00:12 and 01:43) and MEWS counted the FIRST as an
            # arrival and the second not. No field on the two reservations
            # separates them - same state, origin, rate, creator, both walk-ins
            # created the evening before - so no hour cutoff can express it and
            # Chinatown reads one arrival short on days like that.
            night_tail = False
            if day_use:
                started = parse_utc(res.get("StartUtc"))
                # A walk-in is an arrival at any hour - somebody physically
                # turned up at the desk and took a room - so it is never the
                # tail of the night before, however early it starts. See
                # _ST_WALK_IN_RATE_RE for the 14 measured cases this beats the
                # bare hour cutoff on.
                walk_in = bool(_ST_WALK_IN_RATE_RE.search(
                    (rates_by_id.get(res.get("RateId"), {}) or {}).get("Name") or ""))
                night_tail = (not walk_in
                              and started.astimezone(property_tz).hour < _ST_DAY_USE_NIGHT_END_HOUR)
            # Test seam. When _ST_ARRIVAL_RULE is set (only ever by the
            # offline rule-sweep in scripts/, never in production - it is None
            # here and nothing in the app assigns it), that callable decides
            # the arrival outright and the day-use/night-tail pair above is
            # bypassed. It exists so a candidate rule can be scored using this
            # function's REAL category and space-unit resolution rather than a
            # re-implementation of it, which is how the previous rules got
            # scored against numbers that were subtly not the report's own.
            if _ST_ARRIVAL_RULE is not None:
                arrives = bool(_ST_ARRIVAL_RULE(
                    res, actual_start.get(res.get("Id")), in_window, parse_utc,
                    property_tz, rates_by_id, day_start_utc, day_end_utc))
                day_use = night_tail = False
            if arrives and not (day_use and night_tail):
                arrivals.append({**reservation_row(res), "spaces": units})
                arrivals_count += units
            elif arrives and day_use and night_tail:
                day_use_arrivals_excluded += units
            # Departures take EVERY zero-night stay, whichever side of the
            # cutoff it started. The earlier "one side only" split was wrong
            # here: Patong on 30-Aug-2026 filed 63 arrivals AND 80 departures,
            # and the only way to reach both is for its two 04:30 day rooms to
            # appear on both sides. Scored against all nine measured day-use
            # stays (26-Aug, 27-Aug and 30-Aug), this form gets 16 of 18
            # arrival/departure decisions right where either split form got 15.
            if departs:
                departures.append({**reservation_row(res), "spaces": units})
                departures_count += units
            # MEWS counts a guest against the day they slept there, plus
            # same-day stays that never saw a night at all. So the only
            # reservations left out are the ones that merely checked OUT
            # during the day - they belong to the previous day's file.
            end_utc = parse_utc(res.get("EndUtc"))
            stays_the_night = end_utc is not None and end_utc > day_end_utc
            if (stays_the_night or sched_day_use) and units and requested_is_space_type(res):
                customers_count += headcount(res)
                for cid in ([res.get("CustomerId")] + (res.get("CompanionIds") or [])):
                    if cid:
                        night_guest_ids.add(cid)

        customers = []
        for cid in night_guest_ids:
            c = customers_map.get(cid)
            if not c:
                continue
            customers.append({
                "name": f"{c.get('FirstName', '')} {c.get('LastName', '')}".strip(),
                "nationality": c.get("NationalityCode", ""),
                "email": c.get("Email", ""),
                "phone": c.get("Phone", ""),
            })
        customers.sort(key=lambda c: c["name"])
        arrivals.sort(key=lambda r: r["room"] or "zzz")
        departures.sort(key=lambda r: r["room"] or "zzz")

        # Reservation tab - every reservation colliding with the day (the
        # same population complimentary_count is computed from above), with
        # Rate/State/Complimentary shown per-row so the aggregate Complimentary
        # number on the Statistic Data tabs and the ST Files List table can be
        # audited reservation-by-reservation instead of taken on faith.
        reservation_audit_rows = []
        for res in reservations:
            customer = customers_map.get(res.get("CustomerId"), {})
            room = resources_map.get(res.get("AssignedResourceId"), {})
            cat = categories.get(res.get("RequestedCategoryId"), {})
            rate = rates_by_id.get(res.get("RateId"), {})
            reservation_audit_rows.append({
                "number": res.get("Number", ""),
                "guest": f"{customer.get('FirstName', '')} {customer.get('LastName', '')}".strip(),
                "room": room.get("Name", ""),
                "category": cat.get("short_name") or cat.get("name", ""),
                "rate": rate.get("Name", ""),
                # Carried alongside the rate because it is half of what decides
                # the flag below, and on some properties it is the ONLY half
                # that fires - without it, a row flagged complimentary on a
                # perfectly ordinary-looking rate reads as a bug.
                "segment": segments_by_id.get(res.get("BusinessSegmentId"), ""),
                "state": res.get("State", ""),
                "check_in": res.get("StartUtc", ""),
                "check_out": res.get("EndUtc", ""),
                "complimentary": self._is_complimentary(res, rates_by_id, segments_by_id),
            })
        reservation_audit_rows.sort(key=lambda r: r["room"] or "zzz")

        spaces = category_rows(metric["ActiveResources"])
        report = {
            "parameters": {
                "property": property_name,
                "service": stay.get("Name", ""),
                "date": date,
                "space_types": list(space_types),
                "generated_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            },
            "spaces": spaces,
            "occupied": category_rows(metric["Occupied"]),
            "house_use": category_rows(metric["HouseUse"]),
            "house_use_blocks": block_rows("InternalUse"),
            "out_of_order": category_rows(metric["OutOfOrderBlocks"]),
            "out_of_order_blocks": block_rows("OutOfOrder"),
            "availability": category_rows(availability_by_cat),
            "customers": customers,
            "arrivals": arrivals,
            "departures": departures,
            # Counts are stored rather than derived from the row lists: one
            # reservation can be several spaces (dorms), and a night's guests
            # outnumber the profiles MEWS actually names.
            "customers_count": customers_count,
            "arrivals_count": arrivals_count,
            "departures_count": departures_count,
            "complimentary": complimentary_count,
            "reservations": reservation_audit_rows,
            # See day_use_arrivals_excluded's own comment above - how many
            # arrival units the night-tail rule held back today.
            "day_use_arrivals_excluded": day_use_arrivals_excluded,
        }
        return report

    # MEWS hard-caps a single services/getAvailability/2024-01-22 call's
    # interval: FirstTimeUnitStartUtc/LastTimeUnitStartUtc 99 days apart
    # succeeds, 100 is rejected outright ("The interval must not exceed
    # 100D" - confirmed empirically against a live property). Chunk width
    # stays comfortably under that; the max total range is "about a year"
    # with a little slack for leap years, not tied to the chunk size -
    # get_occupancy_report stitches as many chunks together as the range
    # needs.
    _OCCUPANCY_CHUNK_DAYS = 95
    # rates/getPricing takes a far wider interval than getAvailability but is
    # still capped: "The interval must not exceed 367D". Chunked well under
    # that so a whole-month-aligned year (up to 395 days) splits cleanly in
    # two rather than sitting one day inside the limit.
    _RATE_CHUNK_DAYS = 300
    # A year, with room for the daily snapshot's whole-month alignment: it
    # runs 1st-of-this-month through end-of-the-same-month-next-year (see
    # occupancy.snapshot_range), which tops out at 395 days.
    _OCCUPANCY_MAX_RANGE_DAYS = 400

    async def get_occupancy_report(self, property_name: str, start_date: str, end_date: str) -> dict:
        """
        Occupancy % per space category per night, over a date range - the
        same shape as MEWS's own "Availability report" Occupancy tab
        (categories down, dates across, a weighted Total row at the bottom).

        Built from services/getAvailability/2024-01-22, which returns a
        value PER TIME UNIT - so a whole range costs one request per
        _OCCUPANCY_CHUNK_DAYS-wide slice, not one per day. MEWS hard-caps a
        single call's interval at 99 days apart (100 is rejected outright:
        "The interval must not exceed 100D", confirmed empirically), so a
        year-long sweep is stitched together from several calls rather than
        one - see the chunking loop below. Occupied / ActiveResources per
        category per day is occupancy, and the Total row is the sum of both
        across the counted categories rather than an average of the
        percentages - a 40-bed dorm at 20% must not weigh the same as a
        4-room category at 100%.

        Categories are filtered to the property's own ST space types
        (Room/Bed by default, see _resolve_st_space_types) for the same
        reason get_st_files_report does it: a parent Dorm/Apartment
        category covers the very same physical beds as its child Bed
        category, so counting both double-counts the property. Verified
        against Lub d Bangkok Chinatown, 20-Aug-2026 - 126 occupied of 176
        active = 71.59%, matching that day's ST Files Occupied/Spaces
        exactly.

        FirstTimeUnitStartUtc has to land on a real time-unit boundary or
        MEWS answers "is not start of TimeUnit" - it is the property's own
        local midnight, not UTC midnight.
        """
        tz = await self._resolve_property_timezone(property_name)
        try:
            first = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=tz)
            last = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=tz)
        except ValueError:
            raise ValueError("start_date and end_date must be YYYY-MM-DD")
        if last < first:
            raise ValueError("end_date is before start_date")
        if (last - first).days > self._OCCUPANCY_MAX_RANGE_DAYS:
            raise ValueError("Range too large - about a year at a time at most")

        services_res = await mews_client.post(
            "/api/connector/v1/services/getAll", {}, property_name=property_name)
        stay = self._resolve_stay_service(services_res.get("Services", []))
        if not stay:
            raise ValueError(f"No active bookable service found for {property_name}")
        service_id = stay["Id"]

        cats_res = await mews_client.post(
            "/api/connector/v1/resourceCategories/getAll",
            {"ServiceIds": [service_id]}, property_name=property_name)
        space_types = await self._resolve_st_space_types(property_name)
        categories = {}
        for c in cats_res.get("ResourceCategories", []):
            categories[c["Id"]] = {
                "short_name": (c.get("ShortNames") or {}).get("en-US") or "",
                "name": (c.get("Names") or {}).get("en-US") or c.get("Name") or "",
                "type": c.get("Type"),
                "ordering": c.get("Ordering") if c.get("Ordering") is not None else 9999,
            }

        # Split into <=_OCCUPANCY_CHUNK_DAYS-wide slices for MEWS's 99-day
        # cap. Both bounds of a getAvailability call are inclusive day
        # markers (a 92-day gap comes back as 93 dates), so consecutive
        # chunks are built to TOUCH - chunk i's Last equals chunk i+1's
        # First - and the shared boundary day is kept only once, dropped
        # from the tail of every chunk but the final one.
        chunk_bounds: list[tuple[datetime, datetime]] = []
        if last == first:
            chunk_bounds = [(first, last)]
        else:
            cursor = first
            while cursor < last:
                nxt = min(cursor + timedelta(days=self._OCCUPANCY_CHUNK_DAYS), last)
                chunk_bounds.append((cursor, nxt))
                cursor = nxt

        chunk_infos = []
        seen_cids: set = set()
        for idx, (c_first, c_last) in enumerate(chunk_bounds):
            avail_res = await mews_client.post(
                "/api/connector/v1/services/getAvailability/2024-01-22",
                {
                    "ServiceId": service_id,
                    "FirstTimeUnitStartUtc": c_first.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "LastTimeUnitStartUtc": c_last.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "Metrics": ["Occupied", "ActiveResources"],
                },
                property_name=property_name,
            )
            # MEWS echoes the time units it actually used; deriving the
            # column labels from those rather than from our own loop keeps
            # the header honest if it ever snaps a chunk to different
            # boundaries.
            chunk_dates = []
            for ts in avail_res.get("TimeUnitStartsUtc") or []:
                try:
                    chunk_dates.append(datetime.fromisoformat(ts.replace("Z", "+00:00")).astimezone(tz).strftime("%Y-%m-%d"))
                except Exception:
                    continue
            if not chunk_dates:
                chunk_dates = [(c_first + timedelta(days=i)).strftime("%Y-%m-%d")
                               for i in range((c_last - c_first).days + 1)]
            is_final = idx == len(chunk_bounds) - 1
            keep = len(chunk_dates) if is_final else max(0, len(chunk_dates) - 1)

            by_cat = {}
            for entry in avail_res.get("ResourceCategoryAvailabilities", []):
                cid = entry.get("ResourceCategoryId")
                meta = categories.get(cid)
                if not meta or meta["type"] not in space_types:
                    continue
                by_cat[cid] = entry.get("Metrics") or {}
                seen_cids.add(cid)
            chunk_infos.append({"dates": chunk_dates, "keep": keep, "by_cat": by_cat})

        dates = []
        for ci in chunk_infos:
            dates.extend(ci["dates"][:ci["keep"]])
        n = len(dates)

        rows = []
        total_occupied = [0] * n
        total_active = [0] * n
        for cid in seen_cids:
            meta = categories[cid]
            occupied: list = []
            active: list = []
            # Every chunk contributes exactly `keep` values for every
            # category we've seen anywhere in the range - even a chunk
            # where MEWS's response omitted this category entirely - so
            # the merged series always lines up with `dates` regardless of
            # which chunk actually returned data for it.
            for ci in chunk_infos:
                m = ci["by_cat"].get(cid, {})
                clen = len(ci["dates"])
                occ = (m.get("Occupied") or [])[:clen]
                act = (m.get("ActiveResources") or [])[:clen]
                occ += [0] * (clen - len(occ))
                act += [0] * (clen - len(act))
                occupied.extend(occ[:ci["keep"]])
                active.extend(act[:ci["keep"]])
            for i in range(n):
                total_occupied[i] += occupied[i]
                total_active[i] += active[i]
            rows.append({
                "short_name": meta["short_name"],
                "name": meta["name"],
                "type": meta["type"],
                "_ordering": meta["ordering"],
                "occupied": occupied,
                "active": active,
                # None (not 0) for a category with no active spaces that
                # day, so the page can show "-" instead of a false 0%.
                "percent": [round(100 * occupied[i] / active[i], 2) if active[i] else None
                            for i in range(n)],
            })
        rows.sort(key=lambda r: (r["_ordering"], r["short_name"] or r["name"]))
        for r in rows:
            r.pop("_ordering", None)

        return {
            "property": property_name,
            "service": stay.get("Name"),
            "space_types": list(space_types),
            "start_date": start_date,
            "end_date": end_date,
            "dates": dates,
            "categories": rows,
            "total": {
                "occupied": total_occupied,
                "active": total_active,
                "percent": [round(100 * total_occupied[i] / total_active[i], 2) if total_active[i] else None
                            for i in range(n)],
            },
        }

    async def get_rate_report(self, property_name: str, start_date: str, end_date: str) -> dict:
        """
        Nightly sales rate (net value) per space category, over a date
        range - the Rate tab of MEWS's own "Availability report" (see a
        real export's Parameters sheet: Rate mode "Sales rate", Rates
        "Flexible Rate Room Only", Amount part "Net value").

        The rate priced is the property's own IsDefault rate - MEWS's own
        flag for "the" rate, distinct from IsBaseRate (which ~449 rates on
        Chinatown alone carry, since it just means "not derived from
        another rate" - every independent promo/corporate rate qualifies).
        IsDefault, by contrast, resolved to exactly one rate on every one
        of six properties checked, always named "Flexible Rate Room Only"
        with ShortName "BAR" - Best Available Rate, matching the reference
        export exactly.

        Chunked like get_occupancy_report, just far less finely:
        rates/getPricing accepts a much wider interval than
        services/getAvailability (99 days) but is still capped -
        "The interval must not exceed 367D" - so the whole-month-aligned
        year the daily snapshot asks for, up to 395 days, does not fit in
        one call. Chunks are built to touch and the shared boundary day is
        kept once, exactly as in get_occupancy_report.

        Categories are filtered to the same space_types get_occupancy_report
        uses (Room/Bed by default) for the same reason: a parent Dorm/
        Apartment category's price would otherwise repeat its own child
        Bed category's price as a second row.

        Deliberately no Total row: Occupied/Active sum into a meaningful
        portfolio occupancy percentage, but summing or averaging list
        prices across different room types has no standard meaning, and
        the reference export's Rate mode doesn't compute one either.
        """
        tz = await self._resolve_property_timezone(property_name)
        try:
            first = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=tz)
            last = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=tz)
        except ValueError:
            raise ValueError("start_date and end_date must be YYYY-MM-DD")
        if last < first:
            raise ValueError("end_date is before start_date")
        if (last - first).days > self._OCCUPANCY_MAX_RANGE_DAYS:
            raise ValueError("Range too large - about a year at a time at most")

        services_res = await mews_client.post(
            "/api/connector/v1/services/getAll", {}, property_name=property_name)
        stay = self._resolve_stay_service(services_res.get("Services", []))
        if not stay:
            raise ValueError(f"No active bookable service found for {property_name}")
        service_id = stay["Id"]

        cats_res = await mews_client.post(
            "/api/connector/v1/resourceCategories/getAll",
            {"ServiceIds": [service_id]}, property_name=property_name)
        space_types = await self._resolve_st_space_types(property_name)
        categories = {}
        for c in cats_res.get("ResourceCategories", []):
            categories[c["Id"]] = {
                "short_name": (c.get("ShortNames") or {}).get("en-US") or "",
                "name": (c.get("Names") or {}).get("en-US") or c.get("Name") or "",
                "type": c.get("Type"),
                "ordering": c.get("Ordering") if c.get("Ordering") is not None else 9999,
            }

        rates_res = await mews_client.post(
            "/api/connector/v1/rates/getAll", {}, property_name=property_name)
        default_rate = next(
            (r for r in rates_res.get("Rates", [])
             if r.get("ServiceId") == service_id and r.get("IsDefault")),
            None,
        )
        if not default_rate:
            raise ValueError(f"No default (BAR) rate found for {property_name}")

        chunk_bounds: list = []
        if last == first:
            chunk_bounds = [(first, last)]
        else:
            cursor = first
            while cursor < last:
                nxt = min(cursor + timedelta(days=self._RATE_CHUNK_DAYS), last)
                chunk_bounds.append((cursor, nxt))
                cursor = nxt

        currency = None
        chunk_infos = []
        seen_cids: set = set()
        for idx, (c_first, c_last) in enumerate(chunk_bounds):
            pricing_res = await mews_client.post(
                "/api/connector/v1/rates/getPricing",
                {
                    "RateId": default_rate["Id"],
                    "FirstTimeUnitStartUtc": c_first.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "LastTimeUnitStartUtc": c_last.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                },
                property_name=property_name,
            )
            currency = currency or pricing_res.get("Currency")
            chunk_dates = []
            for ts in pricing_res.get("DatesUtc") or []:
                try:
                    chunk_dates.append(datetime.fromisoformat(ts.replace("Z", "+00:00")).astimezone(tz).strftime("%Y-%m-%d"))
                except Exception:
                    continue
            if not chunk_dates:
                chunk_dates = [(c_first + timedelta(days=i)).strftime("%Y-%m-%d")
                               for i in range((c_last - c_first).days + 1)]
            is_final = idx == len(chunk_bounds) - 1
            keep = len(chunk_dates) if is_final else max(0, len(chunk_dates) - 1)

            by_cat = {}
            for entry in pricing_res.get("CategoryPrices", []):
                cid = entry.get("CategoryId")
                meta = categories.get(cid)
                if not meta or meta["type"] not in space_types:
                    continue
                by_cat[cid] = entry.get("Prices") or []
                seen_cids.add(cid)
            chunk_infos.append({"dates": chunk_dates, "keep": keep, "by_cat": by_cat})

        dates = []
        for ci in chunk_infos:
            dates.extend(ci["dates"][:ci["keep"]])

        rows = []
        for cid in seen_cids:
            meta = categories[cid]
            # Every chunk contributes exactly `keep` values for every category
            # seen anywhere in the range - including a chunk whose response
            # omitted it - so the merged series always lines up with `dates`.
            prices: list = []
            for ci in chunk_infos:
                clen = len(ci["dates"])
                p = list(ci["by_cat"].get(cid, []))[:clen]
                p += [None] * (clen - len(p))
                prices.extend(p[:ci["keep"]])
            rows.append({
                "short_name": meta["short_name"],
                "name": meta["name"],
                "type": meta["type"],
                "_ordering": meta["ordering"],
                "prices": prices,
            })
        rows.sort(key=lambda r: (r["_ordering"], r["short_name"] or r["name"]))
        for r in rows:
            r.pop("_ordering", None)

        return {
            "property": property_name,
            "rate_name": default_rate.get("Name"),
            "currency": currency,
            "space_types": list(space_types),
            "start_date": start_date,
            "end_date": end_date,
            "dates": dates,
            "categories": rows,
        }

    async def get_st_files_list(self, property_name: str) -> list:
        """
        ST Files List's per-day summary rows - reads exclusively from
        st_files_sync (Database-sourced, per request), not live MEWS -
        recomputing every historical day live would mean ~6 MEWS calls per
        row for what could be months of history. Each row's totals are
        summed from that day's already-stored report blob (the same
        numbers the single-day tabs above show), not recomputed here.
        """
        if not self.supabase:
            return []
        res = self.supabase.table("st_files_sync") \
            .select("report_date, data, synced_at") \
            .eq("property", property_name) \
            .order("report_date", desc=True) \
            .execute()
        rows = []
        for row in res.data or []:
            blob = (row.get("data") or {}).get("blob", "")
            if not blob:
                continue
            try:
                report = json.loads(encryption_service.decrypt(blob))
            except Exception as e:
                logger.warning(f"ST Files List: failed to decrypt {property_name}/{row.get('report_date')}: {e}")
                continue
            rows.append({
                "date": row.get("report_date"),
                "spaces": sum(c.get("count", 0) for c in report.get("spaces", [])),
                "occupied": sum(c.get("count", 0) for c in report.get("occupied", [])),
                "house_use": sum(c.get("count", 0) for c in report.get("house_use", [])),
                "out_of_order": sum(c.get("count", 0) for c in report.get("out_of_order", [])),
                "availability": sum(c.get("count", 0) for c in report.get("availability", [])),
                **self._st_report_counts(report),
                # Reports imported before the complimentary field existed
                # (see get_st_files_report) simply show 0 here, not an error.
                "complimentary": report.get("complimentary", 0),
                # Same fallback reasoning: a report imported before this
                # field existed shows 0 rather than an error.
                "day_use_arrivals_excluded": report.get("day_use_arrivals_excluded", 0),
                "synced_at": row.get("synced_at"),
            })
        return rows

    def _get_st_report_row(self, property_name: str, date_str: str) -> dict:
        """
        Property Code + the 10 ST metric totals for one already-imported
        (property, date) report - shared by get_st_report_export (the
        pipe-delimited CSV) and the ST Files daily digest email's summary
        table, so both read the exact same numbers off the same blob.
        Raises the same two ValueErrors get_st_report_export always has:
        missing Property Code, or no imported report for that date.
        """
        prop_res = self.supabase.table("property_api_settings").select("st_property_code").eq(
            "property_name", property_name).limit(1).execute()
        property_code = (prop_res.data[0].get("st_property_code") if prop_res.data else None)
        if not property_code:
            raise ValueError(f"No ST Property Code configured yet for {property_name} - set it in Admin > API Settings")
        res = self.supabase.table("st_files_sync").select("data").eq(
            "property", property_name).eq("report_date", date_str).limit(1).execute()
        if not res.data:
            raise ValueError(f"No imported report for {property_name} on {date_str} - import it first")
        blob = (res.data[0].get("data") or {}).get("blob", "")
        report = json.loads(encryption_service.decrypt(blob))
        totals = {
            "spaces": sum(c.get("count", 0) for c in report.get("spaces", [])),
            "occupied": sum(c.get("count", 0) for c in report.get("occupied", [])),
            "house_use": sum(c.get("count", 0) for c in report.get("house_use", [])),
            "out_of_order": sum(c.get("count", 0) for c in report.get("out_of_order", [])),
            "availability": sum(c.get("count", 0) for c in report.get("availability", [])),
            **self._st_report_counts(report),
            "complimentary": report.get("complimentary", 0),
        }
        return {"property_code": property_code, "totals": totals}

    # The 10 metric columns for the ST Files daily digest email's summary
    # table, in the same order _ST_REPORT_METRICS uses (confirmed against
    # the source Google Sheet). "No. of Day" has no totals key - like the
    # CSV export, it's always 1 (this row represents one day of data).
    _ST_FILES_TABLE_METRIC_COLUMNS = [
        ("Spaces", "spaces"),
        ("Occupied", "occupied"),
        ("House Uses", "house_use"),
        ("Out of Order", "out_of_order"),
        ("Availability", "availability"),
        ("Customers", "customers"),
        ("Arrivals", "arrivals"),
        ("Departures", "departures"),
        ("Complimentary", "complimentary"),
        ("No. of Day", None),
    ]

    @classmethod
    def _build_st_files_summary_table(cls, rows: list) -> str:
        """
        Pre-built HTML for the <<StatsTable>> token in the ST Files daily
        digest email (Admin > Templates > ST Files Email) - same "dynamic
        row count, can't be a simple <<Token>> substitution" reasoning as
        RR3's <<IdBoxes>>. rows: list of {"property_name", "property_code",
        "totals": {...}} in display order (one per property included in
        that day's send). Table-based layout + inline styles throughout,
        matching every other email in this app, for the same Outlook
        compatibility reason.
        """
        if not rows:
            return '<p style="margin:0; font-size:12px; color:#152A00; opacity:0.6;">No properties included.</p>'

        header_cells = "".join(
            f'<th style="padding:8px 6px; text-align:center; font-size:8px; font-weight:700; '
            f'text-transform:uppercase; letter-spacing:0.03em; color:#FFEFD2; line-height:1.3;">{label}</th>'
            for label, _ in cls._ST_FILES_TABLE_METRIC_COLUMNS
        )
        body_rows = []
        for i, row in enumerate(rows):
            totals = row["totals"]
            bg = "#ffffff" if i % 2 == 0 else "#FFEFD2"
            data_cells = "".join(
                f'<td style="padding:7px 6px; text-align:center; font-size:12px; color:#152A00; '
                f'font-variant-numeric:tabular-nums;">{1 if key is None else totals.get(key, 0)}</td>'
                for _, key in cls._ST_FILES_TABLE_METRIC_COLUMNS
            )
            body_rows.append(
                f'<tr style="background:{bg}; border-bottom:1px solid rgba(21,42,0,0.08);">'
                f'<td style="padding:7px 10px; text-align:left; font-size:12px; color:#152A00; '
                f'font-weight:700; white-space:nowrap;">{_escape_html(row["property_name"])}</td>'
                f'<td style="padding:7px 6px; text-align:center; font-size:12px; color:#152A00; '
                f'font-variant-numeric:tabular-nums;">{_escape_html(row["property_code"])}</td>'
                f'{data_cells}</tr>'
            )
        return (
            '<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="margin:0 0 20px 0;">'
            '<thead><tr style="background:#152A00;">'
            '<th style="padding:8px 10px; text-align:left; font-size:8px; font-weight:700; text-transform:uppercase; '
            'letter-spacing:0.03em; color:#FFEFD2; white-space:nowrap;">Property</th>'
            '<th style="padding:8px 6px; text-align:center; font-size:8px; font-weight:700; text-transform:uppercase; '
            'letter-spacing:0.03em; color:#FFEFD2;">Code</th>'
            f'{header_cells}</tr></thead>'
            f'<tbody>{"".join(body_rows)}</tbody>'
            '</table>'
        )

    def get_st_report_export(self, property_name: str, date_str: str) -> tuple:
        """
        Builds the legacy pipe-delimited "ST" statistics submission file
        (record types PMSST/RMSST) for one already-imported day - 10 fixed
        metric rows, 41 fields each, values taken from that day's
        st_files_sync blob (same totals get_st_files_list shows).

        Field layout confirmed directly against the source Google Sheet
        formula (not just inferred from sample output):
          ="PMSST|RMSST|"&"90107"&"|"&text(date,"ddmmyyyy")&"|"&"0"&text(date,"mm")
            &"|"&text(date,"yyyy")&"|"&"ST"&text(date,"yyyymmdd")&"|"&"Spaces"&"|"
            &text(date,"ddmmyyyy")&"|"&"THB"&"|"&Master!B4&"|1||||D|"&Master!$B$3
            &"|111|ZZ|ZZ|..."
        Field 5 is "0"+month (a report-period code, NOT a property code -
        "008" in the original sample was just August, field 6's year makes
        it unambiguous). Field 17 IS the real per-property value (Master!$B$3
        in their sheet, e.g. "SM" for Lub d Bangkok Siam) - sourced from
        property_api_settings.st_property_code here; raises if that property
        hasn't been given one yet rather than guessing at it for what could
        be a real government/PMS submission. Fields 12/16/18/19/20 are
        hardcoded literals in the formula itself, confirmed constant for
        every property.

        Returns (csv_text, filename) - filename is "{property_code}_ST_
        {yyyymmdd}.csv", e.g. "SM_ST_20260805.csv" for Lub d Bangkok Siam.
        """
        if not self.supabase:
            raise Exception("Supabase not initialized")
        row = self._get_st_report_row(property_name, date_str)
        property_code, totals = row["property_code"], row["totals"]
        day = datetime.strptime(date_str, "%Y-%m-%d")
        ddmmyyyy = day.strftime("%d%m%Y")
        yyyymmdd = day.strftime("%Y%m%d")
        month_code = "0" + day.strftime("%m")
        currency = self._ST_REPORT_CURRENCY.get(property_name, "THB")
        main_ref = ddmmyyyy if property_name in self._ST_REPORT_DDMMYYYY_REF_PROPERTIES else yyyymmdd
        lines = []
        for code, label, key, record_type in self._ST_REPORT_METRICS:
            value = 1 if key is None else totals.get(key, 0)
            # "No. of Day" always references ddmmyyyy regardless of property;
            # every other row follows main_ref (see _ST_REPORT_DDMMYYYY_REF_PROPERTIES).
            ref_suffix = ddmmyyyy if key is None else main_ref
            fields = [
                "PMSST", record_type, code, ddmmyyyy, month_code, str(day.year),
                f"ST{ref_suffix}", label, ddmmyyyy, currency, str(value), "1",
                "", "", "", "D", property_code, "111", "ZZ", "ZZ",
            ] + [""] * 21
            # ASCII-folded for the same reason the RV file is (see
            # _ascii_fold) - the labels here are fixed literals, but the
            # property code and currency come from admin-entered config that
            # nothing else constrains to ASCII.
            lines.append(_ascii_fold("|".join(fields)))
        filename = f"{property_code}_ST_{yyyymmdd}.csv"
        return "\n".join(lines), filename

    # ------------------------------------------------------------------ RV files
    #
    # The RV ("Revenue") file is the money-side sibling of the ST statistics
    # file - same 41-field Infor SunSystems journal layout, same per-property
    # code in field 17, but journal type PMSRV and one row per GL account
    # instead of ten fixed metric rows.
    #
    # Payment Type (+ the External sub-type) -> GL account + department.
    # Confirmed row-by-row against a real MEWS-generated RV file
    # (MS_RV_20260807.csv, Lub d Bangkok Chinatown). The payment half of the
    # file needs no extra MEWS permission: every distinction the file draws is
    # already on the payment object itself.
    # Each property posts to its OWN chart of accounts - they are not variations
    # on a theme. Siem Reap books Guest Ledger to 11401 where Chinatown uses
    # 21203, online payments to 11208 rather than 11399, laundry to 30450 rather
    # than 30445, and carries accounts Chinatown has no equivalent for at all
    # (11416 breakfast, 11704 in-country VAT, 30505 commission). Every entry
    # below was read off that property's own MEWS-generated RV file.
    #
    # MEWS puts an AccountingCategoryId GUID on each item; accountingCategories/
    # getAll (permission enabled 2026-08-17) now resolves it to a real GL code,
    # via sync_rv_gl_mappings_from_mews -> rv_gl_mappings, checked first in
    # _rv_revenue_gl. This chart is the fallback for whichever categories that
    # table doesn't (yet) cover: first BillingName substring hit wins, so
    # specific entries come before generic ones. A property with neither a
    # chart entry here nor any rv_gl_mappings rows cannot be exported at all
    # (see get_rv_export) rather than silently borrowing another's accounts.
    _RV_CHARTS = {
        "Lub d Bangkok Chinatown": {
            "payments": {
                "CashPayment":                   ("11110", "000"),
                "CreditCardPayment":             ("11403", "000"),
                "ExternalPayment/OnlinePayment": ("11399", "000"),
                "ExternalPayment/WireTransfer":  ("11399", "000"),
                "ExternalPayment/Prepayment":    ("21201", ""),
                "ExternalPayment/Complimentary": ("60300", "241"),
            },
            "revenue": [
                ("room adjustment", "30005", "111"),
                ("room upgrade",    "30005", "111"),
                ("no show",         "30010", "111"),
                ("early check in",  "30020", "111"),
                ("late check out",  "30025", "111"),
                ("breakfast",       "30105", "121"),
                ("laundry",         "30445", "141"),
                ("night",           "30001", "111"),
                # A tiny rounding remainder MEWS posts as its own order item
                # rather than folding into the Night line - confirmed against
                # the file, which books it to the same 30001/111 as Night.
                ("room revenue",    "30001", "111"),
            ],
            "fallback": ("30550", "122"),        # products / minibar / retail
            # Service charge is its own account but keeps the DEPARTMENT of
            # whatever it was charged on: "Included Breakfast SVC (Adults)"
            # books to 30805/121 while a plain room service charge is 30805/111.
            "service_charge": ("30805", "111"),
            "vat": ("21600", ""),
            "guest_ledger": ("21203", "000"),
        },
        "Lub d Siem Reap": {
            "payments": {
                "CashPayment":                   ("11110", "0"),
                "CreditCardPayment":             ("11403", "0"),
                "ExternalPayment/OnlinePayment": ("11208", "000"),
                "ExternalPayment/Invoice":       ("11402", "0"),
            },
            "revenue": [
                ("rounding adjustment", "30005", "111"),
                ("no show",            "30010", "111"),
                ("included breakfast", "11416", "000"),
                ("10% vat",            "11704", "000"),
                ("laundry",            "30450", "141"),
                ("commission",         "30505", "134"),
                ("night",              "30001", "111"),
                ("room revenue",       "30001", "111"),
                ("extra person",       "30001", "111"),
            ],
            # Everything else at Siem Reap is an activity or incidental (tours,
            # lost key card, blanket upgrade) and books to 30815 - the one
            # inferred entry here, from every remaining item on both verified
            # days landing there. Siem Reap posts no service charge at all.
            "fallback": ("30815", "141"),
            "service_charge": None,
            "vat": ("21600", ""),
            "guest_ledger": ("11401", "000"),
        },
        "Lub d Philippines Makati": {
            "payments": {
                "CashPayment":                   ("11110", "0"),
                "CreditCardPayment":             ("11403", "0"),
                # Agoda's own prepayment account code, not a numbered GL - the
                # file writes it literally as "DAG01" in field 3, confirmed
                # against every one of Makati's 50 prepayment rows.
                "ExternalPayment/Prepayment":    ("DAG01", ""),
                "ExternalPayment/OnlinePayment": ("11406", "0"),
                "ExternalPayment/WireTransfer":  ("11406", "0"),
                "ExternalPayment/Invoice":       ("11402", "0"),
                # MEWS's catch-all external sub-type. Makati is the only
                # property that actually uses it (an inflight-menu charge
                # settled outside every other method); its own account, not
                # folded into 11406 with the online/wire externals.
                "ExternalPayment/Unspecified":   ("11800", "0"),
            },
            "revenue": [
                ("room adjustment", "30005", "111"),
                ("room upgrade",    "30005", "111"),
                ("room charge discount", "30005", "111"),
                ("early check in",  "30020", "111"),
                ("late check out",  "30025", "111"),
                ("breakfast",       "11416", "000"),
                ("laundry",         "30450", "141"),
                ("transfer",        "30505", "134"),
                # Not really "revenue" - a local-tax accrual liability - but it
                # shares the same BillingName-driven journal mechanism, so it's
                # listed here to keep it off the 30815 fallback, which is
                # genuine miscellaneous revenue.
                ("accrued - local tax", "21620", "0"),
                # A real order item, not a TaxValues component - a correction
                # against the VAT account itself, so it belongs on 21600 like
                # the VAT line it's adjusting, not the misc fallback.
                ("vat adjustment",  "21600", "0"),
                ("night",           "30001", "111"),
            ],
            "fallback": ("30815", "141"),
            "service_charge": ("30805", "111"),
            "service_charge_exempt_gl": ("11416",),
            "vat": ("21600", ""),
            # PH-S is the real 12% VAT and is all "VAT" is in the file. The
            # PH-MA-CUSTOM* municipal surcharges share the same GL (21600) but
            # post as their OWN separate line labeled "Makati local tax" -
            # confirmed against the file, which shows both as distinct rows
            # under 21600, not one combined total - so they ride the
            # secondary_tax mechanism below instead of being summed in here.
            "vat_tax_codes": {"PH-S"},
            # Makati has no per-reservation "Service Charge" order item at all
            # - unlike every other property, its 10% service charge exists
            # ONLY as this TaxRateCode, and the file turns it into a single
            # synthetic revenue line rather than a real transaction. Confirmed
            # by both the amount (matches to within a day's normal drift) and
            # field 20 carrying the same "ABBSU"/"V07" tax marker the VAT line
            # itself carries.
            "service_charge_tax_code": "PH-MA-SERVICE-1%",
            "secondary_tax": {"codes": {"PH-MA-CUSTOM-0.6%", "PH-MA-CUSTOM-0.75%"}, "gl_code": "21600",
                              "department": "", "label": "Makati local tax"},
            "guest_ledger": ("11401", "000"),
        },
        "Lub d Phuket Patong": {
            "payments": {
                "CashPayment":                   ("11110", "000"),
                "CreditCardPayment":             ("11403", "000"),
                # Patong books its online gateway to the card account (11403),
                # not a separate one - confirmed against all 5 online rows in
                # its file, distinct from every other property checked so far.
                "ExternalPayment/OnlinePayment": ("11403", "000"),
                "ExternalPayment/Complimentary": ("11399", ""),
                "ExternalPayment/Invoice":       ("11402", "000"),
            },
            "revenue": [
                ("room adjustment",     "30005", "111"),
                ("room upgrade",        "30005", "111"),
                ("rounding adjustment", "30005", "111"),
                ("no show",             "30010", "111"),
                ("breakfast",           "11416", "000"),
                ("laundry",             "30445", "141"),
                ("transfer",            "30505", "134"),
                ("cancellation",        "30705", "111"),
                ("night",               "30001", "111"),
            ],
            "fallback": ("30815", "141"),
            "service_charge": ("30805", "111"),
            "service_charge_exempt_gl": ("11416",),
            "vat": ("21600", ""),
            "vat_tax_codes": {"TH-2024-7%"},
            # Provincial Tax is a genuinely separate 1% levy MEWS reports as
            # its own TaxRateCode - unlike Makati's duplicate, this one really
            # is money collected and belongs on its own line, confirmed exact
            # against the file (Koh Samui and Marasca both matched to the
            # satang on this same split).
            "secondary_tax": {"codes": {"TH-PROVINCIAL-1%"}, "gl_code": "21609", "department": "", "label": "Provincial Tax"},
            "guest_ledger": ("11401", "000"),
        },
        "Lub d Bangkok Siam": {
            "payments": {
                "CashPayment":                   ("11110", "0"),
                "CreditCardPayment":             ("11403", "0"),
                "ExternalPayment/OnlinePayment": ("11399", "0"),
                # City-ledger invoices, the same 11402 every other property
                # books them to - department "000" here rather than the "0"
                # this property's other payment rows carry, taken from the
                # file's own KILROY invoice row rather than assumed.
                "ExternalPayment/Invoice":       ("11402", "000"),
            },
            "revenue": [
                ("no show",     "30010", "111"),
                ("rounding adjustment", "30005", "111"),
                # Siam calls its room service charge "Accommodation Service",
                # which the generic service-charge matcher does not catch, so
                # it is routed explicitly. Listed before "accommodation charge"
                # since both share a prefix.
                ("accommodation service", "30805", "111"),
                ("accommodation charge",  "30001", "111"),
                # "Service Charge F&B" doesn't mention "breakfast" or any other
                # base-product keyword, so the generic same-department
                # inference (base[1] from a matched base product) can't find
                # it and it fell back to department 111 - confirmed against
                # the file, which books it to 121 like breakfast itself.
                ("service charge f&b", "30805", "121"),
                ("breakfast",   "30145", "121"),
                ("laundry",     "30445", "141"),
                ("miscellaneous", "30815", "141"),
                ("night",       "30001", "111"),
            ],
            "fallback": ("30550", "122"),
            "service_charge": ("30805", "111"),
            "vat": ("21600", ""),
            # Unlike every other property, Siam's real file books VAT as one
            # row PER market segment (100/101/102/blank) rather than one
            # combined total - the four rows sum to exactly the same figure a
            # single row would, confirmed against the file, so this is purely
            # how the line is split, not a different total.
            "vat_by_segment": True,
            "guest_ledger": ("21203", "000"),
        },
        "Lub d Koh Samui Chaweng Beach": {
            "payments": {
                "CashPayment":                   ("11110", "000"),
                "CreditCardPayment":             ("11403", "000"),
                # Wire and PayPal are the only externals with their own GL;
                # a card brand routed through ExternalPayment (Visa/MasterCard/
                # Amex sub-type, seen when a card is charged manually rather
                # than through the terminal) is booked to the same 11403 as a
                # normal card payment, not split out.
                "ExternalPayment/WireTransfer":  ("11204", "000"),
                "ExternalPayment/PayPal":        ("11200", "000"),
                "ExternalPayment/Visa":          ("11403", "000"),
                "ExternalPayment/MasterCard":    ("11403", "000"),
                "ExternalPayment/Amex":          ("11403", "000"),
                # City-ledger invoices (this property's file books DTR02 and
                # the SLH/Hilton settlements here), same 11402 as elsewhere.
                "ExternalPayment/Invoice":       ("11402", "000"),
            },
            "revenue": [
                ("room adjustment", "30005", "111"),
                ("room upgrade",    "30005", "111"),
                # Matches both "Late Check Out" and the file's actual
                # "Late Check-out" (hyphenated, lowercase o) - confirmed live.
                ("late check",      "30025", "111"),
                ("no show",         "30010", "111"),
                ("breakfast",       "11416", "000"),
                ("laundry",         "30445", "141"),
                ("transfer",        "30505", "134"),
                ("cancellation",    "30705", "111"),
                ("night",           "30001", "111"),
            ],
            "fallback": ("30815", "141"),
            "service_charge": ("30805", "111"),
            "service_charge_exempt_gl": ("11416",),
            "vat": ("21600", ""),
            "vat_tax_codes": {"TH-2024-7%"},
            "secondary_tax": {"codes": {"TH-PROVINCIAL-1%"}, "gl_code": "21609", "department": "", "label": "VAT"},
            "guest_ledger": ("21203", "000"),
        },
        # Koh Tao books almost everything through its outlet tills, so its
        # revenue GLs come entirely from the synced accountingCategories
        # overrides (verified exact against its own file: room 237,226.22,
        # F&B 22,793.93/30,834.54, laundry, transport, retail all to the
        # satang) and this chart carries only what an override cannot supply -
        # the payment accounts, the two tax buckets and the balancing row.
        "Lub d Koh Tao Tanote Bay": {
            "payments": {
                "CashPayment":       ("11110", "0"),
                "CreditCardPayment": ("11403", "0"),
                "ExternalPayment/Invoice":    ("11402", "0"),
                # A card externally charged (not through the terminal) still
                # settles to the same 11403 as a normal card payment - the
                # same rule Koh Samui's chart already uses for its own
                # ExternalPayment/Visa|MasterCard|Amex rows. Found via a
                # 21-day sweep (2026-08-23) that also caught Visa going
                # unmapped 10 of those days.
                "ExternalPayment/MasterCard": ("11403", "0"),
                "ExternalPayment/Visa":       ("11403", "0"),
                # Same sweep: WireTransfer went unmapped 17 of 21 days -
                # this property's single biggest export-blocking gap. Not a
                # guess: Koh Tao's own outlet till posts a front-desk-style
                # bank transfer as the literal line "BANK TRANSFER" to this
                # exact GL/department, confirmed against the real file on
                # both 21- and 22-Aug-2026. A wire transfer settled through
                # MEWS's payments feed instead of the till is the same money
                # by a different door.
                "ExternalPayment/WireTransfer": ("11399", "0"),
            },
            "revenue": [],
            "vat": ("21600", ""),
            "vat_tax_codes": {"TH-2024-7%"},
            "secondary_tax": {"codes": {"TH-PROVINCIAL-1%"}, "gl_code": "21609", "department": "", "label": "VAT"},
            "guest_ledger": ("21203", "000"),
        },
        "Marasca Samui": {
            "payments": {
                "CashPayment":                   ("11110", "000"),
                "CreditCardPayment":             ("11403", "000"),
                "ExternalPayment/WireTransfer":  ("11203", "000"),
                "ExternalPayment/Invoice":       ("11402", "000"),
            },
            "revenue": [
                ("room adjustment", "30005", "111"),
                ("extra person charge", "30005", "111"),
                # A real order item literally named "VAT", separate from the
                # TaxValues-derived total already summed elsewhere - confirmed
                # against the file, which carries both as distinct 21600 rows.
                ("vat",             "21600", "000"),
                # "transfer" (any wording) and "airport" both land on the same
                # GL/department - order between them doesn't matter.
                ("transfer",        "30505", "134"),
                ("airport",         "30505", "134"),
                ("breakfast include in room", "30100", "121"),
                ("breakfast",       "30105", "121"),
                # Order matters: the "- Dinner"/"- Lunch"/outlet-specific
                # variants must be checked before their plain "Food"/
                # "Beverage" counterparts, which would otherwise swallow them
                # (both share the same substring).
                ("beverage cabanas - dinner",       "30215", "121"),
                ("discount beverage cabanas - dinner", "30215", "121"),
                ("bottle of sparkling wine", "30215", "128"),
                ("beverage in room dining", "30210", "128"),
                ("mixology masterclass", "30210", "121"),
                ("beverage",         "30210", "121"),
                ("food cabanas - dinner",  "30115", "121"),
                ("discount food cabanas - dinner", "30115", "121"),
                ("food in room dining - dinner", "30115", "128"),
                ("food in room dining - lunch", "30110", "128"),
                ("food the pantry - dinner", "30115", "126"),
                ("food the pantry - lunch", "30110", "126"),
                ("social dining",   "30115", "121"),
                ("food",              "30110", "121"),
                # Spa treatments are named after the treatment (Aromatherapy,
                # Thai Yoga Massage, ...), never containing the word "spa"
                # itself - listed explicitly. "spa " (with the trailing
                # space) still catches "Spa Service Charge" without also
                # matching unrelated items like "Bottle of Sparkling Wine".
                ("aromatherapy",     "30545", "136"),
                ("thai yoga massage", "30545", "136"),
                ("promotion of the month", "30545", "136"),
                ("spa ",              "30545", "136"),
                # A tip is a guest-owed liability the property passes through
                # to staff, not the property's own revenue - it belongs on
                # 21668, matching how Koh Tao's file treats the same line.
                ("tip",                "21668", "0"),
                # Only beer/alcohol brands have their own account (30240);
                # every other snack item (Salted Mixed Roots/Potato/Cashews/
                # Peanuts, ...) shares the generic 30140 fallback below.
                ("white claw",       "30240", "127"),
                ("asahi beer",       "30240", "127"),
                ("san miguel",       "30240", "127"),
                ("singha",             "30240", "127"),
                # Marasca's per-outlet service charge lines don't share a word
                # with their own outlet's food/beverage line ("Service Charge
                # Cabanas" vs "Food Cabanas"), so the generic same-department
                # inference below can't find them - listed explicitly instead.
                ("service charge cabanas",         "30805", "121"),
                ("service charge in room dining",  "30805", "128"),
                ("service charge the pantry",      "30805", "126"),
                ("service charge nightly",         "30805", "111"),
                ("service charge fb", "30805", "121"),
                # A distinct GL from the 30140 retail-snacks fallback -
                # confirmed against the file's "Miscellaneous Cabanas" line.
                ("miscellaneous cabanas",           "30815", "121"),
                # 61000 is Marasca's own Resort Credit account - doesn't exist
                # on any other property's chart.
                ("resort credit",    "61000", "242"),
                ("night",            "30001", "111"),
            ],
            "fallback": ("30140", "127"),   # retail snacks bucket
            "service_charge": ("30805", "111"),
            "vat": ("21600", ""),
            "vat_tax_codes": {"TH-2024-7%"},
            "secondary_tax": {"codes": {"TH-PROVINCIAL-1%"}, "gl_code": "21609", "department": "", "label": "VAT"},
            "guest_ledger": ("21203", "000"),
        },
    }

    # Some properties assign a NUMBERED code to a segment that every other
    # property leaves blank - Koh Tao books "Travel Agent" to 107 rather than
    # blank, confirmed against its own file (five segments reconciled exactly,
    # with no blank Night bucket at all). This overrides the shared table
    # above for that one property/segment pair; Koh Tao itself is not on the
    # verified-properties list below for an unrelated reason (see the note on
    # _RV_CHARTS), but the override is recorded now so it's ready once that's
    # resolved.
    _RV_MARKET_SEGMENT_OVERRIDES = {
        "Lub d Koh Tao Tanote Bay": {"Travel Agent": "107"},
    }

    # Market segment codes ARE shared across the group by default - eight
    # properties' files agree on the common ones - so this is keyed by segment
    # name, not by property, with _RV_MARKET_SEGMENT_OVERRIDES above for the
    # one known exception. "104" is a coarser SunSystems-side bucket that more
    # than one MEWS business segment rolls into (Makati books both "Corporate
    # FIT" and "Government" to it). "106" is likewise a Group bucket that both
    # the Leisure and Business pairs roll into - confirmed against Koh Samui's
    # 09-Aug-2026 file, where a "Group Business Series" reservation (Night
    # 1,769.16 + its 10% service charge) posted to segment 106 exactly like
    # a Group Leisure one; the code previously only listed the Leisure pair,
    # so a Group Business booking fell through to blank instead. "Travel
    # Agent" really does map to a blank code at every property except the
    # override. A segment not listed here also yields blank, which is a value
    # the files themselves use and so cannot break the import.
    _RV_MARKET_SEGMENT_CODES = {
        "Own Web":              "100",
        "Online Travel Agent":  "101",
        "Direct Host":          "102",
        "Direct Reservation":   "103",
        "Corporate FIT":        "104",
        "Government":           "104",
        "Social Chats":         "105",
        "Group Leisure Series": "106",
        "Group Business Series": "106",
        "Group Business Ad-Hoc": "106",
        "Group Leisure Ad-Hoc": "106",
        "Travel Agent":         "",
    }

    # MEWS caps a single getAll response and hands back a Cursor to continue
    # from. Both of the RV report's own feeds really do overflow it: Lub d Koh
    # Samui posted 1,133 order items on 21-Aug-2026, so the un-paged call
    # silently returned the first 1,000 and dropped 133 - revenue that simply
    # vanished from the journal with nothing to show anything was missing.
    # Stop on a short page rather than on a missing Cursor: MEWS returns a
    # Cursor on the LAST page too (Chinatown's 642-item day comes back with
    # one), so trusting it alone would loop forever re-reading the same tail.
    _RV_PAGE = 1000

    async def _rv_fetch_paged(self, property_name: str, endpoint: str,
                              collection: str, body: dict) -> list:
        out, cursor = [], None
        while True:
            limitation = {"Count": self._RV_PAGE}
            if cursor:
                limitation["Cursor"] = cursor
            res = await mews_client.post(
                endpoint, {**body, "Limitation": limitation}, property_name=property_name)
            chunk = res.get(collection) or []
            out.extend(chunk)
            cursor = res.get("Cursor")
            if not cursor or len(chunk) < self._RV_PAGE:
                return out

    @staticmethod
    def _rv_payment_key(payment: dict) -> str:
        """Payments split across three different GL accounts purely on their
        External sub-type (Online/Wire -> 11399 but Prepayment -> 21201 and
        Complimentary -> 60300), so the sub-type is part of the lookup key."""
        ptype = payment.get("Type") or ""
        if ptype == "ExternalPayment":
            sub = ((payment.get("Data") or {}).get("External") or {}).get("Type") or ""
            return f"{ptype}/{sub}"
        return ptype

    # MEWS spells its card brands differently from the journal file
    # ("MasterCard" on the API, "Mastercard" in the file); anything not listed
    # is passed through as MEWS spells it.
    _RV_CARD_TYPE_LABELS = {"MasterCard": "Mastercard", "Jcb": "JCB"}
    _RV_EXTERNAL_LABELS = {
        "OnlinePayment": "Online payment",
        "WireTransfer": "Wire transfer",
        "Prepayment": "Prepayment",
        "Complimentary": "Complimentary",
        "MasterCard": "Mastercard",   # a card externally charged (not via terminal) still gets MEWS's card-brand spelling
    }

    # Was {"Marasca Samui": "Cash on Sales"}, on the reading that Marasca
    # labels every cash payment that way. It doesn't: "Cash on Sales" is the
    # name its POS till gives its OWN cash line, and those rows now arrive
    # from outletItems carrying that name themselves. Marasca's front-desk
    # cash payments read "Cash payment THB (Bill#05477)" like everywhere
    # else, so the override was renaming real payments to a label that
    # belongs to a different ledger. Kept as an empty hook for any property
    # that genuinely does relabel cash.
    _RV_CASH_LABEL_OVERRIDES: dict = {}

    async def _rv_market_segments(self, property_name: str, items: list, stay_service_id) -> dict:
        """Maps stay-service ServiceOrderIds to their market segment code.

        Only items sold on the Reservable stay service carry a segment.
        Verified against the real file: accommodation, its service charge and
        its included breakfast all carry one, while everything billed on the
        other services (Accommodation Extras, Merchandise, Food & Beverage,
        Laundry) leaves field 22 blank. The filter is required rather than an
        optimisation - those other services' ServiceOrderIds are not
        reservations, and reservations/getAll rejects an entire batch that
        contains a single non-reservation id.

        Degrades to {} (every row blank) if the lookups fail."""
        codes = {**self._RV_MARKET_SEGMENT_CODES, **self._RV_MARKET_SEGMENT_OVERRIDES.get(property_name, {})}
        if not stay_service_id:
            return {}
        ids = sorted({i["ServiceOrderId"] for i in items
                      if i.get("ServiceId") == stay_service_id and i.get("ServiceOrderId")})
        if not ids:
            return {}
        try:
            business_segment_ids = {}
            for start in range(0, len(ids), 100):
                res = await mews_client.post(
                    "/api/connector/v1/reservations/getAll",
                    {"ReservationIds": ids[start:start + 100], "Limitation": {"Count": 100}},
                    property_name=property_name,
                )
                for r in res.get("Reservations", []):
                    business_segment_ids[r.get("Id")] = r.get("BusinessSegmentId")
            segments_res = await mews_client.post(
                "/api/connector/v1/businessSegments/getAll",
                {"Limitation": {"Count": 200}}, property_name=property_name,
            )
        except Exception as e:
            logger.info(f"RV: market segments unavailable for {property_name} ({e})")
            return {}
        # Stripped, because MEWS stores whatever was typed into the segment
        # name and several are saved with a trailing space - Koh Tao's are
        # literally "Travel Agent " and "Group Business Series ". An exact
        # lookup silently missed those and left the code blank, which is how
        # six of its rows (its whole 107 bucket, Night 37,500.51 and the
        # service charge and breakfast that follow the same reservations)
        # went out unsegmented against a file that segments them.
        names = {s.get("Id"): (s.get("Name") or "").strip()
                 for s in segments_res.get("BusinessSegments", [])}
        return {rid: codes.get(names.get(bsid) or "", "")
                for rid, bsid in business_segment_ids.items()}

    async def _rv_load_credit_cards(self, property_name: str, card_ids) -> dict:
        """Card rows name the card itself - "Card payment (Mastercard ****8119
        Virtual, B46FWB)" - but payments/getAll carries only a CreditCardId;
        the brand, obfuscated number and Virtual flag come from
        creditCards/getAll. Returns {} if the property's token lacks that
        permission, in which case the description falls back to the payment's
        own identifier rather than failing the whole report."""
        ids = [i for i in (card_ids or []) if i]
        if not ids:
            return {}
        cards = {}
        for start in range(0, len(ids), 100):
            try:
                res = await mews_client.post(
                    "/api/connector/v1/creditCards/getAll",
                    {"CreditCardIds": ids[start:start + 100], "Limitation": {"Count": 100}},
                    property_name=property_name,
                )
            except Exception as e:
                logger.info(f"RV: creditCards/getAll unavailable for {property_name} ({e})")
                return {}
            for card in res.get("CreditCards", []):
                cards[card.get("Id")] = card
        return cards

    def _rv_payment_description(self, pay: dict, cards: dict, property_name: str = "",
                                  force_payment_verb: bool = False) -> str:
        """Rebuilds the per-transaction narrative MEWS writes into field 8,
        verified line-by-line against the real file."""
        ptype = pay.get("Type") or ""
        amount = pay.get("Amount") or {}
        # Kept verbatim rather than trimmed: front desk often types a trailing
        # space ("PAID EXTEND ") and the real file preserves it inside the
        # brackets. Only the emptiness test ignores whitespace.
        notes = pay.get("Notes") or ""
        if not notes.strip():
            notes = ""
        # MEWS returns money in as negative, so a positive amount is money
        # going back out. Every payment kind is renamed for that case in the
        # real file - "Cash refund THB", "Card refund (...)", "External refund
        # (Prepayment)" - and it is only the wording that changes; the D/C flag
        # is derived from the amount separately.
        #
        # GhostPayment is the one exception: MEWS narrates BOTH sides of a
        # Ghost pair as "payment", never "refund", regardless of which way the
        # amount points - confirmed against Chinatown's 22-Aug-2026 retry pair,
        # whose credit-side Ghost row still reads "Card payment (...)" not
        # "Card refund (...)". A Ghost isn't a real refund action a guest
        # authorised; it's MEWS's own internal correction, and its wording
        # says so.
        verb = "payment" if force_payment_verb else (
            "refund" if (amount.get("NetValue") or 0.0) > 0 else "payment")

        if ptype == "CashPayment":
            override = self._RV_CASH_LABEL_OVERRIDES.get(property_name)
            if override:
                return override
            text = f"Cash {verb} {amount.get('Currency') or ''}".strip()
            return f"{text} ({notes})" if notes else text

        if ptype == "CreditCardPayment":
            card = cards.get(((pay.get("Data") or {}).get("CreditCard") or {}).get("CreditCardId"))
            parts = []
            if card:
                brand = self._RV_CARD_TYPE_LABELS.get(card.get("Type"), card.get("Type") or "")
                # ObfuscatedNumber is usually just the last 4 digits ("9428"),
                # but a Physical card can carry the full BIN+last4 masked PAN
                # ("478448******7470") and a Virtual token can be all mask
                # with no real digits at all ("****...****", 32 chars) - the
                # real file always normalizes to "****" + the last 4 real
                # digits, or the literal 8-asterisk placeholder when there are
                # none, never the raw field verbatim.
                obfuscated = card.get("ObfuscatedNumber") or ""
                digit_runs = re.findall(r"\d+", obfuscated)
                if digit_runs:
                    detail = f"{brand} ****{digit_runs[-1][-4:]}"
                elif obfuscated:
                    detail = f"{brand} ********"
                else:
                    detail = brand
                if card.get("Format") == "Virtual":
                    detail += " Virtual"
                if detail:
                    parts.append(detail)
            if pay.get("Identifier"):
                parts.append(str(pay["Identifier"]))
            # Terminal payments often carry a second reference (the acquirer's
            # own trace number) in Notes, which the file appends after the
            # identifier - e.g. "(Mastercard ****2839, 642238, 0985064)".
            if notes:
                parts.append(notes)
            return f"Card {verb} ({', '.join(parts)})" if parts else f"Card {verb}"

        if ptype == "ExternalPayment":
            ext = (pay.get("Data") or {}).get("External") or {}
            label = self._RV_EXTERNAL_LABELS.get(ext.get("Type"), ext.get("Type") or "")
            identifier = ext.get("ExternalIdentifier")
            inner = f"{label} - {identifier}" if identifier else label
            return f"External {verb} ({inner})" if inner else f"External {verb}"

        return ptype or "Payment"

    @staticmethod
    def _rv_display_name(name: str, report_day=None) -> str:
        """The item's BillingName becomes the journal description, and is used
        for grouping too.

        Deliberately not trimmed: several MEWS products are named with a
        trailing space ("Coke ", "Dental Kit ") and the real file preserves it,
        so trimming here would make otherwise-identical rows differ.

        The one thing that IS normalised is the date accommodation carries.
        MEWS formats it per property - Chinatown returns "Night 8/6/2026" but
        Siem Reap returns "Night 06/08/2026" for the very same night - while
        both properties' RV files write it as M/D/YYYY. Rather than guess which
        way round an ambiguous date is, a date is only rewritten when it parses
        (either order) to the day being reported, which is the only date
        accommodation for that day can carry. A date for any OTHER day - a
        rebate reversing an earlier night, say - is left exactly as MEWS wrote
        it, since nothing here can tell 8/9 from 9/8 on a day that isn't this
        one.

        The date is matched wherever it sits, not just at the end: MEWS
        suffixes the line when a charge is amended ("Night 8/21/2026 Rebate
        (Bill 28328)", "Night 8/21/2026 (Originally consumed: 8/21/2026)"),
        and an end-anchored match skipped exactly those rows - Siam's own
        rebate line went out as "Night 21/08/2026 Rebate (Bill 28328)" where
        the real file says "Night 8/21/2026 Rebate (Bill 28328)". Every
        occurrence is rewritten, so the doubled-date form lands right too."""
        name = name or ""
        if report_day is None:
            return name

        def rewrite(match):
            first, second, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
            if year != report_day.year:
                return match.group(0)
            if not ((first, second) == (report_day.month, report_day.day)
                    or (second, first) == (report_day.month, report_day.day)):
                return match.group(0)
            return f"{report_day.month}/{report_day.day}/{report_day.year}"

        return re.sub(r"\b(\d{1,2})/(\d{1,2})/(\d{4})\b", rewrite, name)

    def _rv_gl_overrides(self, property_name: str) -> dict:
        """Per-AccountingCategoryId GL overrides from rv_gl_mappings, keyed by
        category id. Returns {} when the table doesn't exist yet or nothing has
        been mapped - the BillingName defaults above then apply, so RV keeps
        working before anyone has touched Admin (same graceful-degrade shape as
        ftp_service.get_ftp_settings)."""
        if not self.supabase:
            return {}
        try:
            res = self.supabase.table("rv_gl_mappings").select(
                "accounting_category_id, gl_code, department").eq("property", property_name).execute()
        except Exception as e:
            logger.info(f"RV: no rv_gl_mappings overrides for {property_name} ({e})")
            return {}
        return {r["accounting_category_id"]: r for r in (res.data or []) if r.get("accounting_category_id")}

    async def sync_rv_gl_mappings_from_mews(self, property_name: str) -> int:
        """Fetches this property's real chart of accounts from MEWS
        (accountingCategories/getAll - permission enabled 2026-08-17 on all
        8 properties) and upserts it into rv_gl_mappings, keyed by
        AccountingCategoryId - the exact table _rv_gl_overrides reads,
        already checked BEFORE the hand-maintained BillingName chart in
        _rv_revenue_gl. So populating this table needs no other code
        change: every property immediately starts using MEWS's own GL
        codes in place of (or on top of, for an unverified property) the
        BillingName fallback.

        LedgerAccountCode is the real GL account, not Code: confirmed
        identical to Code/ExternalCode/PostingAccountCode for every
        category on Marasca Samui and Koh Tao, but on Chinatown 6 travel-
        agency categories ("Tropic Sun", "Trip Advisor", ...) share one
        generic Code ("9009") while their LedgerAccountCode/
        PostingAccountCode hold each category's real distinct account.
        CostCenterCode is the department - cross-checked against
        Chinatown's own hand-verified chart (Complimentary Services and
        Gifts: 60300/241 in both).

        Every call overwrites the row for a given (property, category id) -
        MEWS's chart of accounts is treated as authoritative on every
        sync, since there's no admin UI yet to hand-edit a row this could
        clobber. Returns the number of categories upserted."""
        all_categories = []
        cursor = None
        while True:
            payload = {"Limitation": {"Count": 1000}}
            if cursor:
                payload["Limitation"]["Cursor"] = cursor
            res = await mews_client.post(
                "/api/connector/v1/accountingCategories/getAll", payload, property_name=property_name)
            chunk = res.get("AccountingCategories") or []
            all_categories.extend(chunk)
            cursor = res.get("Cursor")
            if not cursor or not chunk:
                break

        rows = [
            {
                "property": property_name,
                "accounting_category_id": c["Id"],
                "category_name": c.get("Name") or "",
                "gl_code": c.get("LedgerAccountCode") or c.get("Code") or "",
                "department": c.get("CostCenterCode") or "",
            }
            for c in all_categories
            if c.get("Id") and (c.get("LedgerAccountCode") or c.get("Code"))
        ]
        if not rows:
            return 0
        self.supabase.table("rv_gl_mappings").upsert(
            rows, on_conflict="property,accounting_category_id").execute()
        return len(rows)

    @staticmethod
    def _rv_is_service_charge(low: str) -> bool:
        """MEWS names service-charge lines either "Service Charge [Room]" or
        "<product> SVC", e.g. "Included Breakfast SVC (Adults)". SVC is matched
        as a whole word so a product that merely contains those letters isn't
        swept up."""
        return "service charge" in low or re.search(r"\bsvc\b", low) is not None

    def _rv_chart(self, property_name: str) -> dict:
        """The property's chart of accounts, or {} if it has never been
        verified against that property's own RV file."""
        return self._RV_CHARTS.get(property_name) or {}

    def _rv_revenue_gl(self, billing_name: str, category_id: str, overrides: dict, chart: dict) -> tuple:
        override = overrides.get(category_id or "")
        if override and override.get("gl_code"):
            return override["gl_code"], override.get("department") or ""
        low = (billing_name or "").lower()
        base = next(((gl, dept) for keyword, gl, dept in chart.get("revenue", []) if keyword in low), None)
        service_charge = chart.get("service_charge")
        # Most properties book a product's own service charge to the separate
        # service-charge account (breakfast SVC -> 30805 at Chinatown/Siam).
        # Makati, Patong and Koh Samui do the opposite for breakfast - "Included
        # Breakfast SVC (Adults)" and its VAT stay on breakfast's own account
        # (11416), never splitting out - confirmed against every SVC/VAT line
        # in their files. service_charge_exempt_gl lists any base account a
        # property keeps intact like this.
        exempt = base and base[0] in chart.get("service_charge_exempt_gl", ())
        if service_charge and self._rv_is_service_charge(low) and not exempt:
            # Service charge on a product keeps that product's department
            # (breakfast SVC -> 121); a bare room service charge has no base
            # product and falls back to accommodation's own department.
            return service_charge[0], (base[1] if base else service_charge[1])
        if base:
            return base
        return chart.get("fallback") or ("", "")

    async def get_rv_report(self, property_name: str, date: str) -> dict:
        """
        Builds the daily RV (Revenue) report for one property + one calendar
        date in that property's own MEWS timezone - the revenue/payment
        counterpart to get_st_files_report.

        Two MEWS calls, both filtered to the property's own calendar day:
          * orderItems/getAll over ConsumedUtc -> the revenue lines
          * payments/getAll   over CreatedUtc  -> the settlement lines

        The one non-obvious rule, and the one that had this off by 2.4x until
        it was checked against a real RV file: order items whose
        AccountingState is "Canceled" must be dropped. MEWS keeps the
        superseded posting alongside its replacement (on 07-Aug Chinatown that
        was 95 of 224 accommodation items, 197,478.21 of phantom revenue), and
        the RV file counts only the survivors. Payments have no such twins -
        every payment MEWS returned for that day was live.

        VAT is summed from each surviving item's Amount.TaxValues rather than
        derived from a rate, because MEWS already splits mixed-rate items for
        us. Amounts are stored positive with a separate "dc" (Debit/Credit)
        flag, matching how the Infor file itself represents sign.
        """
        property_tz = await self._resolve_property_timezone(property_name)
        day = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=property_tz)
        start_iso = day.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        end_iso = (day + timedelta(days=1)).astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

        all_items, all_payments, all_outlet_items, services_res = await asyncio.gather(
            self._rv_fetch_paged(
                property_name, "/api/connector/v1/orderItems/getAll", "OrderItems",
                {"ConsumedUtc": {"StartUtc": start_iso, "EndUtc": end_iso}}),
            self._rv_fetch_paged(
                property_name, "/api/connector/v1/payments/getAll", "Payments",
                {"CreatedUtc": {"StartUtc": start_iso, "EndUtc": end_iso}}),
            # Outlet (POS) tills are a SEPARATE ledger in MEWS and never
            # appear in orderItems/payments - see _rv_outlet_rows.
            self._rv_fetch_paged(
                property_name, "/api/connector/v1/outletItems/getAll", "OutletItems",
                {"ConsumedUtc": {"StartUtc": start_iso, "EndUtc": end_iso}}),
            mews_client.post(
                "/api/connector/v1/services/getAll",
                {"Limitation": {"Count": 100}},
                property_name=property_name,
            ),
        )

        overrides = self._rv_gl_overrides(property_name)
        chart = self._rv_chart(property_name)
        stay = self._resolve_stay_service(services_res.get("Services", []))

        # --- Revenue -------------------------------------------------------
        live_items = [i for i in all_items
                      if i.get("AccountingState") != "Canceled"]
        skipped_canceled = len(all_items) - len(live_items)
        segments = await self._rv_market_segments(
            property_name, live_items, stay.get("Id") if stay else None)

        # A property's TaxValues array isn't always "just VAT". Makati stacks a
        # duplicate service-charge component onto every item under the same
        # array (PH-MA-SERVICE), which would double the real 12% VAT if summed
        # blindly; vat_tax_codes, when the chart sets it, restricts the primary
        # bucket to the codes that are genuinely output tax. secondary_tax
        # captures a second, truly separate levy several properties report
        # under its own TaxRateCode (Thailand's 1% provincial tax) - real
        # money, its own account, not a duplicate of anything.
        vat_codes = chart.get("vat_tax_codes")
        secondary = chart.get("secondary_tax")
        secondary_codes = (secondary or {}).get("codes") or set()
        # Makati has no real order item for its service charge at all - see
        # the chart comment - so its total is accumulated from this one
        # TaxRateCode instead and turned into a synthetic revenue row below.
        sc_tax_code = chart.get("service_charge_tax_code")
        sc_tax_total = 0.0

        revenue, vat_total, secondary_tax_total = {}, 0.0, 0.0
        vat_by_segment = Counter()
        currencies = Counter()
        for item in live_items:
            amount = item.get("Amount") or {}
            if amount.get("Currency"):
                currencies[amount["Currency"]] += 1
            net = amount.get("NetValue") or 0.0
            name = self._rv_display_name(item.get("BillingName") or "", day)
            gl, dept = self._rv_revenue_gl(name, item.get("AccountingCategoryId"), overrides, chart)
            segment = segments.get(item.get("ServiceOrderId"), "") or ""
            key = (gl, dept, name, segment)
            row = revenue.setdefault(key, {"gl_code": gl, "department": dept, "name": name,
                                           "market_segment": segment, "amount": 0.0, "count": 0})
            row["amount"] += net
            row["count"] += 1
            for tax in (amount.get("TaxValues") or []):
                code, value = tax.get("Code"), tax.get("Value") or 0.0
                if sc_tax_code and code == sc_tax_code:
                    sc_tax_total += value
                elif code in secondary_codes:
                    secondary_tax_total += value
                elif vat_codes is None or code in vat_codes:
                    vat_total += value
                    vat_by_segment[segment] += value

        if sc_tax_code and round(sc_tax_total, 2):
            sc_gl, sc_dept = chart.get("service_charge") or ("", "")
            key = (sc_gl, sc_dept, "Service charge", "")
            row = revenue.setdefault(key, {"gl_code": sc_gl, "department": sc_dept, "name": "Service charge",
                                           "market_segment": "", "amount": 0.0, "count": 0,
                                           # Not a real order item - the whole
                                           # row is summed from a TaxValues
                                           # component, like VAT itself, and
                                           # the file marks it the same way.
                                           "tax_derived": True})
            row["amount"] += sc_tax_total
            row["count"] += 1

        # --- Outlet (POS) tills --------------------------------------------
        # Restaurant/bar/retail sales rung up on a MEWS Outlet are a separate
        # ledger: they are NOT order items and their settlements are NOT
        # payments, so both feeds above return nothing for them. Two of our
        # properties run real outlets and were silently short a whole
        # department because of it - Koh Tao books ALL its food and beverage
        # this way (30110/30115/30210/30215/30550 came out as a flat 0.00
        # against 22,793/30,834/2,767/13,052/990 in its own RV file), and
        # Marasca splits F&B across both ledgers (its Food Cabanas - Lunch is
        # 10,373.77 of order items plus 25,488.39 of outlet items, which is
        # the 35,862.16 its file reports).
        #
        # Outlet lines carry no market segment and no TaxValues - an outlet's
        # tax is rung up as its own line ("TAX", "VAT") against its own
        # accounting category - so they never feed vat_total and never take
        # the ABBSU/V07 stamp. GL and department come from the accounting
        # category alone, the same accountingCategories chart the overrides
        # are synced from.
        outlet_payments = []
        for item in all_outlet_items:
            amount = item.get("UnitAmount") or {}
            if amount.get("Currency"):
                currencies[amount["Currency"]] += 1
            net = (amount.get("NetValue") or 0.0) * (item.get("UnitCount") or 1)
            name = item.get("Name") or ""
            category = overrides.get(item.get("AccountingCategoryId") or "") or {}
            gl = category.get("gl_code") or ""
            dept = category.get("department") or ""
            if item.get("Type") == "Payment":
                # Money in, so MEWS signs it negative exactly like a normal
                # payment; the row below flips it back to a magnitude.
                outlet_payments.append({
                    "gl_code": gl, "department": dept, "name": name,
                    "amount": -net, "count": 1, "unmapped": not gl,
                })
                continue
            # Revenue and NonRevenue alike: discounts arrive as NonRevenue
            # negatives and service charge/tax as NonRevenue positives, and
            # both belong on the revenue side of the journal against their own
            # account, which is where the file puts them.
            key = (gl, dept, name, "")
            row = revenue.setdefault(key, {"gl_code": gl, "department": dept, "name": name,
                                           "market_segment": "", "amount": 0.0, "count": 0})
            row["amount"] += net
            row["count"] += 1

        # --- Payments ------------------------------------------------------
        # One journal row per transaction, not one per GL account. The real
        # file does this and it can't be recovered later: the SunSystems PMSRV
        # import profile runs with Consolidation "Not Specified", so whatever
        # granularity arrives in the file is exactly what lands in the ledger.
        # Aggregating here would permanently destroy the per-payment detail
        # finance reconciles against (card, prepayment and voucher numbers all
        # live in the description).
        #
        # MEWS returns payments as negative (they credit the guest's ledger);
        # rows carry the magnitude and the sign moves into the D/C flag.
        # A payment shared across a group booking (e.g. one prepayment
        # covering several linked reservations) can come back from MEWS as
        # the same Payment resource more than once - confirmed against a
        # Makati AGODA group booking where one payment appeared 19 times
        # instead of once. Dedup by Id so it's counted exactly once.
        seen_payment_ids = set()
        live_payments = []
        for pay in all_payments:
            if pay.get("State") == "Canceled":
                continue
            # GhostPayment was assumed to be MEWS's own internal balancing
            # placeholder and excluded on that theory. The real file proves
            # that wrong: Chinatown, 22-Aug-2026, carries a matched
            # +2,403.03/-2,403.03 GhostPayment pair (a failed charge and its
            # retry) as two ordinary journal rows. Kept in now - a Ghost pair
            # still nets to zero on any day that has one, so this can only
            # add the rows the file already expects, never move a balance.
            if not ((pay.get("Amount") or {}).get("NetValue") or 0.0):
                continue
            pay_id = pay.get("Id")
            if pay_id and pay_id in seen_payment_ids:
                continue
            if pay_id:
                seen_payment_ids.add(pay_id)
            live_payments.append(pay)

        # A GhostPayment carries no card/brand/reference of its own - Data.Ghost
        # only points at the real payment it's tracking via OriginalPaymentId,
        # and MEWS's file describes the Ghost row exactly as that original
        # ("Card payment (Mastercard ****8700 Virtual, 459909)" for a Ghost
        # whose original was that Mastercard charge). Resolve it once so the
        # key/description builders below can treat a Ghost as its original's
        # Type+Data+Identifier while still using the GHOST's own Amount/Notes
        # - the two must not be conflated, or the row's own D/C and any
        # front-desk note on the correction itself would be lost.
        ghost_original_ids = {
            ((p.get("Data") or {}).get("Ghost") or {}).get("OriginalPaymentId")
            for p in live_payments if p.get("Type") == "GhostPayment"
        }
        ghost_original_ids.discard(None)
        ghost_originals = {}
        if ghost_original_ids:
            orig_res = await mews_client.post(
                "/api/connector/v1/payments/getAll",
                {"PaymentIds": list(ghost_original_ids), "Limitation": {"Count": len(ghost_original_ids)}},
                property_name=property_name,
            )
            ghost_originals = {o["Id"]: o for o in orig_res.get("Payments", [])}

        def effective_pay(pay: dict) -> dict:
            if pay.get("Type") != "GhostPayment":
                return pay
            original = ghost_originals.get(((pay.get("Data") or {}).get("Ghost") or {}).get("OriginalPaymentId"))
            if not original:
                return pay
            return {**original, "Amount": pay.get("Amount"), "Notes": pay.get("Notes")}

        cards = await self._rv_load_credit_cards(property_name, {
            ((effective_pay(p).get("Data") or {}).get("CreditCard") or {}).get("CreditCardId")
            for p in live_payments if effective_pay(p).get("Type") == "CreditCardPayment"
        })

        payment_rows = []
        for pay in live_payments:
            net = (pay.get("Amount") or {}).get("NetValue") or 0.0
            eff = effective_pay(pay)
            key = self._rv_payment_key(eff)
            mapped = (chart.get("payments") or {}).get(key)
            # An unrecognised payment type must NOT be quietly booked to some
            # real GL account - it's left unmapped so the UI can flag it and
            # somebody decides where it belongs.
            gl, dept = mapped if mapped else ("", "")
            payment_rows.append({
                "gl_code": gl,
                "department": dept,
                "name": self._rv_payment_description(
                    eff, cards, property_name, force_payment_verb=pay.get("Type") == "GhostPayment"
                ) if mapped else (key or "Unknown payment type"),
                "amount": -net,            # flip to positive money-in
                "count": 1,
                "unmapped": not mapped,
            })

        # Appended rather than merged: an outlet till row is already one line
        # per transaction. Zero-value ones are kept deliberately - Koh Tao
        # rings up 41 "ROOM PACKAGE" settlements at 0.00 a day (a package
        # already paid for through the room) and its RV file carries every
        # one of them, unlike the MEWS-payment loop above where a zero is
        # noise.
        payment_rows.extend(outlet_payments)

        # A group whose postings cancel out (a rebate against its own original
        # within the same billing name) nets to zero. MEWS omits those rows and
        # so do we - a zero-value journal line carries no information and would
        # just be noise in the ledger.
        revenue_rows = sorted((r for r in revenue.values() if round(r["amount"], 2)),
                              key=lambda r: (r["gl_code"], r["name"], r.get("market_segment") or ""))
        payment_rows.sort(key=lambda r: (r["gl_code"], r["name"]))
        revenue_net = sum(r["amount"] for r in revenue_rows)
        payments_total = sum(r["amount"] for r in payment_rows)
        total_tax = vat_total + secondary_tax_total

        return {
            "property": property_name,
            "date": date,
            "revenue": revenue_rows,
            "payments": payment_rows,
            "vat": round(vat_total, 2),
            # Per-market-segment VAT breakdown - only Siam's file actually
            # splits the VAT line this way (see vat_by_segment on the chart);
            # every other property posts the total above as one row and
            # ignores this. Rounds away the blank-segment bucket only when
            # it's genuinely zero, same rule as the revenue rows.
            "vat_by_segment": {seg: round(amt, 2) for seg, amt in vat_by_segment.items() if round(amt, 2)},
            # Field 10 of the journal. Siem Reap posts in USD, not THB, so this
            # cannot be assumed - it is the currency MEWS actually stamped on
            # the day's items (the most common one, if a day ever mixes them).
            "currency": (currencies.most_common(1)[0][0] if currencies else "THB"),
            "vat_gl_code": (chart.get("vat") or ("", ""))[0],
            # A second, genuinely separate tax several properties report under
            # its own TaxRateCode (Thailand's 1% provincial tax) - not present
            # for properties whose chart doesn't define secondary_tax.
            "secondary_tax": round(secondary_tax_total, 2) if secondary else None,
            "secondary_tax_gl_code": (secondary or {}).get("gl_code", ""),
            "secondary_tax_label": (secondary or {}).get("label", ""),
            # Guest Ledger is the balancing row: whatever the day earned but
            # didn't settle is still owed on somebody's open bill.
            "guest_ledger": round(revenue_net + total_tax - payments_total, 2),
            "guest_ledger_gl_code": (chart.get("guest_ledger") or ("", ""))[0],
            "totals": {
                "revenue_net": round(revenue_net, 2),
                "vat": round(vat_total, 2),
                # Includes the secondary tax (where the property has one) -
                # this is the true amount the guest was actually charged.
                "revenue_gross": round(revenue_net + total_tax, 2),
                "payments": round(payments_total, 2),
            },
            "counts": {
                "revenue_items": sum(r["count"] for r in revenue_rows),
                "payment_items": sum(r["count"] for r in payment_rows),
                "canceled_items_skipped": skipped_canceled,
            },
            # Surfaced in the UI so nobody has to guess why GL codes look
            # generic: until MEWS enables Accounting Categories, revenue GL
            # codes come from BillingName matching, not from MEWS itself.
            "gl_source": "mews_categories" if overrides else "billing_name_defaults",
            # False means the GL codes shown are Chinatown's, applied to a
            # property that has never been checked against its own RV file -
            # the report is still useful to read, but get_rv_export refuses to
            # produce a financial file from it.
            "gl_verified": bool(overrides) or bool(chart),
        }

    async def get_rv_list(self, property_name: str) -> list:
        """RV List's per-day summary rows - reads exclusively from
        rv_files_sync (recomputing history live would mean 2 MEWS calls per
        row for what could be months), each row's totals taken from that
        day's already-stored blob rather than recomputed here."""
        if not self.supabase:
            return []
        res = self.supabase.table("rv_files_sync") \
            .select("report_date, data, synced_at") \
            .eq("property", property_name) \
            .order("report_date", desc=True) \
            .execute()
        rows = []
        for row in res.data or []:
            blob = (row.get("data") or {}).get("blob", "")
            if not blob:
                continue
            try:
                report = json.loads(encryption_service.decrypt(blob))
            except Exception as e:
                logger.warning(f"RV List: failed to decrypt {property_name}/{row.get('report_date')}: {e}")
                continue
            totals = report.get("totals") or {}
            rows.append({
                "date": row.get("report_date"),
                "revenue_net": totals.get("revenue_net", 0),
                "vat": totals.get("vat", 0),
                "revenue_gross": totals.get("revenue_gross", 0),
                "payments": totals.get("payments", 0),
                "guest_ledger": report.get("guest_ledger", 0),
                "synced_at": row.get("synced_at"),
            })
        return rows

    def get_rv_export(self, property_name: str, date_str: str) -> tuple:
        """
        Builds the pipe-delimited Infor "RV" revenue journal for one
        already-imported day - the same 41-field layout as the ST file
        (see get_st_report_export) with journal type PMSRV in fields 1-2 and
        an "RV{yyyymmdd}" reference in field 7.

        Layout confirmed field-by-field against a real MEWS-generated file
        (MS_RV_20260807.csv):
          PMSRV|PMSRV|{gl}|{ddmmyyyy}|0{mm}|{yyyy}|RV{yyyymmdd}|{description}
            |{ddmmyyyy}|THB|{amount}|1|{amount}|||{D|C}|{code}|{dept}|ZZ|ZZ|...
        Two differences from the ST file worth noting: the amount appears
        again in field 13 (blank in ST), and field 18 is a real per-line
        department code (always "111" in ST).

        Sign convention, taken from the file itself rather than assumed:
        revenue and VAT are Credits, payments and the Guest Ledger balance
        are Debits, and anything negative (rebates, refunds) flips to the
        opposite letter and is written as a positive magnitude.
        """
        if not self.supabase:
            raise Exception("Supabase not initialized")
        prop_res = self.supabase.table("property_api_settings").select("st_property_code").eq(
            "property_name", property_name).limit(1).execute()
        property_code = (prop_res.data[0].get("st_property_code") if prop_res.data else None)
        if not property_code:
            raise ValueError(f"No Property Code configured yet for {property_name} - set it in Admin > API")
        chart = self._rv_chart(property_name)
        if not chart and not self._rv_gl_overrides(property_name):
            raise ValueError(
                f"Cannot export: no chart of accounts has been verified for {property_name}. "
                "Each property posts to its own accounts - Siem Reap books Guest Ledger to "
                "11401 where Chinatown uses 21203 - so no other property's codes can stand "
                "in. Run POST /rv/gl-mappings/sync?property_name=... to pull this property's "
                "real chart of accounts from MEWS (accountingCategories/getAll), or add its "
                "GL mapping to rv_gl_mappings by hand. Viewing the report on screen still "
                "works.")
        res = self.supabase.table("rv_files_sync").select("data").eq(
            "property", property_name).eq("report_date", date_str).limit(1).execute()
        if not res.data:
            raise ValueError(f"No imported RV report for {property_name} on {date_str} - import it first")
        report = json.loads(encryption_service.decrypt((res.data[0].get("data") or {}).get("blob", "")))

        unmapped = [p for p in report.get("payments", []) if p.get("unmapped")]
        if unmapped:
            raise ValueError(
                "Cannot export: unmapped payment type(s) " +
                ", ".join(p.get("name", "?") for p in unmapped) +
                " have no GL account. Map them before exporting.")

        day = datetime.strptime(date_str, "%Y-%m-%d")
        ddmmyyyy = day.strftime("%d%m%Y")
        yyyymmdd = day.strftime("%Y%m%d")
        month_code = "0" + day.strftime("%m")
        currency = report.get("currency") or "THB"

        def row(gl, description, amount, natural_dc, dept, market_segment="", tax_derived=False):
            """natural_dc is the letter this kind of line carries when its
            amount is positive; a negative amount flips it and is written as
            a magnitude.

            Field 8 is capped at 50 characters because that is the declared
            width of TransDesc in Infor's interface spec, and MEWS's own file
            truncates to exactly that. It matters more than it looks: the
            SunSystems import profile posts with "Post if no errors", so one
            over-length line rejects the entire day's journal, not just its
            own row. A raw newline embedded in free-text guest/payment notes
            (seen live: a Notes field of literally "EXTEND\\n") is just as
            fatal - split into "|".join(lines) it breaks one logical row into
            two ragged ones - so each newline/pipe becomes a space; a
            literal "|" would misalign every field after it the same way, so
            that's stripped too. The same all-or-nothing stake applies to
            non-ASCII: see _ascii_fold."""
            # Each separator maps to its OWN space, not a collapsed run - Koh
            # Samui's own Notes field is literally "...C/S\n\n" (two
            # newlines), and the real file's description ends "...C/S  "
            # (two trailing spaces). A run-collapsing regex was tried first
            # and silently ate the second character; every multi-separator
            # case checked against the file confirms 1:1, never N:1.
            # Not .strip() either - a few real product names carry a
            # deliberate trailing space, e.g. "Coke ", that the real file
            # preserves and this must not eat.
            desc_clean = _ascii_fold(re.sub(r"[\r\n|]", " ", description or ""))
            dc = natural_dc if amount >= 0 else ("D" if natural_dc == "C" else "C")
            value = f"{abs(amount):.2f}"
            # Fields 21-41 are blank except Analysis 6 (field 22), which the
            # spec designates MARKET SEGMENT, and - on lines whose amount was
            # summed from TaxValues rather than posted as its own order item
            # (VAT, the secondary/local tax, and Makati's synthetic Service
            # Charge row) - Analysis 1/5 (fields 21/25), which the real file
            # always stamps "ABBSU"/"V07" on those lines and only those.
            tail = [""] * 21
            tail[1] = market_segment or ""
            if tax_derived:
                tail[0] = "ABBSU"
                tail[4] = "V07"
            # Folded once more over the assembled line so the guarantee covers
            # every field, not just the description - Analysis 6 (market
            # segment) is free text from MEWS too, and codes/currency come
            # from config that nothing else validates.
            return _ascii_fold("|".join([
                "PMSRV", "PMSRV", gl, ddmmyyyy, month_code, str(day.year),
                f"RV{yyyymmdd}", desc_clean[:50], ddmmyyyy, currency, value, "1", value,
                "", "", dc, property_code, dept or "", "ZZ", "ZZ",
            ] + tail))

        lines = []
        for r in report.get("revenue", []):
            lines.append(row(r["gl_code"], r["name"], r["amount"], "C", r.get("department"),
                             r.get("market_segment"), tax_derived=bool(r.get("tax_derived"))))
        vat_by_segment = report.get("vat_by_segment")
        if chart.get("vat_by_segment") and vat_by_segment:
            vat_gl = chart.get("vat") or ("", "")
            for segment, amount in vat_by_segment.items():
                if round(amount, 2):
                    lines.append(row(vat_gl[0], "VAT", amount, "C", vat_gl[1], segment, tax_derived=True))
        elif report.get("vat"):
            vat_gl = chart.get("vat") or ("", "")
            lines.append(row(vat_gl[0], "VAT", report["vat"], "C", vat_gl[1], tax_derived=True))
        if report.get("secondary_tax"):
            lines.append(row(report["secondary_tax_gl_code"], report["secondary_tax_label"] or "VAT",
                             report["secondary_tax"], "C", "", tax_derived=True))
        for p in report.get("payments", []):
            lines.append(row(p["gl_code"], p["name"], p["amount"], "D", p.get("department")))
        if report.get("guest_ledger"):
            ledger_gl = chart.get("guest_ledger") or ("", "")
            lines.append(row(ledger_gl[0], "Guest Ledger",
                             report["guest_ledger"], "D", ledger_gl[1]))

        filename = f"{property_code}_RV_{yyyymmdd}.csv"
        return "\n".join(lines), filename

    def _build_st_property_email(self, prop: str, date_display: str,
                                  row: dict, per_property_settings: dict) -> tuple:
        """Subject/body builder for one property's individual ST Files
        email, split out of send_st_files_property_email for readability."""
        subject = per_property_settings["subject"] \
            .replace("<<Property>>", prop) \
            .replace("<<PropertyCode>>", row["property_code"]) \
            .replace("<<Date>>", date_display)
        html_body = per_property_settings["html_template"] \
            .replace("<<Property>>", prop) \
            .replace("<<PropertyCode>>", row["property_code"]) \
            .replace("<<Date>>", date_display) \
            .replace("<<StatsTable>>", self._build_st_files_summary_table(
                [{"property_name": prop, "property_code": row["property_code"], "totals": row["totals"]}]))
        return subject, html_body

    async def send_st_files_property_email(self, property_name: str, date_str: str,
                                            mark_sent: bool = True, sent_date_str: str = None,
                                            sync_type: str = "auto") -> dict:
        """
        Builds and sends ONE property's own ST Files email (Admin >
        Templates > ST Files Email (Per-Property)) - used by main.py's
        send_st_files_per_property_emails, which checks each opted-in
        property's own st_files_email_hour/_minute independently (Admin >
        Templates' Per-Property panel) rather than one shared clock.

        No st_files_email_recipients (To) configured is a skip, not fatal -
        same as a missing Property Code or that day's report not being
        imported yet. mark_sent writes THIS property's own
        st_files_email_last_sent_date (not the bundled st_files_daily row's
        last_sent_date, which this path never touches) - same
        sent_date_str-vs-date_str separation send_st_files_bundled_digest
        uses, for the same reason (dedup marker is keyed on the send day,
        the report itself is yesterday's).

        Every outcome (sent or skipped) is logged to sync_logs under
        target_table="ST Files Email (Per-Property)", same table/pattern
        _log_sync_row already uses for ST Files FTP uploads - this is what
        powers the History section on /st-files. sync_type defaults to
        "auto" for the real per-property scheduler; its own manual "Send
        Test Now" (admin.py's send_st_files_per_property_email_now) passes
        "manual" instead.

        Subject/HTML are per-property (property_api_settings.
        st_files_email_subject/_template, edited on the same Admin >
        Templates > Statistic Files > Per-Property panel as the recipients
        below) - null falls back to the built-in
        DEFAULT_ST_FILES_DAILY_PER_PROPERTY_SUBJECT/TEMPLATE, same pattern
        every other template in this app uses. Each property can now have
        a fully different subject/body, not just a shared one with tokens.
        """
        p_res = self.supabase.table("property_api_settings").select(
            "id, st_files_email_recipients, st_files_email_cc, st_files_email_bcc, "
            "st_files_email_subject, st_files_email_template"
        ).eq("property_name", property_name).limit(1).execute()
        p = p_res.data[0] if p_res.data else {}
        prop_id = p.get("id")
        per_property_settings = {
            "subject": p.get("st_files_email_subject") or DEFAULT_ST_FILES_DAILY_PER_PROPERTY_SUBJECT,
            "html_template": p.get("st_files_email_template") or DEFAULT_ST_FILES_DAILY_PER_PROPERTY_TEMPLATE,
        }
        date_display = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d/%m/%Y")

        try:
            text, filename = self.get_st_report_export(property_name, date_str)
            row = self._get_st_report_row(property_name, date_str)
        except Exception as e:
            reason = f"{property_name}: {str(e)[:150]}"
            self._log_sync_row(property_name, prop_id, "ST Files Email (Per-Property)", "error", 0, reason, sync_type)
            return {"sent": False, "skipped": reason}

        recipients = [e.strip() for e in (p.get("st_files_email_recipients") or "").split(",") if e.strip()]
        if not recipients:
            reason = f"{property_name}: no ST Files Email Recipients configured (Admin > Templates > Statistic Files > Per-Property)"
            self._log_sync_row(property_name, prop_id, "ST Files Email (Per-Property)", "error", 0, reason, sync_type)
            return {"sent": False, "skipped": reason}
        cc = [e.strip() for e in (p.get("st_files_email_cc") or "").split(",") if e.strip()]
        bcc = [e.strip() for e in (p.get("st_files_email_bcc") or "").split(",") if e.strip()]

        subject, html_body = self._build_st_property_email(property_name, date_display, row, per_property_settings)
        email_service.send_email_with_attachments(
            recipients, subject, html_body, [(filename, text.encode("utf-8"))],
            cc_emails=cc, bcc_emails=bcc)
        self._log_sync_row(property_name, prop_id, "ST Files Email (Per-Property)", "success", 1,
                            f"Email sent to {', '.join(recipients)}", sync_type)

        if mark_sent:
            try:
                self.supabase.table("property_api_settings").update(
                    {"st_files_email_last_sent_date": sent_date_str or date_str}
                ).eq("property_name", property_name).execute()
            except Exception as e:
                logger.warning(f"ST Files per-property email: failed to record last_sent_date for {property_name}: {e}")

        return {"sent": True, "skipped": None}

    async def send_st_files_bundled_digest(self, date_str: str, mark_sent: bool = True,
                                            sent_date_str: str = None, sync_type: str = "auto") -> dict:
        """
        The bundled ST Files email (Admin > Templates > ST Files Email) -
        one CSV attachment per EVERY property that has a Property Code
        configured and already-imported st_files_sync data for date_str.
        This is a standing master copy independent of each property's own
        st_files_email_enabled opt-in (see send_st_files_property_email for
        that separate path) - a property having its own per-property email
        does not exclude it here, by design, so this recipient list always
        gets every property regardless of who else is separately opted in.
        Used by main.py's send_st_files_daily_email, gated by this row's own
        shared send_hour/send_minute - independent of whatever individual
        times any per-property-opted-in properties are using.

        sent_date_str/date_str/sync_type semantics match
        send_st_files_property_email - see its own docstring. Logs one
        sync_logs row per property (success or error), target_table="ST
        Files Email" - a single combined email still gets one history row
        per participating property so each property's own History section
        on /st-files shows it, same as the per-property path's own rows.
        """
        settings_row = email_service.get_st_files_daily_settings()
        props_res = self.supabase.table("property_api_settings").select(
            "id, property_name"
        ).order("property_name").execute()
        date_display = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d/%m/%Y")

        attachments, included, skipped, table_rows = [], [], [], []
        prop_ids = {}
        for p in (props_res.data or []):
            prop = p["property_name"]
            prop_ids[prop] = p.get("id")
            try:
                text, filename = self.get_st_report_export(prop, date_str)
                attachments.append((filename, text.encode("utf-8")))
                included.append(prop)
                row = self._get_st_report_row(prop, date_str)
                table_rows.append({"property_name": prop, "property_code": row["property_code"], "totals": row["totals"]})
            except Exception as e:
                reason = f"{prop}: {str(e)[:150]}"
                skipped.append(reason)
                self._log_sync_row(prop, prop_ids.get(prop), "ST Files Email", "error", 0, reason, sync_type)

        if not attachments:
            return {"sent": False, "included": included, "skipped": skipped}

        recipients = [e.strip() for e in (settings_row["recipients"] or "").split(",") if e.strip()]
        if not recipients:
            for prop in included:
                self._log_sync_row(prop, prop_ids.get(prop), "ST Files Email", "error", 0,
                                    "Bundled ST Files Email has no recipients configured (Admin > Templates)", sync_type)
            return {"sent": False, "included": [], "skipped": skipped + [f"{p}: bundled email has no recipients configured" for p in included]}

        subject = settings_row["subject"].replace("<<Date>>", date_display)
        html_body = settings_row["html_template"] \
            .replace("<<Date>>", date_display) \
            .replace("<<PropertyCount>>", str(len(included))) \
            .replace("<<PropertyList>>", ", ".join(included)) \
            .replace("<<StatsTable>>", self._build_st_files_summary_table(table_rows))

        email_service.send_email_with_attachments(recipients, subject, html_body, attachments)
        for prop in included:
            self._log_sync_row(prop, prop_ids.get(prop), "ST Files Email", "success", 1,
                                f"Bundled email sent to {', '.join(recipients)}", sync_type)

        if mark_sent:
            self._mark_st_files_daily_sent(settings_row, sent_date_str or date_str)

        return {"sent": True, "included": included, "skipped": skipped}

    def _mark_st_files_daily_sent(self, settings_row: dict, marker_date: str):
        """
        Same-day dedup guard for send_st_files_bundled_digest - writes
        marker_date onto the single st_files_daily row's last_sent_date.
        """
        try:
            existing = self.supabase.table("email_templates").select("id") \
                .eq("template_key", ST_FILES_DAILY_TEMPLATE_KEY).limit(1).execute()
            if existing.data:
                self.supabase.table("email_templates").update({"last_sent_date": marker_date}) \
                    .eq("id", existing.data[0]["id"]).execute()
            else:
                self.supabase.table("email_templates").insert({
                    "template_key": ST_FILES_DAILY_TEMPLATE_KEY,
                    "subject": settings_row["subject"],
                    "html_template": settings_row["html_template"],
                    "recipients": settings_row["recipients"],
                    "send_hour": settings_row["send_hour"],
                    "send_minute": settings_row["send_minute"],
                    "enabled": True,
                    "last_sent_date": marker_date,
                }).execute()
        except Exception as e:
            logger.warning(f"ST Files daily email: failed to record last_sent_date: {e}")

    # ------------------------------------------------------------- RR4/TM30 email
    #
    # Thai Hotel Act filings only - Lub d Siem Reap (Cambodia) and Lub d
    # Philippines Makati don't file RR4/TM30 at all, so both the bundled
    # digest and the per-property email skip them outright rather than
    # sending a report neither property actually needs.
    _RR4_TM30_EMAIL_EXCLUDED_PROPERTIES = {"Lub d Siem Reap", "Lub d Philippines Makati"}

    def _read_rr4_tm30_cached_day(self, property_name: str, date_str: str) -> Optional[dict]:
        """Decrypts one already-imported (property, date) row out of
        rr4_tm30_sync - same table/blob shape api/app/routers/rr4.py's own
        read_managed_day reads, duplicated here (a router importing into a
        service is the wrong direction) so send_rr4_tm30_property_email/
        send_rr4_tm30_bundled_digest can render straight from whatever the
        daily import job already fetched, with no live MEWS call of their
        own - same reasoning get_st_report_export has for reading
        st_files_sync instead of re-fetching live. Returns None if that day
        hasn't been imported yet; callers treat that as "not ready", same as
        ST Files treats a missing st_files_sync row."""
        res = self.supabase.table("rr4_tm30_sync").select("data").eq(
            "property", property_name).eq("report_date", date_str).limit(1).execute()
        if not res.data:
            return None
        blob = (res.data[0].get("data") or {}).get("blob", "")
        payload = json.loads(encryption_service.decrypt(blob))
        # The stored blob is MEWS's raw answer (sync_rr4_tm30_day opts out of
        # overrides on purpose), so the manual corrections have to be laid
        # over it here - otherwise the emailed attachment would be the only
        # copy of the register that still shows the uncorrected data.
        for kind in ("rr4", "tm30"):
            if isinstance(payload.get(kind), dict):
                self.apply_rr4_tm30_overrides(payload[kind], property_name, date_str, kind)
        return payload

    def _resolve_property_code(self, property_name: str) -> str:
        """One-off st_property_code lookup, used by the RR4/TM30 email
        functions so both attachments for one property share a single query
        instead of the two _render_rr4_export/_render_tm30_export would
        otherwise each run when called through get_rr4_export/
        get_tm30_export directly."""
        property_code = property_name
        if self.supabase:
            try:
                prop_res = self.supabase.table("property_api_settings").select("st_property_code").eq(
                    "property_name", property_name).limit(1).execute()
                property_code = (prop_res.data[0].get("st_property_code") if prop_res.data else None) or property_name
            except Exception:
                pass
        return property_code

    def _build_rr4_tm30_summary_table(self, rows: list) -> str:
        """
        Pre-built HTML for the <<StatsTable>> token in the RR4/TM30 daily
        email (Admin > Templates > RR4/TM30 Files) - same reasoning as
        _build_st_files_summary_table. rows: list of {"property_name",
        "property_code", "rr4_count", "tm30_count"} in display order.
        """
        if not rows:
            return '<p style="margin:0; font-size:12px; color:#152A00; opacity:0.6;">No properties included.</p>'
        body_rows = []
        for i, row in enumerate(rows):
            bg = "#ffffff" if i % 2 == 0 else "#FFEFD2"
            body_rows.append(
                f'<tr style="background:{bg}; border-bottom:1px solid rgba(21,42,0,0.08);">'
                f'<td style="padding:7px 10px; text-align:left; font-size:12px; color:#152A00; '
                f'font-weight:700; white-space:nowrap;">{_escape_html(row["property_name"])}</td>'
                f'<td style="padding:7px 6px; text-align:center; font-size:12px; color:#152A00; '
                f'font-variant-numeric:tabular-nums;">{_escape_html(row["property_code"])}</td>'
                f'<td style="padding:7px 6px; text-align:center; font-size:12px; color:#152A00; '
                f'font-variant-numeric:tabular-nums;">{row["rr4_count"]}</td>'
                f'<td style="padding:7px 6px; text-align:center; font-size:12px; color:#152A00; '
                f'font-variant-numeric:tabular-nums;">{row["tm30_count"]}</td>'
                f'</tr>'
            )
        return (
            '<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="margin:0 0 20px 0;">'
            '<thead><tr style="background:#152A00;">'
            '<th style="padding:8px 10px; text-align:left; font-size:8px; font-weight:700; text-transform:uppercase; '
            'letter-spacing:0.03em; color:#FFEFD2; white-space:nowrap;">Property</th>'
            '<th style="padding:8px 6px; text-align:center; font-size:8px; font-weight:700; text-transform:uppercase; '
            'letter-spacing:0.03em; color:#FFEFD2;">Code</th>'
            '<th style="padding:8px 6px; text-align:center; font-size:8px; font-weight:700; text-transform:uppercase; '
            'letter-spacing:0.03em; color:#FFEFD2;">RR4 Guests</th>'
            '<th style="padding:8px 6px; text-align:center; font-size:8px; font-weight:700; text-transform:uppercase; '
            'letter-spacing:0.03em; color:#FFEFD2;">TM30 Arrivals</th>'
            '</tr></thead>'
            f'<tbody>{"".join(body_rows)}</tbody>'
            '</table>'
        )

    def _build_rr4_tm30_property_email(self, prop: str, date_display: str,
                                        property_code: str, rr4_count: int, tm30_count: int,
                                        per_property_settings: dict) -> tuple:
        """Subject/body builder for one property's individual RR4/TM30
        email, split out for readability - same shape as
        _build_st_property_email."""
        subject = per_property_settings["subject"] \
            .replace("<<Property>>", prop) \
            .replace("<<PropertyCode>>", property_code) \
            .replace("<<Date>>", date_display)
        html_body = per_property_settings["html_template"] \
            .replace("<<Property>>", prop) \
            .replace("<<PropertyCode>>", property_code) \
            .replace("<<Date>>", date_display) \
            .replace("<<StatsTable>>", self._build_rr4_tm30_summary_table(
                [{"property_name": prop, "property_code": property_code,
                  "rr4_count": rr4_count, "tm30_count": tm30_count}]))
        return subject, html_body

    async def send_rr4_tm30_property_email(self, property_name: str, date_str: str,
                                            mark_sent: bool = True, sent_date_str: str = None,
                                            sync_type: str = "auto") -> dict:
        """
        Builds and sends ONE property's own RR4/TM30 email (Admin >
        Templates > RR4/TM30 Files (Per-Property)) - same shape as
        send_st_files_property_email, reading rr4_tm30_sync (populated by
        the daily RR4/TM30 auto-import) instead of hitting MEWS live, so
        the email can never disagree with that day's stored/verified export.

        No rr4_tm30_email_recipients (To) configured is a skip, not fatal -
        same as no imported data for that day. mark_sent writes THIS
        property's own rr4_tm30_email_last_sent_date.

        Every outcome logs to sync_logs under target_table="RR4/TM30 Email
        (Per-Property)".
        """
        p_res = self.supabase.table("property_api_settings").select(
            "id, rr4_tm30_email_recipients, rr4_tm30_email_cc, rr4_tm30_email_bcc, "
            "rr4_tm30_email_subject, rr4_tm30_email_template"
        ).eq("property_name", property_name).limit(1).execute()
        p = p_res.data[0] if p_res.data else {}
        prop_id = p.get("id")
        per_property_settings = {
            "subject": p.get("rr4_tm30_email_subject") or DEFAULT_RR4_TM30_DAILY_PER_PROPERTY_SUBJECT,
            "html_template": p.get("rr4_tm30_email_template") or DEFAULT_RR4_TM30_DAILY_PER_PROPERTY_TEMPLATE,
        }
        date_display = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d/%m/%Y")

        try:
            payload = self._read_rr4_tm30_cached_day(property_name, date_str)
            if not payload:
                raise ValueError(f"no imported RR4/TM30 report for {date_str} - import it first")
            property_code = self._resolve_property_code(property_name)
            rr4_bytes, rr4_filename = self._render_rr4_export(payload["rr4"], date_str, property_code)
            tm30_bytes, tm30_filename = self._render_tm30_export(payload["tm30"], date_str, property_code)
        except Exception as e:
            reason = f"{property_name}: {str(e)[:150]}"
            self._log_sync_row(property_name, prop_id, "RR4/TM30 Email (Per-Property)", "error", 0, reason, sync_type)
            return {"sent": False, "skipped": reason}

        recipients = [e.strip() for e in (p.get("rr4_tm30_email_recipients") or "").split(",") if e.strip()]
        if not recipients:
            reason = f"{property_name}: no RR4/TM30 Email Recipients configured (Admin > Templates > RR4/TM30 Files > Per-Property)"
            self._log_sync_row(property_name, prop_id, "RR4/TM30 Email (Per-Property)", "error", 0, reason, sync_type)
            return {"sent": False, "skipped": reason}
        cc = [e.strip() for e in (p.get("rr4_tm30_email_cc") or "").split(",") if e.strip()]
        bcc = [e.strip() for e in (p.get("rr4_tm30_email_bcc") or "").split(",") if e.strip()]

        subject, html_body = self._build_rr4_tm30_property_email(
            property_name, date_display, property_code,
            len(self._rr4_filed_rows(payload["rr4"].get("rows", []))),
            len(payload["tm30"].get("rows", [])),
            per_property_settings)
        email_service.send_email_with_attachments(
            recipients, subject, html_body,
            [(rr4_filename, rr4_bytes), (tm30_filename, tm30_bytes)],
            cc_emails=cc, bcc_emails=bcc)
        self._log_sync_row(property_name, prop_id, "RR4/TM30 Email (Per-Property)", "success", 1,
                            f"Email sent to {', '.join(recipients)}", sync_type)

        if mark_sent:
            try:
                self.supabase.table("property_api_settings").update(
                    {"rr4_tm30_email_last_sent_date": sent_date_str or date_str}
                ).eq("property_name", property_name).execute()
            except Exception as e:
                logger.warning(f"RR4/TM30 per-property email: failed to record last_sent_date for {property_name}: {e}")

        return {"sent": True, "skipped": None}

    async def send_rr4_tm30_bundled_digest(self, date_str: str, mark_sent: bool = True,
                                            sent_date_str: str = None, sync_type: str = "auto") -> dict:
        """
        The bundled RR4/TM30 email (Admin > Templates > RR4/TM30 Files) -
        one RR4+TM30 .xlsx pair per Thai property (see
        _RR4_TM30_EMAIL_EXCLUDED_PROPERTIES) that already has that day's
        report imported into rr4_tm30_sync. Standing master copy, same
        independent-of-per-property-opt-in shape as
        send_st_files_bundled_digest.

        Logs one sync_logs row per included property, target_table=
        "RR4/TM30 Email".
        """
        settings_row = email_service.get_rr4_tm30_daily_settings()
        props_res = self.supabase.table("property_api_settings").select(
            "id, property_name"
        ).order("property_name").execute()
        date_display = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d/%m/%Y")

        attachments, included, skipped, table_rows = [], [], [], []
        prop_ids = {}
        for p in (props_res.data or []):
            prop = p["property_name"]
            if prop in self._RR4_TM30_EMAIL_EXCLUDED_PROPERTIES:
                continue
            prop_ids[prop] = p.get("id")
            try:
                payload = self._read_rr4_tm30_cached_day(prop, date_str)
                if not payload:
                    raise ValueError(f"no imported RR4/TM30 report for {date_str} - import it first")
                property_code = self._resolve_property_code(prop)
                rr4_bytes, rr4_filename = self._render_rr4_export(payload["rr4"], date_str, property_code)
                tm30_bytes, tm30_filename = self._render_tm30_export(payload["tm30"], date_str, property_code)
                attachments.append((rr4_filename, rr4_bytes))
                attachments.append((tm30_filename, tm30_bytes))
                included.append(prop)
                table_rows.append({
                    "property_name": prop, "property_code": property_code,
                    # The count in the email must match the attachment beside
                    # it, so it counts the rows that survive _rr4_filed_rows -
                    # not the report's, which still carries MEWS's nameless
                    # occupant slots.
                    "rr4_count": len(self._rr4_filed_rows(payload["rr4"].get("rows", []))),
                    "tm30_count": len(payload["tm30"].get("rows", [])),
                })
            except Exception as e:
                reason = f"{prop}: {str(e)[:150]}"
                skipped.append(reason)
                self._log_sync_row(prop, prop_ids.get(prop), "RR4/TM30 Email", "error", 0, reason, sync_type)

        if not attachments:
            return {"sent": False, "included": included, "skipped": skipped}

        recipients = [e.strip() for e in (settings_row["recipients"] or "").split(",") if e.strip()]
        if not recipients:
            for prop in included:
                self._log_sync_row(prop, prop_ids.get(prop), "RR4/TM30 Email", "error", 0,
                                    "Bundled RR4/TM30 Email has no recipients configured (Admin > Templates)", sync_type)
            return {"sent": False, "included": [], "skipped": skipped + [f"{p}: bundled email has no recipients configured" for p in included]}

        subject = settings_row["subject"].replace("<<Date>>", date_display)
        html_body = settings_row["html_template"] \
            .replace("<<Date>>", date_display) \
            .replace("<<PropertyCount>>", str(len(included))) \
            .replace("<<PropertyList>>", ", ".join(included)) \
            .replace("<<StatsTable>>", self._build_rr4_tm30_summary_table(table_rows))

        email_service.send_email_with_attachments(recipients, subject, html_body, attachments)
        for prop in included:
            self._log_sync_row(prop, prop_ids.get(prop), "RR4/TM30 Email", "success", 1,
                                f"Bundled email sent to {', '.join(recipients)}", sync_type)

        if mark_sent:
            self._mark_rr4_tm30_daily_sent(settings_row, sent_date_str or date_str)

        return {"sent": True, "included": included, "skipped": skipped}

    def _mark_rr4_tm30_daily_sent(self, settings_row: dict, marker_date: str):
        """Same-day dedup guard for send_rr4_tm30_bundled_digest - mirrors
        _mark_st_files_daily_sent."""
        try:
            existing = self.supabase.table("email_templates").select("id") \
                .eq("template_key", RR4_TM30_DAILY_TEMPLATE_KEY).limit(1).execute()
            if existing.data:
                self.supabase.table("email_templates").update({"last_sent_date": marker_date}) \
                    .eq("id", existing.data[0]["id"]).execute()
            else:
                self.supabase.table("email_templates").insert({
                    "template_key": RR4_TM30_DAILY_TEMPLATE_KEY,
                    "subject": settings_row["subject"],
                    "html_template": settings_row["html_template"],
                    "recipients": settings_row["recipients"],
                    "send_hour": settings_row["send_hour"],
                    "send_minute": settings_row["send_minute"],
                    "enabled": True,
                    "last_sent_date": marker_date,
                }).execute()
        except Exception as e:
            logger.warning(f"RR4/TM30 daily email: failed to record last_sent_date: {e}")

    def _log_sync_row(self, prop, prop_id, target, status, count, msg, sync_type="auto"):
        """
        Same insert main.py's private _log_sync does (Admin > Activity Log /
        the Sync page's History widget both read sync_logs), duplicated here
        rather than imported - main.py already imports sync_service, so the
        reverse import would be circular. Swallows its own errors so a
        logging failure never breaks the action it's recording.
        """
        if not self.supabase:
            return
        try:
            self.supabase.table("sync_logs").insert({
                "property": prop,
                "property_id": prop_id,
                "target_table": target,
                "sync_type": sync_type,
                "status": status,
                "records_synced": count,
                "message": msg,
            }).execute()
        except Exception as e:
            logger.warning(f"sync_logs insert failed ({target}): {e}")

    async def send_ftp_upload(self, date_str: str, mark_sent: bool = True,
                               sent_date_str: str = None, sync_type: str = "auto") -> dict:
        """
        Uploads each property's report CSV(s) to the single global FTP
        destination (Admin > Sync > FTP Upload). Which report type(s) get
        included is controlled by ftp_settings.upload_st_files/
        upload_rv_files (that card's two checkboxes) - independent of each
        other, both can be on at once, uploading both files for a property
        in the same connection. ST uses get_st_report_export (same file the
        email digest/manual Download button build); RV uses get_rv_export,
        which can additionally raise for reasons ST never does (no verified
        GL chart of accounts, unmapped payment types) - caught by the same
        per-property try/except tolerance ST already used, just doubled.
        Properties not ready for a given type are silently skipped for that
        type only - the upload still runs for whoever/whatever's ready.

        Shared by main.py's scheduled job (mark_sent=True, sync_type="auto")
        and admin.py's manual "Upload Test Now" button (mark_sent=False so
        testing never suppresses that day's real scheduled upload via the
        last_sent_date guard below, sync_type="manual") - same split, same
        reasoning as the email digest.

        Logs one sync_logs row per property+type attempted, target_table=
        "ST Files FTP" or "RV Files FTP" so each page's own History widget
        (and the Activity Log) shows only the type it cares about.
        """
        settings_row = ftp_service.get_ftp_settings()
        if not settings_row["enabled"] or not settings_row["host"]:
            return {"uploaded": False, "included": [], "skipped": [],
                    "reason": "FTP upload is not configured or not enabled"}
        upload_st = settings_row.get("upload_st_files", True)
        upload_rv = settings_row.get("upload_rv_files", False)
        if not upload_st and not upload_rv:
            return {"uploaded": False, "included": [], "skipped": [],
                    "reason": "No file type selected to upload (Admin > Sync > FTP Upload)"}

        props_res = self.supabase.table("property_api_settings").select("id, property_name").order("property_name").execute()
        prop_ids = {p["property_name"]: p["id"] for p in (props_res.data or [])}

        files, included, skipped = [], [], []
        for prop, prop_id in prop_ids.items():
            if upload_st:
                try:
                    text, filename = self.get_st_report_export(prop, date_str)
                    files.append((filename, text.encode("utf-8"), prop, prop_id, "ST Files FTP"))
                    included.append(f"{prop} (ST)")
                except Exception as e:
                    skipped.append(f"{prop} (ST): {str(e)[:150]}")
            if upload_rv:
                try:
                    text, filename = self.get_rv_export(prop, date_str)
                    files.append((filename, text.encode("utf-8"), prop, prop_id, "RV Files FTP"))
                    included.append(f"{prop} (RV)")
                except Exception as e:
                    skipped.append(f"{prop} (RV): {str(e)[:150]}")

        if not files:
            return {"uploaded": False, "included": included, "skipped": skipped}

        result = ftp_service.upload_files(settings_row, [(f, d) for f, d, _, _, _ in files])
        if result.get("connection_error"):
            for _, _, prop, prop_id, target_table in files:
                self._log_sync_row(prop, prop_id, target_table, "error", 0,
                                    f"FTP Upload Failed: {result['connection_error']}", sync_type)
            return {"uploaded": False, "included": [], "skipped": skipped,
                    "reason": f"FTP connection failed: {result['connection_error']}"}

        failed_names = {name for name, _ in result["failed"]}
        for filename, _, prop, prop_id, target_table in files:
            if filename in failed_names:
                err = next(e for name, e in result["failed"] if name == filename)
                self._log_sync_row(prop, prop_id, target_table, "error", 0,
                                    f"FTP Upload Failed: {filename}: {err}", sync_type)
            else:
                self._log_sync_row(prop, prop_id, target_table, "success", 1,
                                    f"FTP Upload: {filename} -> {settings_row['host']}", sync_type)

        if mark_sent:
            marker_date = sent_date_str or date_str
            try:
                self.supabase.table("ftp_settings").update({"last_sent_date": marker_date}) \
                    .eq("id", settings_row["id"]).execute()
            except Exception as e:
                logger.warning(f"FTP upload: failed to record last_sent_date: {e}")

        return {
            "uploaded": True,
            "included": result["uploaded"],
            "skipped": skipped + [f"{filename}: {err}" for filename, err in result["failed"]],
        }

    # BCP timeline window: how far back/forward from "today" the reservations
    # grid covers. Wide enough to show most stays without blowing up capture
    # time or the encrypted snapshot's payload size.
    _BCP_WINDOW_DAYS_BACK = 7
    _BCP_WINDOW_DAYS_FORWARD = 7

    async def get_bcp_snapshot(self, property_name: str):
        """
        Builds the BCP (Business Continuity Plan) snapshot for one property,
        captured every 5 minutes: a MEWS-style reservation timeline (rooms x
        dates, today -7 to today +7) for the front desk to keep working from
        on paper if MEWS goes down, plus today's (Asia/Bangkok) customer
        profiles (tagged Arrival/In-house/Departure) and payments.

        Reservation-level notes come from serviceOrderNotes/getAll, which not
        every Connector token has enabled - that part degrades gracefully to
        empty notes rather than failing the snapshot.
        """
        bkk = ZoneInfo("Asia/Bangkok")
        now_bkk = datetime.now(bkk)
        day = now_bkk.replace(hour=0, minute=0, second=0, microsecond=0)
        day_start_utc = day.astimezone(timezone.utc)
        day_end_utc = (day + timedelta(days=1)).astimezone(timezone.utc)
        start_iso = day_start_utc.strftime("%Y-%m-%dT%H:%M:%SZ")
        end_iso = day_end_utc.strftime("%Y-%m-%dT%H:%M:%SZ")

        window_start = day - timedelta(days=self._BCP_WINDOW_DAYS_BACK)
        window_end = day + timedelta(days=self._BCP_WINDOW_DAYS_FORWARD + 1)  # +1: exclusive end
        window_start_iso = window_start.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        window_end_iso = window_end.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

        resv_res = await mews_client.post(
            "/api/connector/v1/reservations/getAll",
            {
                "StartUtc": window_start_iso,
                "EndUtc": window_end_iso,
                "Extent": {"Reservations": True, "Customers": True, "Resources": True},
            },
            property_name=property_name,
        )
        reservations = [r for r in resv_res.get("Reservations", []) if r.get("State") != "Canceled"]
        customers_map = {c["Id"]: c for c in resv_res.get("Customers", []) if c.get("Id")}
        resources_map = {r["Id"]: r for r in resv_res.get("Resources", []) if r.get("Id")}

        def in_window(ts):
            if not ts:
                return False
            try:
                t = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            except Exception:
                return False
            return day_start_utc <= t < day_end_utc

        # Today-specific subsets - used only for the customer Arrival/In-house/
        # Departure tags below, not for the timeline itself (which shows the
        # whole window).
        arrival_rs = [r for r in reservations if in_window(r.get("StartUtc"))]
        departure_rs = [r for r in reservations if in_window(r.get("EndUtc"))]
        inhouse_rs = [r for r in reservations if r.get("State") == "Started" and not in_window(r.get("EndUtc"))]

        # Product order items for every reservation in the window (breakfast
        # add-ons etc.) - any bar on the timeline can be clicked, not just
        # today's arrivals, so every reservation needs this fetched.
        items_by_reservation: dict = {}
        # Rate (room-charge) vs Items (product) net-value sums per reservation,
        # for the detail panel's Rate/Items/Total/Avg breakdown - MEWS's own
        # UI derives these from order items too (RequestedPaymentAmount on
        # the reservation itself is never populated in this property's data,
        # confirmed live), bucketed by the item's Type since NightRebate/
        # ProductOrderRebate need to net against their own category. Each
        # bucket also keeps its individual line items (date+amount for Rate,
        # name+amount for Items) for the panel's expand-to-see-detail rows.
        rate_amount_by_reservation: dict = {}
        items_amount_by_reservation: dict = {}
        rate_lines_by_reservation: dict = {}
        item_lines_by_reservation: dict = {}
        gross_amount_by_reservation: dict = {}
        currency_by_reservation: dict = {}
        # One BillId per reservation (first one seen among its order items -
        # a reservation's own charges all sit on the same bill in the simple,
        # non-grouped case this page handles) - resolved to the bill's own
        # display name below, for the Guest Profile's Billing tab.
        bill_id_by_reservation: dict = {}
        all_res_ids = [r["Id"] for r in reservations]
        for i in range(0, len(all_res_ids), 100):
            chunk = all_res_ids[i:i + 100]
            try:
                oi_res = await mews_client.post(
                    "/api/connector/v1/orderItems/getAll",
                    {"ServiceOrderIds": chunk, "Limitation": {"Count": 1000}},
                    property_name=property_name,
                )
                for item in oi_res.get("OrderItems", []):
                    if item.get("AccountingState") == "Canceled":
                        continue
                    order_id = item.get("ServiceOrderId")
                    if order_id and item.get("BillId") and order_id not in bill_id_by_reservation:
                        bill_id_by_reservation[order_id] = item.get("BillId")
                    item_type = item.get("Type")
                    item_amount = item.get("Amount") or {}
                    # MEWS's own Billing screen shows gross (tax-inclusive)
                    # everywhere - it has no separate net display at all.
                    # This used to sum NetValue here, so every rate/item
                    # line, the Rate/Items subtotals, and Total amount all
                    # came out ~7% (Thai VAT) short of what MEWS itself
                    # shows for the same reservation - confirmed line-by-line
                    # against a live bill. gross_amount_by_reservation below
                    # was already correct; everything else now matches it.
                    gross = item_amount.get("GrossValue") or 0
                    currency = item_amount.get("Currency")
                    if currency:
                        currency_by_reservation.setdefault(order_id, currency)
                    if item_type in ("SpaceOrder", "NightRebate", "ProductOrder", "ProductOrderRebate"):
                        gross_amount_by_reservation[order_id] = gross_amount_by_reservation.get(order_id, 0) + gross
                    if item_type in ("SpaceOrder", "NightRebate"):
                        rate_amount_by_reservation[order_id] = rate_amount_by_reservation.get(order_id, 0) + gross
                        start_utc = item.get("StartUtc")
                        if start_utc:
                            night_label = datetime.fromisoformat(start_utc.replace("Z", "+00:00")) \
                                .astimezone(ZoneInfo("Asia/Bangkok")).strftime("%d/%m")
                        else:
                            night_label = item.get("BillingName") or "Night"
                        rate_lines_by_reservation.setdefault(order_id, []).append({"label": night_label, "amount": gross, "_start": start_utc or ""})
                    elif item_type in ("ProductOrder", "ProductOrderRebate"):
                        items_amount_by_reservation[order_id] = items_amount_by_reservation.get(order_id, 0) + gross
                        product_label = item.get("BillingName") or item.get("Name") or "Product"
                        item_lines_by_reservation.setdefault(order_id, []).append({"label": product_label, "amount": gross, "_start": item.get("StartUtc") or ""})

                    if item_type != "ProductOrder":
                        continue
                    label = item.get("BillingName") or item.get("Name") or "Product"
                    count = item.get("UnitCount") or 1
                    items_by_reservation.setdefault(order_id, []).append(
                        f"{count}x {label}" if count != 1 else label
                    )
            except Exception as e:
                logger.warning(f"BCP order items fetch failed for {property_name}: {e}")

        # Resolve each reservation's BillId to the bill's own display name
        # (e.g. "LE-27-7-6043") for the Guest Profile's Billing tab - MEWS's
        # "Number" field is only populated once a bill is formally issued/
        # closed, staying null for an open one (confirmed live), so this
        # reads "Name" instead, which bills/getAll returns regardless of
        # State. Filtering by BillIds (not a date range) also works
        # regardless of IssuedUtc being null.
        bill_name_by_id: dict = {}
        all_bill_ids = list({bid for bid in bill_id_by_reservation.values() if bid})
        for i in range(0, len(all_bill_ids), 1000):
            chunk = all_bill_ids[i:i + 1000]
            try:
                bill_res = await mews_client.post(
                    "/api/connector/v1/bills/getAll",
                    {"BillIds": chunk, "Limitation": {"Count": 1000}},
                    property_name=property_name,
                )
                for b in bill_res.get("Bills", []):
                    if b.get("Id"):
                        bill_name_by_id[b["Id"]] = b.get("Name", "")
            except Exception as e:
                logger.warning(f"BCP bill name lookup failed for {property_name}: {e}")

        # orderItems/getAll doesn't return items in chronological order per
        # reservation (confirmed live: a 3-night stay came back as
        # 26/07, 25/07, 27/07) - sort each reservation's lines by the
        # underlying StartUtc, then drop the sort-only field.
        for lines_by_reservation in (rate_lines_by_reservation, item_lines_by_reservation):
            for lines in lines_by_reservation.values():
                lines.sort(key=lambda x: x["_start"])
                for line in lines:
                    del line["_start"]

        # Reservation-level notes (separate endpoint; permission-dependent)
        notes_by_reservation: dict = {}
        for i in range(0, len(all_res_ids), 100):
            chunk = all_res_ids[i:i + 100]
            try:
                notes_res = await mews_client.post(
                    "/api/connector/v1/serviceOrderNotes/getAll",
                    {"ServiceOrderIds": chunk, "Limitation": {"Count": 1000}},
                    property_name=property_name,
                )
                for note in notes_res.get("ServiceOrderNotes", []):
                    text = (note.get("Text") or "").strip()
                    if text:
                        # The field is OrderId, not ServiceOrderId despite the
                        # endpoint's own name - confirmed against a live
                        # response. Getting this wrong silently drops every
                        # note (they all bucket under the None key and never
                        # match a real reservation Id).
                        notes_by_reservation.setdefault(note.get("OrderId"), []).append({
                            "text": text,
                            "type": note.get("Type", ""),
                            "created_utc": note.get("CreatedUtc", ""),
                        })
            except Exception:
                break  # endpoint not enabled for this token - skip quietly
        # MEWS's own Notes panel lists newest first.
        for notes_list in notes_by_reservation.values():
            notes_list.sort(key=lambda n: n["created_utc"], reverse=True)

        # Per-guest payment history (Guest Profile's own Payments tab) -
        # MEWS links a payment to the paying Customer's AccountId, not a
        # ReservationId (confirmed live: every sampled payment - including
        # one matched to a specific guest by AccountId - had
        # ReservationId: null), so this is fetched per customer across the
        # whole window instead of per reservation.
        payments_by_customer: dict = {}
        all_customer_ids = list(customers_map.keys())
        for i in range(0, len(all_customer_ids), 100):
            chunk = all_customer_ids[i:i + 100]
            try:
                pay_res = await mews_client.post(
                    "/api/connector/v1/payments/getAll",
                    {"AccountIds": chunk, "Limitation": {"Count": 1000}},
                    property_name=property_name,
                )
                for p in pay_res.get("Payments", []):
                    account_id = p.get("AccountId")
                    if not account_id:
                        continue
                    amount = p.get("Amount") or {}
                    external = (p.get("Data") or {}).get("External") or {}
                    payments_by_customer.setdefault(account_id, []).append({
                        "created": p.get("CreatedUtc", ""),
                        # MEWS's own GrossValue is negative for money coming
                        # in (a payment reducing what's owed) - flipped here
                        # to the positive amount MEWS's own UI shows.
                        "amount": -(amount.get("GrossValue") or 0),
                        "currency": amount.get("Currency", ""),
                        "type": p.get("Type", ""),
                        "sub_type": external.get("Type", ""),
                        "identifier": p.get("Identifier") or external.get("ExternalIdentifier", ""),
                        "state": p.get("State", ""),
                        "notes": (p.get("Notes") or "").strip(),
                    })
            except Exception as e:
                logger.warning(f"BCP per-guest payments fetch failed for {property_name}: {e}")
        for pays in payments_by_customer.values():
            pays.sort(key=lambda x: x["created"], reverse=True)

        # Extra lookups for the "Manage" detail view (group name, requested
        # category, rate, company/travel agency) - IDs deduplicated across
        # the whole widened window so this stays a handful of calls per
        # property/hour, not one per reservation. Note: the Reservation
        # object's field is `GroupId` (confirmed against a live response),
        # not `ReservationGroupId` as get_mapped_reservations assumes -
        # that's a separate pre-existing bug in that method, left alone here
        # since it's out of scope for this change.
        group_ids = list({r.get("GroupId") for r in reservations if r.get("GroupId")})
        rate_ids = list({r.get("RateId") for r in reservations if r.get("RateId")})
        company_ids = {r.get("CompanyId") for r in reservations if r.get("CompanyId")}
        ta_ids = {r.get("TravelAgencyId") for r in reservations if r.get("TravelAgencyId")}
        all_company_ids = list(company_ids | ta_ids)
        segment_ids = list({r.get("BusinessSegmentId") for r in reservations if r.get("BusinessSegmentId")})

        async def fetch_entity(endpoint, payload_key, ids, response_key):
            if not ids:
                return {}
            try:
                # Limitation is required by some of these endpoints (e.g.
                # reservationGroups/getAll rejects its absence with a 400
                # "Invalid Limitation") even when filtering by an explicit Id
                # list - include it unconditionally rather than special-casing.
                res = await mews_client.post(
                    endpoint,
                    {payload_key: ids[:200], "Limitation": {"Count": 200}},
                    property_name=property_name,
                )
                return {item["Id"]: item for item in res.get(response_key, [])}
            except Exception as e:
                logger.warning(f"BCP {response_key} lookup failed for {property_name}: {e}")
                return {}

        groups_dict, rates_dict, companies_dict, segments_dict = await asyncio.gather(
            fetch_entity("/api/connector/v1/reservationGroups/getAll", "ReservationGroupIds", group_ids, "ReservationGroups"),
            fetch_entity("/api/connector/v1/rates/getAll", "RateIds", rate_ids, "Rates"),
            # companies/getAll filters by "Ids", not "CompanyIds" (confirmed
            # against MEWS docs) - see the comment on the other fetch_entity
            # call to this same endpoint above in this file.
            fetch_entity("/api/connector/v1/companies/getAll", "Ids", all_company_ids, "Companies"),
            fetch_entity("/api/connector/v1/businessSegments/getAll", "BusinessSegmentIds", segment_ids, "BusinessSegments"),
        )

        # Resource categories (+ per-room assignment) - needed for both the
        # reservation detail's "Requested category" and, unconditionally, for
        # grouping/ordering the Timeline's room rows (e.g. "The Duo | King"
        # spanning every room of that type, whether or not it currently has a
        # reservation in this window). Same ServiceIds quirk
        # get_st_files_report already works around - resolve the bookable
        # service first. Retries once on any failure (a transient MEWS
        # error/rate-limit here doesn't just lose category names - it also
        # wipes category_ordering, which silently collapses the Timeline's
        # whole group order back to an arbitrary/wrong one instead of
        # matching MEWS - confirmed against a real production capture where
        # a single failed cycle put "LADIES TRIBE HIDEOUT" ahead of "The Duo
        # | King" instead of the reverse). Still degrades to
        # empty/ungrouped after a second failure, or if the token lacks the
        # Resource Categories permission.
        categories_dict: dict = {}      # category_id -> localized name
        category_short_names: dict = {}  # category_id -> short code (e.g. "TNK"), may be empty
        category_ordering: dict = {}    # category_id -> MEWS's own display-order integer
        resource_category_id: dict = {}  # resource_id -> category_id
        stay_service_name: str = ""
        for attempt in range(2):
            categories_dict.clear()
            category_short_names.clear()
            category_ordering.clear()
            resource_category_id.clear()
            stay_service_name = ""
            try:
                services_res = await mews_client.post(
                    "/api/connector/v1/services/getAll",
                    {"Limitation": {"Count": 100}},
                    property_name=property_name,
                )
                stay_service = self._resolve_stay_service(services_res.get("Services", []))
                if stay_service:
                    stay_service_name = stay_service.get("Name", "")
                    cats_res = await mews_client.post(
                        "/api/connector/v1/resourceCategories/getAll",
                        {"ServiceIds": [stay_service["Id"]], "Limitation": {"Count": 200}},
                        property_name=property_name,
                    )
                    for c in cats_res.get("ResourceCategories", []):
                        names = c.get("Names") or {}
                        shorts = c.get("ShortNames") or {}
                        categories_dict[c["Id"]] = names.get("en-US") or names.get("en-GB") or next(iter(names.values()), "")
                        category_short_names[c["Id"]] = shorts.get("en-US") or shorts.get("en-GB") or next(iter(shorts.values()), "")
                        category_ordering[c["Id"]] = c.get("Ordering", 0)

                    all_cat_ids = list(categories_dict.keys())
                    if all_cat_ids:
                        assign_res = await mews_client.post(
                            "/api/connector/v1/resourceCategoryAssignments/getAll",
                            {"ResourceCategoryIds": all_cat_ids, "Limitation": {"Count": 1000}},
                            property_name=property_name,
                        )
                        for a in assign_res.get("ResourceCategoryAssignments", []):
                            if a.get("IsActive", True):
                                resource_category_id[a["ResourceId"]] = a["CategoryId"]
                break  # success - no need to retry
            except Exception as e:
                if attempt == 0:
                    logger.warning(f"BCP resource categories lookup failed for {property_name}, retrying once: {e}")
                    await asyncio.sleep(1)
                    continue
                logger.warning(f"BCP resource categories lookup failed for {property_name}: {e}")

        # MEWS's Reservation.Origin is a combined string like "CommanderInPerson"
        # (Origin enum + CommanderOrigin sub-enum concatenated - confirmed
        # against a live response, not split into two fields as the public
        # docs describe). Reconstructs MEWS's own full label ("Mews Operations
        # In person") from the documented enum value lists.
        #
        # Channel-manager (OTA) bookings get a richer, different Origin format
        # entirely - confirmed against a live Agoda-via-SiteMinder reservation:
        # "SiteMinder 1750892592: AGO-1750892592-01" is
        # f"{ChannelManager} {ChannelNumber}: {ChannelManagerNumber}", not the
        # generic prefix+suffix split used for Commander/Distributor/etc.
        #
        # "Reservation source" prefers OriginDetails when MEWS populates it
        # (it did for that same Agoda reservation - "TravelBundle, Agoda"),
        # falling back to the derived suffix/channel-manager name otherwise -
        # confirmed against the walk-in reservation, whose OriginDetails was
        # null but still showed a "Reservation source" ("In person").
        _ORIGIN_PREFIX_LABELS = {
            "Distributor": "Booking Engine",
            "ChannelManager": "Channel Manager",
            "Commander": "Mews Operations",
            "Import": "Import",
            "Connector": "Connector API",
            "Navigator": "Guest Services",
        }

        def format_origin(res: dict):
            raw = res.get("Origin") or ""
            origin_details = (res.get("OriginDetails") or "").strip()
            if not raw:
                return "", origin_details

            if raw.startswith("ChannelManager"):
                cm = res.get("ChannelManager")
                cn = res.get("ChannelNumber")
                cmn = res.get("ChannelManagerNumber")
                if cm and cn and cmn:
                    return f"{cm} {cn}: {cmn}", (origin_details or cm)

            for prefix, label in _ORIGIN_PREFIX_LABELS.items():
                if raw.startswith(prefix):
                    suffix = raw[len(prefix):]
                    if not suffix:
                        return label, origin_details
                    words = re.findall(r"[A-Z][a-z]*", suffix)
                    if not words:
                        return label, (origin_details or suffix)
                    source = words[0] + ("" if len(words) == 1 else " " + " ".join(w.lower() for w in words[1:]))
                    return f"{label} {source}", (origin_details or source)
            return raw, (origin_details or raw)

        def extract_guest_identity(c):
            """
            MEWS Customer -> the guest-identity fields the Reg Card/RR3 form
            and the Guest Profile page need. Shared by the primary guest
            (CustomerId) and each companion (CompanionIds) below - same
            extraction either way, just applied to a different Customer
            record, so a reservation's companions get full profiles too
            instead of just a bare name.

            occupation/address_details are the RAW MEWS values only (empty
            if MEWS has none) - the Guest Profile page shows exactly this,
            and must never show anything MEWS didn't actually provide. The
            Reg Card/RR3 print form's own print-safe defaults ("นักธุรกิจ"
            for a blank occupation, nationality as a last-resort address so
            the government form field isn't literally empty) are applied
            separately, only at Reg Card token-build time in the frontend
            (buildRegCardTokens) - never baked in here, or the Guest Profile
            page would show fabricated data as if MEWS provided it.
            """
            identity_card_value = ""
            identity_card = c.get("IdentityCard")
            identity_cards = c.get("IdentityCards")
            if isinstance(identity_card, dict):
                identity_card_value = identity_card.get("Number", "")
            elif isinstance(identity_cards, list) and identity_cards:
                identity_card_value = identity_cards[0].get("Number", "")
            passport = c.get("Passport") or {}
            cust_address = c.get("Address") or {}
            # MEWS's own Addresses table has separate Address/City/Country
            # columns - a guest can have just a Country set with Address and
            # City both blank (confirmed live: this exact guest's Address
            # record has only CountryCode "VN", everything else null) and
            # MEWS still shows that row rather than "No data available".
            # Checking Line1/Line2 alone missed this - only counted as "no
            # address" when actually MEWS had a country on file.
            addr_country_name = _rr3_country_name(cust_address.get("CountryCode")) if cust_address.get("CountryCode") else ""
            addr_parts = [cust_address.get("Line1"), cust_address.get("Line2"),
                          cust_address.get("City"), cust_address.get("PostalCode"), addr_country_name]
            if any(addr_parts):
                address_details = " ".join(p for p in addr_parts if p)
            elif (c.get("BirthPlace") or "").strip():
                address_details = c.get("BirthPlace")
            else:
                address_details = ""
            return {
                "name": f"{c.get('FirstName', '')} {c.get('LastName', '')}".strip(),
                # Real MEWS fields, not a guess split off the combined name
                # above - a guest with a multi-word first OR last name (e.g.
                # FirstName "Minh Dat", LastName "Le") would otherwise split
                # wrong (confirmed against this exact guest live).
                "first_name": c.get("FirstName", ""),
                "last_name": c.get("LastName", ""),
                "second_last_name": c.get("SecondLastName", ""),
                "title": _mews_title_display(c.get("Title")),
                "sex": c.get("Sex", ""),
                "language": c.get("PreferredLanguageCode") or c.get("LanguageCode") or "",
                "birth_date": c.get("BirthDate", ""),
                "birth_country_name": _rr3_country_name(c.get("BirthCountryCode")) if c.get("BirthCountryCode") else "",
                "birth_place": c.get("BirthPlace", ""),
                "nationality": c.get("NationalityCode", ""),
                "nationality_name": _rr3_country_name(c.get("NationalityCode")),
                "email": c.get("Email", ""),
                "phone": c.get("Phone", ""),
                "identity_card_number": identity_card_value,
                "passport_number": passport.get("Number", ""),
                "occupation": c.get("Occupation", ""),
                "address_details": address_details,
                "alien_book": c.get("IdentityDocumentSupportNumber", ""),
                "mews_customer_id": c.get("Id", ""),
                "payments": payments_by_customer.get(c.get("Id", ""), []),
            }

        def reservation_row(res):
            customer = customers_map.get(res.get("CustomerId"), {})
            room = resources_map.get(res.get("AssignedResourceId"), {})
            group = groups_dict.get(res.get("GroupId"), {})
            category_name = categories_dict.get(res.get("RequestedCategoryId"), "")
            rate = rates_dict.get(res.get("RateId"), {})
            company = companies_dict.get(res.get("CompanyId"), {})
            travel_agency = companies_dict.get(res.get("TravelAgencyId"), {})
            segment = segments_dict.get(res.get("BusinessSegmentId"), {})
            res_id = res.get("Id")
            rate_amount = rate_amount_by_reservation.get(res_id, 0)
            items_amount = items_amount_by_reservation.get(res_id, 0)
            gross_total = gross_amount_by_reservation.get(res_id, 0)
            origin_label, reservation_source = format_origin(res)

            # Same customer-profile extraction as get_rr3_cards - the Extent
            # here already includes Customers:True, so this is just reading
            # more fields off data already fetched for the Timeline, not an
            # extra live call. Captured into the snapshot at capture time, so
            # it's still there for the Reg Card even once MEWS is down.
            guest_identity = extract_guest_identity(customer)

            # Additional named guests on the same reservation (MEWS attaches
            # them via CompanionIds, alongside the primary CustomerId/"Owner"
            # surfaced as `guest` below) - already in customers_map since the
            # Extent's Customers:True includes them (used below to build the
            # Arrival/In-house/Departure customer list), just never resolved
            # to full profiles on the reservation itself before, so a
            # reservation with more than 1 adult/child only ever showed its
            # Owner. CompanionIds includes the Owner's own CustomerId as one
            # of its entries (confirmed live: reservation 89110's
            # CompanionIds started with its own CustomerId) - excluded here,
            # or the Owner shows up a second time with no "Owner" label.
            customer_id = res.get("CustomerId")
            companions = [
                extract_guest_identity(customers_map[cid])
                for cid in (res.get("CompanionIds") or [])
                if cid != customer_id and customers_map.get(cid)
            ]

            return {
                "number": res.get("Number", ""),
                # MEWS's own internal Id (a GUID, distinct from the
                # human-readable Number above) - needed as ServiceOrderId
                # when pushing a locally-added note back into MEWS via
                # serviceOrderNotes/add (see sync_pending_reservation_notes).
                "mews_reservation_id": res.get("Id", ""),
                "guest": guest_identity["name"],
                "first_name": guest_identity["first_name"],
                "last_name": guest_identity["last_name"],
                "second_last_name": guest_identity["second_last_name"],
                "title": guest_identity["title"],
                "sex": guest_identity["sex"],
                "language": guest_identity["language"],
                "birth_date": guest_identity["birth_date"],
                "birth_country_name": guest_identity["birth_country_name"],
                "birth_place": guest_identity["birth_place"],
                "nationality": guest_identity["nationality"],
                "nationality_name": guest_identity["nationality_name"],
                "email": guest_identity["email"],
                "phone": guest_identity["phone"],
                "identity_card_number": guest_identity["identity_card_number"],
                "passport_number": guest_identity["passport_number"],
                "occupation": guest_identity["occupation"],
                "address_details": guest_identity["address_details"],
                "alien_book": guest_identity["alien_book"],
                "mews_customer_id": guest_identity["mews_customer_id"],
                "payments": guest_identity["payments"],
                "bill_name": bill_name_by_id.get(bill_id_by_reservation.get(res_id), ""),
                "room": room.get("Name", ""),
                "check_in": res.get("StartUtc", ""),
                "check_out": res.get("EndUtc", ""),
                "state": res.get("State", ""),
                "adults": res.get("AdultCount", 0),
                "children": res.get("ChildCount", 0),
                "companions": companions,
                "products": items_by_reservation.get(res_id, []),
                "notes": notes_by_reservation.get(res_id, []),
                "group_name": group.get("Name", ""),
                "category": category_name,
                "rate": rate.get("Name", ""),
                "company": company.get("Name", ""),
                "travel_agency": travel_agency.get("Name", ""),
                # ChannelNumber is the OTA/channel's own booking reference,
                # e.g. Agoda's own confirmation number for the stay - not
                # MEWS's own "Number" - confirmed against a live Agoda
                # reservation matching this exactly.
                "travel_agency_confirmation_number": res.get("ChannelNumber", ""),
                "rate_amount": rate_amount,
                "items_amount": items_amount,
                "rate_lines": rate_lines_by_reservation.get(res_id, []),
                "item_lines": item_lines_by_reservation.get(res_id, []),
                "total_amount": rate_amount + items_amount,
                "total_amount_gross": gross_total,
                # This used to read Reservation.RequestedPaymentAmount, which
                # turned out to be a specific prepayment/deposit request field,
                # not a running balance - it's almost never populated, so
                # "To be paid" showed 0 even for reservations MEWS's own
                # Billing screen shows a real outstanding amount on (confirmed
                # by the user against a live reservation). MEWS's Billing
                # screen itself derives the outstanding amount as charges
                # minus payments, so this does the same: gross charges (same
                # total shown above) minus payments received by the guest.
                # Can go negative if the guest overpaid (a credit balance).
                "to_be_paid": gross_total - sum(p["amount"] for p in guest_identity["payments"]),
                "currency": currency_by_reservation.get(res_id, ""),
                "service": stay_service_name,
                "segment": segment.get("Name", ""),
                "origin": origin_label,
                "reservation_source": reservation_source,
                "purpose": res.get("Purpose", ""),
                "created_utc": res.get("CreatedUtc", ""),
                # MEWS shows this as a padlock toggle next to the room number
                # (locked = this exact room is guaranteed; unlocked = MEWS
                # may still reassign the room before check-in) - read-only
                # here, same as everything else in BCP.
                "room_locked": bool(res.get("AssignedResourceLocked")),
            }

        def sort_key(row):
            return row["room"] or "zzz"

        # The whole window's reservations, flat - this is what the frontend's
        # Timeline positions as bars across room rows x date columns.
        timeline_reservations = sorted((reservation_row(r) for r in reservations), key=sort_key)

        # Customer profiles for everyone attached to today's reservations,
        # tagged by how they relate to today (arrival/departure/in-house).
        def guest_ids(res_list):
            ids = set()
            for r in res_list:
                for cid in [r.get("CustomerId")] + (r.get("CompanionIds") or []):
                    if cid:
                        ids.add(cid)
            return ids

        arrival_guests = guest_ids(arrival_rs)
        departure_guests = guest_ids(departure_rs)
        inhouse_guests = guest_ids(inhouse_rs)
        customers = []
        for cid in arrival_guests | departure_guests | inhouse_guests:
            c = customers_map.get(cid)
            if not c:
                continue
            tags = []
            if cid in arrival_guests:
                tags.append("Arrival")
            if cid in inhouse_guests:
                tags.append("In-house")
            if cid in departure_guests:
                tags.append("Departure")
            customers.append({
                "name": f"{c.get('FirstName', '')} {c.get('LastName', '')}".strip(),
                "tags": tags,
                "nationality": c.get("NationalityCode", ""),
                "email": c.get("Email", ""),
                "phone": c.get("Phone", ""),
                "notes": (c.get("Notes") or "").strip(),
            })
        customers.sort(key=lambda c: c["name"])

        # Today's payments
        payments = []
        try:
            pay_res = await mews_client.post(
                "/api/connector/v1/payments/getAll",
                {"CreatedUtc": {"StartUtc": start_iso, "EndUtc": end_iso}, "Limitation": {"Count": 1000}},
                property_name=property_name,
            )
            res_number_by_id = {r["Id"]: r.get("Number", "") for r in reservations}
            for p in pay_res.get("Payments", []):
                amount = p.get("Amount") or {}
                customer = customers_map.get(p.get("AccountId"), {})
                payments.append({
                    "created": p.get("CreatedUtc", ""),
                    "type": p.get("Type", ""),
                    "state": p.get("State", ""),
                    "amount": amount.get("GrossValue", 0),
                    "currency": amount.get("Currency", ""),
                    "guest": f"{customer.get('FirstName', '')} {customer.get('LastName', '')}".strip(),
                    "reservation": res_number_by_id.get(p.get("ReservationId"), ""),
                    "notes": (p.get("Notes") or "").strip(),
                })
            payments.sort(key=lambda x: x["created"], reverse=True)
        except Exception as e:
            logger.warning(f"BCP payments fetch failed for {property_name}: {e}")

        # Room list (the Timeline's Y-axis): every active resource + today's
        # housekeeping state. Who's in/arriving/departing each room is now
        # derived client-side from `reservations` (the whole window), not
        # computed here - the timeline bars already show that.
        #
        # MEWS sells some physical spaces multiple ways - e.g. room 210 as a
        # whole-dorm buyout ("TRIBE HIDEOUT - ALL YOURS!") AND its individual
        # beds 211-219 as their own bookable category ("1 BED IN OUR TRIBE
        # HIDEOUT") - and nests the beds directly under the parent room in
        # its own Timeline regardless of the beds' different category.
        # Mirror that here: top-level (parentless) resources are grouped/
        # sorted by their own category as before, but a resource with a
        # ParentResourceId is nested immediately after its parent instead of
        # sorting into its own category's slot. `group_category`
        # carries the category used for the vertical group-label span (the
        # parent's, for children) separately from `category` (the room's own
        # true category, still used for its Room Properties/tooltip) so a
        # family of parent+children reads as one section even though the
        # children's real category differs.
        all_rooms_res = await mews_client.post(
            "/api/connector/v1/resources/getAll",
            {"Extent": {"Resources": True}, "Limitation": {"Count": 1000}},
            property_name=property_name,
        )
        raw_resources = []
        for r in all_rooms_res.get("Resources", []):
            if not r.get("IsActive", True):
                continue
            data_val = (r.get("Data") or {}).get("Value") or {}
            cat_id = resource_category_id.get(r["Id"])
            raw_resources.append({
                "id": r["Id"],
                "parent_id": r.get("ParentResourceId") or "",
                "room": r.get("Name", ""),
                "floor": data_val.get("FloorNumber", ""),
                "state": r.get("State", ""),
                "category": categories_dict.get(cat_id, ""),
                "category_short": category_short_names.get(cat_id, ""),
                "category_ordering": category_ordering.get(cat_id, 0),
            })

        by_id = {res["id"]: res for res in raw_resources}
        children_by_parent: dict = {}
        top_level = []
        for res in raw_resources:
            # An orphaned child (parent inactive/missing from this list)
            # falls back to top-level so it doesn't silently vanish.
            if res["parent_id"] and res["parent_id"] in by_id:
                children_by_parent.setdefault(res["parent_id"], []).append(res)
            else:
                top_level.append(res)

        def room_sort_key(res):
            name = res["room"]
            return (int(name), name) if name.isdigit() else (float("inf"), name)

        for kids in children_by_parent.values():
            kids.sort(key=room_sort_key)
        # Group order follows MEWS's own resourceCategories.Ordering field
        # (confirmed against a real property: alphabetical grouping put
        # unrelated categories first and buried "The Duo | King" mid-list,
        # while Ordering reproduces MEWS's actual on-screen sequence).
        # Ordering ties (MEWS sometimes assigns the same number to a few
        # categories) fall back to the category name, then numeric room
        # order, so the result is still fully deterministic.
        top_level.sort(key=lambda res: (
            1 if not res["category"] else 0,
            res["category_ordering"],
            res["category"],
            room_sort_key(res),
        ))

        def to_room_row(res, parent_room_name, group_category, group_category_short, is_child):
            return {
                "room": res["room"],
                "floor": res["floor"],
                "state": res["state"],
                "category": res["category"],
                "category_short": res["category_short"],
                "parent_room": parent_room_name,
                "service": stay_service_name,
                "group_category": group_category,
                "group_category_short": group_category_short,
                "is_child": is_child,
            }

        rooms = []
        for parent in top_level:
            rooms.append(to_room_row(parent, "", parent["category"], parent["category_short"], False))
            for child in children_by_parent.get(parent["id"], []):
                rooms.append(to_room_row(child, parent["room"], parent["category"], parent["category_short"], True))

        return {
            "property": property_name,
            "date": day.strftime("%Y-%m-%d"),
            "captured_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "window": {
                "start": window_start.strftime("%Y-%m-%d"),
                "end": (window_end - timedelta(days=1)).strftime("%Y-%m-%d"),  # last inclusive day
            },
            "counts": {
                "customers": len(customers),
                "payments": len(payments),
                "rooms": len(rooms),
            },
            "rooms": rooms,
            "reservations": timeline_reservations,
            "customers": customers,
            "payments": payments,
        }

    async def sync_pending_reservation_notes(self, property_name: str):
        """
        Pushes every note added through our own system while MEWS was down
        (bcp_reservation_notes, synced_to_mews = False) into MEWS itself via
        serviceOrderNotes/add, then marks it synced. Called from
        capture_snapshot right after a successful get_bcp_snapshot for the
        same property - reaching that point already proves MEWS is
        reachable, so there's no separate "is MEWS back up" check needed.

        This is the one field in BCP that writes back to MEWS automatically,
        per explicit instruction - and strictly an addition; nothing here
        ever edits or deletes anything already in MEWS. A note that fails to
        push (MEWS flaked again mid-sync) simply stays unsynced and is
        retried on the next successful capture.
        """
        if not self.supabase:
            return
        try:
            pending = self.supabase.table("bcp_reservation_notes") \
                .select("id, mews_reservation_id, text, created_by") \
                .eq("property", property_name) \
                .eq("synced_to_mews", False) \
                .execute()
        except Exception as e:
            logger.warning(f"BCP pending-notes lookup failed for {property_name}: {e}")
            return
        for row in (pending.data or []):
            if not row.get("mews_reservation_id") or not row.get("text"):
                continue
            # Attribute the note to whoever actually typed it into our
            # system - MEWS's own Text field is plain text with no separate
            # author field to carry this in, so it's appended here rather
            # than dropped. Only affects what's sent to MEWS; the row's own
            # `text` column (and everywhere we display it ourselves) stays
            # exactly as typed.
            text = row["text"]
            if row.get("created_by"):
                text = f"{text}\n\nAdded via NHGOne by {row['created_by']}"
            try:
                await mews_client.post(
                    "/api/connector/v1/serviceOrderNotes/add",
                    {"ServiceOrderNotes": [{"ServiceOrderId": row["mews_reservation_id"], "Text": text}]},
                    property_name=property_name,
                )
                self.supabase.table("bcp_reservation_notes").update({
                    "synced_to_mews": True,
                    "synced_at": datetime.now(timezone.utc).isoformat(),
                }).eq("id", row["id"]).execute()
            except Exception as e:
                logger.warning(f"BCP note sync-to-MEWS failed for {property_name} note {row['id']}: {e}")

sync_service = SyncService()
