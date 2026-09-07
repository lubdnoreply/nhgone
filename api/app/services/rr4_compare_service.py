"""RR4 / TM30 verification: our stored register vs each property's own
"RR4-TM30-<Name>-Gen" Google Sheet, which is the ground truth for what
actually gets filed with the authorities.

The RR4/TM30 counterpart to st_compare_service, and deliberately the same
shape - but three things differ, each forced by the source sheets:

1. **Thailand only.** Lub d Siem Reap and Lub d Philippines Makati don't file
   under the Thai Hotel Act and have no generator sheet at all (same exclusion
   sync_service._RR4_TM30_EMAIL_EXCLUDED_PROPERTIES already applies to the
   daily RR4/TM30 digest).

2. **Each property is compared at ITS OWN date**, not one date shared by all -
   unlike the ST sheets, which are all pasted within an hour of each other.
   These windows run to the property's own cutoff hour, and Chinatown's is
   12:15 where everyone else's is ~02:00, so at any given moment Chinatown's
   sheet is a full day behind the rest. Demanding one common date would mean
   never sending a mail at all.

3. **Rows are paired by identity, then compared column by column.** A
   key-based diff that stops at "we found a row with this passport" hides real
   defects (on 2026-08-22 a 4-field key called Patong's TM30 a perfect match
   while a full-field diff surfaced 6 differing rows). So the passport/PID is
   used ONLY to pair the two sides up; every column of a paired row is then
   compared, and a name that changed shows up as a column difference rather
   than as two unmatched rows.

Known drift is reported separately from real differences - see _KNOWN_DRIFT
below. Those four patterns have each been investigated and confirmed to be
the sheet or MEWS moving on after the fact, not our bug; counting them as
mismatches every single day would bury the differences that do matter.
"""
import asyncio
import io
import logging
import re
from collections import OrderedDict
from datetime import datetime, timedelta, timezone

import httpx
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# The six Thai properties' generator workbooks. Marasca is the "MRCS"-style
# sheet, NOT `1tY_kX...`/MRCKY-Gen, whose ImportInhouse was empty when checked.
SHEETS = OrderedDict([
    ("Lub d Bangkok Chinatown",       ("Chinatown", "1qT4ZClqvTLVUW9Bc4Oaxx2oy6u8QZo0dCVlyM35JRy4")),
    ("Lub d Bangkok Siam",            ("Siam",      "1liiB8tqYGCgAyKqDsRCCaZSonubHgnMr2YSZQr7-SlQ")),
    ("Lub d Koh Samui Chaweng Beach", ("Samui",     "1nanCOqRnRjiFzkQ_l0RyZqJ0oZ_LGJ6-RTB0qDMTAyg")),
    ("Lub d Koh Tao Tanote Bay",      ("Koh Tao",   "1akGkOIoHKURs6DihwkCRw5zx37HkWVlaYFSdjj-KI6c")),
    ("Lub d Phuket Patong",           ("Patong",    "1XKfU7pSyMwSFIiq7g1wKlKJB_gW9d1JtTuniahqBja8")),
    ("Marasca Samui",                 ("Marasca",   "1YZD0CYpaOwxSiHLa7iH_7bK3ED5CAIhdR2GizwKKwuI")),
])

# rowNo is excluded from the comparison: both sides renumber their own rows
# from 1, and the two exports don't emit guests in the same order (the sheet
# lists a room's unnamed occupant slot first, we don't), so it would report a
# difference on almost every row while meaning nothing.
_SKIP_RR4_COLUMNS = {"row_no"}

# Differences that have each been chased down and confirmed as the world
# moving on after the sheet was generated, not a defect on our side. Reported
# in their own column so the "real differences" number stays meaningful.
_KNOWN_DRIFT = {
    "time_check_in":
        "MEWS wrote ActualStartUtc at :59 seconds, right after the sheet was generated (sheet is exactly 1 minute behind)",
    "date_check_out":
        "Guest checked out earlier than scheduled, after the sheet was generated (ours is ahead of the sheet)",
    "check_out_date":
        "Guest checked out earlier than scheduled, after the sheet was generated (ours is ahead of the sheet)",
    "birth_date":
        "Sheet prints 30/12/1899 when MEWS has no birth date (Excel's render of an empty value) - ours leaves it blank, which is correct",
}


def _norm(v) -> str:
    """One cell, normalized for comparison. Strips the leading apostrophe the
    sheets use to force Text formatting (our export carries it too, on the
    date columns only - see the RR4 importer's own rule), collapses 63 vs
    63.0 vs "63", and treats None and "" as the same empty."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    if s.startswith("'"):
        s = s[1:].strip()
    return s


def _dmy(s: str):
    """(y, m, d) from a dd/mm/yyyy string - Buddhist or Christian era, since
    both sides of any one column always use the same one."""
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    return (int(m.group(3)), int(m.group(2)), int(m.group(1))) if m else None


def _hhmm(s: str):
    """Minutes-since-midnight from the sheets' "HH.MM" check-in time."""
    m = re.match(r"^(\d{1,2})[.:](\d{2})$", s)
    return int(m.group(1)) * 60 + int(m.group(2)) if m else None


def _is_known_drift(key: str, ours: str, sheet: str) -> bool:
    if key == "time_check_in":
        a, b = _hhmm(ours), _hhmm(sheet)
        return a is not None and b is not None and b - a == 1
    if key in ("date_check_out", "check_out_date"):
        a, b = _dmy(ours), _dmy(sheet)
        return a is not None and b is not None and a < b
    if key == "birth_date":
        return ours == "" and sheet == "30/12/1899"
    return False


def _pair_key(row: dict, kind: str) -> tuple:
    """What identifies the same guest on both sides. Passport first (the one
    field the filing itself is keyed on), then Thai national ID, then the
    name - and for MEWS's unnamed occupant slots, which carry none of the
    three, the room and check-in they were booked under."""
    if kind == "rr4":
        pp, pid = _norm(row.get("passport")).upper(), _norm(row.get("pid")).upper()
        if pp:
            return ("P", pp)
        if pid:
            return ("I", pid)
        name = (_norm(row.get("name_en")) + "|" + _norm(row.get("surname_en"))).upper()
        if name != "|":
            return ("N", name)
        # Deliberately NOT keyed on time_check_in: the two exports routinely
        # disagree by a minute on it (_is_known_drift has a rule for exactly
        # that), and a field known to drift cannot also be an identity. When
        # it was part of this key the drift stopped the row pairing at all,
        # so one unnamed slot surfaced as "only ours" AND "only sheet"
        # instead of as the known drift it is - Patong room 2313 on
        # 2026-08-26 (ours 06.42, sheet 06.43) was doing exactly that. Room +
        # check-in date still separates the slots; two of them sharing even
        # that are paired by content in _index.
        return ("X", _norm(row.get("room_no")), _norm(row.get("date_check_in")))
    pp = _norm(row.get("passport_no")).upper()
    if pp:
        return ("P", pp)
    return ("N", (_norm(row.get("first_name")) + "|" + _norm(row.get("last_name"))).upper())


def _row_label(row: dict, kind: str) -> str:
    """One guest, named well enough to be found in MEWS from the mail alone.

    Name first because that is what a person recognises, then the number the
    filing is actually keyed on, then (RR4 only) the room - which is all there
    is to go on for MEWS's unnamed occupant slots, since they carry neither of
    the first two. TM30 has no room column at all, so it gets name + passport.
    """
    if kind == "rr4":
        name = " ".join(x for x in (_norm(row.get("name_en")), _norm(row.get("surname_en"))) if x)
        ident = _norm(row.get("passport")) or _norm(row.get("pid"))
        room = _norm(row.get("room_no"))
    else:
        name = " ".join(x for x in (_norm(row.get("first_name")), _norm(row.get("last_name"))) if x)
        ident = _norm(row.get("passport_no"))
        room = ""
    bits = [name or "(no name)"]
    if ident:
        bits.append(ident)
    if room:
        bits.append(f"room {room}")
    return " · ".join(bits)


# How many differing guests the detail table names per column, and per
# one-side-only group. The counts printed beside them are always the FULL
# number and the table's own footnote states these caps, so a truncated list
# can never be read as the complete one.
_EXAMPLES_PER_COLUMN = 3
_EXAMPLES_PER_MISSING_GROUP = 6


def _index(rows: list, kind: str, columns: list) -> dict:
    """Rows by pair key, with an occurrence counter appended so two guests
    sharing a passport (or two unnamed slots in the same room) stay distinct
    instead of one silently overwriting the other.

    Where a key DOES repeat, the duplicates are ordered by their own compared
    values rather than by the order the export happened to emit them. The two
    sides genuinely do emit rows in different orders (see the module
    docstring), so numbering by first-seen paired a guest's first row against
    the sheet's second and reported every column that differs between the
    guest's own two rows as a difference on both of them - twice over, once
    in each direction. Verified 2026-08-26: Chinatown's Andrea Solves Vidal
    (530 @ 21.43 + 404 @ 21.44) and Siam's PATTRAPORN CHANIM (105 @ 13.59 +
    101 @ 14.00) each held identical data on both sides in opposite order,
    and accounted for 4 of the 5 "real" RR4 differences that day. Sorting on
    the compared columns is deterministic and identical on both sides, so
    matching rows line up and genuinely different ones still report.
    """
    groups = {}
    for row in rows:
        groups.setdefault(_pair_key(row, kind), []).append(row)
    out = {}
    for k, group in groups.items():
        if len(group) > 1:
            group = sorted(group, key=lambda r: tuple(_norm(r.get(c)) for c in columns))
        for i, row in enumerate(group, 1):
            out[k + (i,)] = row
    return out


def _compare_rows(ours: list, sheet: list, kind: str, columns: list) -> dict:
    """One property, one register. Pairs the two sides up by identity, then
    compares every column of every paired row."""
    ours_ix, sheet_ix = _index(ours, kind, columns), _index(sheet, kind, columns)
    paired = ours_ix.keys() & sheet_ix.keys()

    diff_rows, drift_rows = 0, 0
    cols, drift_cols = {}, {}
    col_examples, drift_col_examples = {}, {}
    # sorted() rather than the set's own order: which guests end up as the
    # mail's named examples must not move between two runs over identical
    # data, or someone comparing this morning's mail against yesterday's sees
    # a change that never happened.
    for k in sorted(paired):
        o, s = ours_ix[k], sheet_ix[k]
        real, drift = [], []
        for key in columns:
            ov, sv = _norm(o.get(key)), _norm(s.get(key))
            if ov == sv:
                continue
            (drift if _is_known_drift(key, ov, sv) else real).append((key, ov, sv))
        who = _row_label(o, kind)
        for key, ov, sv in drift:
            drift_cols[key] = drift_cols.get(key, 0) + 1
            ex = drift_col_examples.setdefault(key, [])
            if len(ex) < _EXAMPLES_PER_COLUMN:
                ex.append((who, ov, sv))
        for key, ov, sv in real:
            cols[key] = cols.get(key, 0) + 1
            ex = col_examples.setdefault(key, [])
            if len(ex) < _EXAMPLES_PER_COLUMN:
                ex.append((who, ov, sv))
        if real:
            diff_rows += 1
        elif drift:
            drift_rows += 1

    # Named, not just counted. A morning that reads "3 rows only in the sheet"
    # tells nobody which three, and the answer is not derivable from the mail;
    # with the names in hand the same reader can open those guests in MEWS
    # before the register is filed.
    only_ours_keys = sorted(ours_ix.keys() - sheet_ix.keys())
    only_sheet_keys = sorted(sheet_ix.keys() - ours_ix.keys())
    return {
        "ours": len(ours),
        "sheet": len(sheet),
        "paired": len(paired),
        "only_ours": len(only_ours_keys),
        "only_sheet": len(only_sheet_keys),
        "only_ours_rows": [_row_label(ours_ix[k], kind)
                           for k in only_ours_keys[:_EXAMPLES_PER_MISSING_GROUP]],
        "only_sheet_rows": [_row_label(sheet_ix[k], kind)
                            for k in only_sheet_keys[:_EXAMPLES_PER_MISSING_GROUP]],
        "clean_rows": len(paired) - diff_rows - drift_rows,
        "diff_rows": diff_rows,
        "drift_rows": drift_rows,
        "cols": cols,
        "drift_cols": drift_cols,
        # There used to be a separate `samples` list here - up to 3 whole rows
        # with up to 4 of their differing columns each - feeding its own table.
        # It is gone rather than kept alongside col_examples, because two
        # parallel example paths can name different guests for the same
        # difference, and it could never carry a known-drift row at all (it was
        # only appended inside `if real:`).
        "col_examples": col_examples,
        "drift_col_examples": drift_col_examples,
    }


def _parse_sheet(content: bytes) -> dict:
    """One workbook's RR4 and TM30 tabs, plus the window each was exported
    over. The Master tab is the only place every sheet agrees on: A2 is the
    ImportInhouse (RR4) start and B2 the ImportCP (TM30) start. Chinatown also
    carries Parameter-* tabs, but the other five don't."""
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    ms = wb["Master"]
    rr4_start, tm30_start = ms.cell(2, 1).value, ms.cell(2, 2).value
    date = rr4_start.strftime("%Y-%m-%d") if isinstance(rr4_start, datetime) else None

    # RR4 tab: row 3 is the Thai header, row 4 the English field keys, data
    # from row 5. Columns are located BY that field-key row rather than by
    # position, so a sheet that gains a column doesn't silently shift the
    # whole comparison one to the left.
    from app.services.sync_service import sync_service
    rr4_rows = []
    if "RR4" in wb.sheetnames:
        ws = wb["RR4"]
        grid = [[c.value for c in row] for row in ws.iter_rows(min_row=4, max_col=27)]
        keys = {_norm(v): i for i, v in enumerate(grid[0])} if grid else {}
        by_key = {field: keys.get(field) for _k, _l, field in sync_service._RR4_COLUMNS}
        for row in grid[1:]:
            if _norm(row[0]) == "":
                continue
            rr4_rows.append({
                key: (row[by_key[field]] if by_key.get(field) is not None else None)
                for key, _label, field in sync_service._RR4_COLUMNS
            })

    # TM30 tab: one header row of the government form's own bilingual labels,
    # then data. Nine fixed columns in the export's own order - matched
    # positionally because those labels carry embedded newlines and stray
    # spaces that differ between sheets.
    tm30_rows = []
    if "TM30" in wb.sheetnames:
        ws = wb["TM30"]
        for row in ws.iter_rows(min_row=2, max_col=len(sync_service._TM30_COLUMNS), values_only=True):
            if _norm(row[0]) == "":
                continue
            tm30_rows.append({key: row[i] for i, (key, _label) in enumerate(sync_service._TM30_COLUMNS)})

    wb.close()
    return {
        "date": date,
        "rr4_rows": rr4_rows,
        "tm30_rows": tm30_rows,
        "rr4_window": rr4_start.strftime("%H:%M") if isinstance(rr4_start, datetime) else None,
        "tm30_window": tm30_start.strftime("%H:%M") if isinstance(tm30_start, datetime) else None,
    }


async def _fetch_sheets() -> dict:
    """All six workbooks over their public export URL, concurrently. One that
    fails comes back as an error on that property's row rather than failing
    the run - five comparable properties still beat none."""
    async def one(prop, sheet_id):
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        try:
            async with httpx.AsyncClient(follow_redirects=True, timeout=120) as client:
                r = await client.get(url)
                r.raise_for_status()
            return prop, _parse_sheet(r.content), None
        except Exception as e:
            logger.warning(f"RR4 compare: could not read {prop}'s sheet: {e}")
            return prop, None, str(e)

    results = await asyncio.gather(*(one(p, sid) for p, (_, sid) in SHEETS.items()))
    return {p: {"data": d, "error": e} for p, d, e in results}


def _windows() -> dict:
    """Each property's configured RR4 day window, to be shown next to the one
    the sheet was actually exported over. A stale value here silently
    undercounts the register - a leftover 14:00/12:00 on Chinatown once
    produced 160 rows where the real answer was 241 - and the sheets change
    their window without warning, so this is worth a daily look."""
    from app.services.sync_service import sync_service
    try:
        res = sync_service.supabase.table("property_api_settings").select(
            "property_name, rr4_tm30_day_start_hour, rr4_tm30_day_start_minute").execute()
        return {r["property_name"]: f"{r.get('rr4_tm30_day_start_hour') or 0:02d}:"
                                    f"{r.get('rr4_tm30_day_start_minute') or 0:02d}"
                for r in (res.data or [])}
    except Exception as e:
        logger.warning(f"RR4 compare: could not read property windows: {e}")
        return {}


async def _tm30_windows() -> dict:
    """TM30's own configured window per property - a separate setting from
    RR4's, so the mail has to ask for it separately too. Read through
    _resolve_tm30_day_start rather than the table directly, so this column
    can never disagree with the one the register was actually built on."""
    from app.services.sync_service import sync_service
    out = {}
    for prop in SHEETS:
        try:
            h, m = await sync_service._resolve_tm30_day_start(prop)
            out[prop] = f"{h:02d}:{m:02d}"
        except Exception as e:
            logger.warning(f"RR4 compare: could not read {prop}'s TM30 window: {e}")
    return out


def _local(ts: str) -> str:
    """A stored UTC timestamp as Bangkok wall-clock. Every property here is in
    Thailand, so there is only the one offset to apply."""
    if not ts:
        return ""
    ts = re.sub(r"\.(\d+)", lambda m: "." + m.group(1)[:6].ljust(6, "0"), ts)
    return datetime.fromisoformat(ts).astimezone(timezone(timedelta(hours=7))).strftime("%d %b %H:%M")


async def build_comparison(want_date: str = None) -> dict:
    """The whole check, as data.

    `want_date` (the CLI's optional argument) pins every property to one date
    instead of each following its own sheet - useful for reproducing a past
    run, and it simply reports "the sheet holds a different date" for any property whose sheet
    has since moved on, because these workbooks hold one pasted export each
    and a day they no longer hold cannot be reconstructed.
    """
    from app.routers.rr4 import read_managed_day

    sheets = await _fetch_sheets()
    windows = _windows()
    tm30_windows = await _tm30_windows()
    props = []

    for prop, (short, _sid) in SHEETS.items():
        row = {"property": prop, "short": short, "date": None, "status": "error",
               "note": "", "rr4": None, "tm30": None, "synced_at": "",
               "sheet_rr4_window": "", "sheet_tm30_window": "",
               "our_window": windows.get(prop, ""),
               "our_tm30_window": tm30_windows.get(prop, "")}
        sh = sheets[prop]["data"]
        if not sh:
            row["note"] = f"Could not read sheet — {sheets[prop]['error']}"
            props.append(row)
            continue

        row["date"] = sh["date"]
        row["sheet_rr4_window"] = sh["rr4_window"] or ""
        row["sheet_tm30_window"] = sh["tm30_window"] or ""
        if not sh["date"]:
            row["note"] = "Sheet has no date in Master"
            props.append(row)
            continue
        if want_date and want_date != sh["date"]:
            row["status"] = "other_date"
            row["note"] = f"Sheet holds {sh['date']}, not {want_date}"
            props.append(row)
            continue

        try:
            payload = read_managed_day(prop, sh["date"])
        except Exception as e:
            row["note"] = f"Could not read rr4_tm30_sync — {e}"
            props.append(row)
            continue
        if not payload:
            row["status"] = "missing"
            row["note"] = f"{sh['date']} has not been imported yet"
            props.append(row)
            continue

        from app.services.sync_service import sync_service
        rr4_cols = [k for k, _l, _f in sync_service._RR4_COLUMNS if k not in _SKIP_RR4_COLUMNS]
        tm30_cols = [k for k, _l in sync_service._TM30_COLUMNS]
        row["status"] = "ok"
        row["synced_at"] = _local(payload.get("_synced_at") or "")
        row["rr4"] = _compare_rows((payload.get("rr4") or {}).get("rows", []),
                                   sh["rr4_rows"], "rr4", rr4_cols)
        row["tm30"] = _compare_rows((payload.get("tm30") or {}).get("rows", []),
                                    sh["tm30_rows"], "tm30", tm30_cols)
        props.append(row)

    compared = [p for p in props if p["status"] == "ok"]
    if not compared:
        return {"status": "no_data", "properties": props, "date": want_date}

    # The headline date is whichever day most properties are sitting on -
    # Chinatown's later cutoff routinely leaves it a day behind the other
    # five, and naming its date in the subject line would misdescribe the mail.
    counts = {}
    for p in compared:
        counts[p["date"]] = counts.get(p["date"], 0) + 1
    date = max(counts, key=lambda d: (counts[d], d))

    totals = {}
    for kind in ("rr4", "tm30"):
        totals[kind] = {
            f: sum(p[kind][f] for p in compared)
            for f in ("ours", "sheet", "paired", "only_ours", "only_sheet",
                      "clean_rows", "diff_rows", "drift_rows")
        }
    return {
        "status": "ok",
        "date": date,
        "mixed_dates": len(counts) > 1,
        "properties": props,
        "compared": len(compared),
        "totals": totals,
    }


def _title(result: dict) -> str:
    day = datetime.strptime(result["date"], "%Y-%m-%d")
    # "%-d" (unpadded day) is a glibc extension - it raises ValueError on
    # Windows, which made this whole mail impossible to preview from a dev
    # machine. Formatting the day separately is portable and identical.
    return f"RR4/TM30 {day.day} {day.strftime('%b %Y')}"


def render_text(result: dict) -> str:
    """Plain-text form - the CLI output, and the email's text/plain part."""
    if result["status"] != "ok":
        out = ["⛔ Not comparable yet, for any property:"]
        for p in result["properties"]:
            out.append(f"     {p['short']:<12} {p['note']}")
        return "\n".join(out)

    out = ["=" * 88, f"Comparison Summary — {_title(result)}", "=" * 88]
    if result["mixed_dates"]:
        out.append("(Each property is compared at its own sheet's date — Chinatown cuts its day at 12:15, so it runs a day behind the rest)")

    # The same three sections the HTML mail is built from, in the same order,
    # so the text/plain part of a message can never describe a different check
    # from the part most people actually read.
    out += ["", "1. EVERY PROPERTY — Google Sheet / NHGOne", "-" * 88]
    out.append(f"{'Property':<12}{'Date':<12}{'RR4 sheet/ours':<17}{'Diff':<7}"
               f"{'TM30 sheet/ours':<17}Diff")
    for p in result["properties"]:
        if p["status"] != "ok":
            out.append(f"{p['short']:<12}{(p['date'] or '—'):<12}{p['note']}")
            continue
        r, t = p["rr4"], p["tm30"]
        r_count = "{}/{}".format(r["sheet"], r["ours"])
        t_count = "{}/{}".format(t["sheet"], t["ours"])
        # Same rule as the HTML cell: the guest COUNT decides the mark, and
        # what still differs inside matching-sized registers is section 2's.
        r_gap = r["sheet"] - r["ours"]
        t_gap = t["sheet"] - t["ours"]
        out.append(
            f"{p['short']:<12}{p['date']:<12}"
            f"{r_count:<17}{('✓' if r_gap == 0 else f'✗ {r_gap:+d}'):<7}"
            f"{t_count:<17}{'✓' if t_gap == 0 else f'✗ {t_gap:+d}'}")
    out.append("-" * 88)

    tr, tt = result["totals"]["rr4"], result["totals"]["tm30"]
    out.append(f"RR4  total {tr['sheet']}/{tr['ours']} rows · paired {tr['paired']} · fully matched "
               f"{tr['clean_rows']} · real diff {tr['diff_rows']} · known drift {tr['drift_rows']} · "
               f"only in sheet {tr['only_sheet']} · only in NHGOne {tr['only_ours']}")
    out.append(f"TM30 total {tt['sheet']}/{tt['ours']} rows · paired {tt['paired']} · fully matched "
               f"{tt['clean_rows']} · real diff {tt['diff_rows']} · known drift {tt['drift_rows']} · "
               f"only in sheet {tt['only_sheet']} · only in NHGOne {tt['only_ours']}")

    out += ["", "2. WHAT DIFFERS (sheet / NHGOne)", "-" * 88]
    groups = _diff_groups(result)
    if not groups:
        out.append("  ✅ Every column of every paired row matches, and both sides hold the same guests")
    for g in groups:
        tag = {"real": "!!", "expected": "..", "drift": "  "}[g["tone"]]
        shown = "" if len(g["ex"]) >= g["n"] else f" (showing {len(g['ex'])})"
        out.append(f"{tag} {g['short']:<12} {g['reg']:<5} {g['what']}  "
                   f"× {g['n']}{shown}  — {g['why']}")
        for who, ours, sheet in g["ex"]:
            out.append(f"       {_clip(who, 46):<48} {_clip(sheet, 30) or '—'}  /  "
                       f"{_clip(ours, 30) or '—'}")
    if groups:
        out.append("   (!! needs review today · .. expected from a configured window · "
                   "blank = known drift, already explained)")

    out += ["", "3. WHEN EACH SIDE PULLED ITS DATA", "-" * 88]
    for p in result["properties"]:
        bad = (p["sheet_rr4_window"] != p["our_window"]
               or p["sheet_tm30_window"] != p["our_tm30_window"])
        flag = "   ⚠️ mismatch" if bad else ""
        out.append(f"  {p['short']:<12} RR4 sheet {p['sheet_rr4_window'] or '—':<7} ours "
                   f"{p['our_window'] or '—':<7} · TM30 sheet {p['sheet_tm30_window'] or '—':<7} ours "
                   f"{p['our_tm30_window'] or '—':<7} · built {p['synced_at'] or '—':<13}{flag}")
    out.append("   Each side sweeps a 24-hour day that STARTS at these times — both halves of a "
               "pair must match.")
    return "\n".join(out)


def _esc(s) -> str:
    """Guest names, passport numbers and addresses come from MEWS, i.e. from
    whatever a guest typed at check-in - they are never trusted as markup."""
    return (str(s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


_TD = "padding:6px 10px;border:1px solid #e2e8f0;font-size:13px;"
_TH = "padding:6px 10px;border:1px solid #e2e8f0;font-size:11px;font-weight:700;background:#f8fafc;text-align:left;"
_MUTED = "color:#94a3b8;"

# Three states, three colours, used consistently by all three tables so a
# colour means the same thing wherever it appears:
#   green  the sheet and NHGOne agree - nothing to do
#   red    a real difference - somebody has to look at it today
#   amber  a difference that has already been explained (_KNOWN_DRIFT, or a
#          shortfall the property's own configured window is supposed to
#          produce) - worth seeing, not worth chasing
# Amber was the only highlight before this; splitting it means the daily
# known-drift rows stop competing for attention with the rare real ones.
_OK = "color:#166534;font-weight:700;"
_BAD = "background:#fee2e2;color:#b91c1c;font-weight:700;"
_EXPECTED = "background:#fef3c7;color:#92400e;font-weight:700;"

# An em dash as a name rather than inline: an f-string expression cannot
# contain a backslash before Python 3.12, and this file has to import on the
# 3.10 that runs it locally as well as the 3.12 on Vercel.
_DASH = "—"
_TICK = "✓"
_CROSS = "✗"


_TABLE_OPEN = '<table style="border-collapse:collapse">'


def _scroll(parts: list, footnote: str) -> str:
    """Every table here is wider than a phone, so each one scrolls inside its
    own box rather than forcing the email body to - the same wrapper
    st_compare_service.render_grid_table uses, and the reason _TABLE_OPEN
    carries no width:100%: a table pinned to 100% of an overflow-x box can
    never overflow it, so the cells squash instead of the box scrolling.

    Outlook desktop (the Word rendering engine) ignores overflow-x entirely
    and widens the card instead. That is survivable and is why the tables stay
    as narrow as they do; it is not a reason to add columns.
    """
    return ('<div style="overflow-x:auto">' + "".join(parts) + "</div>"
            + f'<p style="font-size:11px;color:#94a3b8;margin:6px 0 0">{footnote}</p>')


def _clip(value: str, limit: int = 60) -> str:
    """MEWS free text can be long - an RR4 `address` is a whole postal address -
    and one long value in a cell widens the whole card in the clients that
    ignore overflow-x. Cut it here rather than trusting a CSS property Outlook
    doesn't implement."""
    s = str(value or "")
    return s if len(s) <= limit else s[:limit - 1].rstrip() + "…"


def _dmy_display(iso: str) -> str:
    """The sheets' own ISO date as DD/MM/YYYY, so the cells agree with the
    <<Date>> token in the line above them instead of printing 2026-09-02
    beside 02/09/2026."""
    try:
        return datetime.strptime(iso, "%Y-%m-%d").strftime("%d/%m/%Y")
    except (TypeError, ValueError):
        return iso or _DASH


def _summary_cell(block: dict, window: str = "") -> str:
    """One register's cell in table 1: a green "✓ N" when the sheet and NHGOne
    hold the SAME NUMBER OF GUESTS, a red "✗ sheet / ours" when they don't.

    The count is the whole test for the colour, by request. Anything else that
    differs - a column whose value disagrees, a guest one side holds and the
    other doesn't, known drift - is reported in grey inside the green cell and
    listed row by row in table 2, but does not turn it red. Two registers of
    the same size are the thing the filing is judged on; which of their rows
    still need a look is the next table's job.

    Sheet first, then ours, in that order everywhere in this mail - it is the
    sheet that is the ground truth being checked against, so it reads as
    "what should be there / what we produced".

    `window` is the property's configured TM30 start. A non-midnight one files
    a shorter day than the sheet holds ON PURPOSE (Chinatown's 12:15 drops
    every guest arriving before noon - 2 to 20 of them on each of the seven
    days measured to 29-Aug-2026), so that shortfall is annotated as the
    configured consequence it is. It still shows red, because it is a real
    difference in the number of guests filed and, unlike _KNOWN_DRIFT, cannot
    be verified row by row - TM30 carries no check-in column to test each
    missing guest against, so a genuine new miss would look identical.
    """
    # Deliberately "is our window non-midnight", NOT "does our window differ
    # from the sheet's". The sheets DECLARE a TM30 window in Master!B2 that
    # they do not actually filter by: on 02-Sep-2026 Chinatown's sheet said
    # 12:15, exactly what we are configured to, and still held 32 arrivals to
    # our 26 - and Patong's said 02:05 to our 02:05, 42 to our 39. Comparing
    # the two windows would therefore find them equal and call both shortfalls
    # unexplained every single morning. Our own non-midnight start is the real
    # cause (an earlier check found midnight reproduces Chinatown's sheet
    # exactly, 67 for 67, where 12:15 gives 56), so that is what is tested.
    shifted = bool(window) and window != "00:00"

    bits = []
    if block["diff_rows"]:
        bits.append(f"{block['diff_rows']} differ")
    if block["only_ours"]:
        bits.append(f"{block['only_ours']} only in NHGOne")
    if block["only_sheet"]:
        bits.append(f"{block['only_sheet']} only in the sheet"
                    + (f", as our {window} window intends" if shifted else ""))
    if block["drift_rows"]:
        bits.append(f"{block['drift_rows']} known drift")

    # THE HEADLINE IS THE GUEST COUNT, and only the guest count. A register
    # holding the same number of guests as the sheet reads green even when
    # some of those rows still disagree on a column - by request, and because
    # a count that matches is the thing the filing is judged on. Whatever
    # still differs is not hidden: it is spelled out in grey right here and
    # listed row by row in the next table.
    #
    # The two are genuinely independent. Patong on 06-Sep-2026 held exactly
    # 233 RR4 rows against the sheet's 233 while three of them differed on a
    # column and one guest existed on each side that the other did not - the
    # unpaired pair cancels in the total.
    if block["sheet"] == block["ours"]:
        note = ""
        if bits:
            note = (f'<span style="{_MUTED}font-weight:400"> '
                    f'({", ".join(bits)})</span>')
        return f'<td style="{_TD}{_OK}">{_TICK} {block["sheet"]}{note}</td>'
    # The counts stay on one line; the note after them is allowed to wrap, or
    # "6 only in the sheet, as our 12:15 window intends" pushes the table past
    # the card in every client that honours nowrap.
    return (f'<td style="{_TD}{_BAD}">'
            f'<span style="white-space:nowrap">{_CROSS} {block["sheet"]} / {block["ours"]}</span>'
            f'<span style="font-weight:400;font-size:11px"> ({", ".join(bits)})</span></td>')


def render_summary_table(result: dict) -> str:
    """TABLE 1 - every property, both registers, Google Sheet / NHGOne with a
    green tick or a red cross. The <<SummaryTable>> token, and the one table
    that has to answer "is anything wrong this morning?" on its own."""
    if result["status"] != "ok":
        return f'<pre style="font-family:ui-monospace,monospace;font-size:12px">{render_text(result)}</pre>'

    h = [f'{_TABLE_OPEN}<tr>'
         f'<th style="{_TH}">Property</th><th style="{_TH}">Date</th>'
         f'<th style="{_TH}">RR4 &mdash; Google Sheet / NHGOne</th>'
         f'<th style="{_TH}">TM30 &mdash; Google Sheet / NHGOne</th></tr>']
    for p in result["properties"]:
        h.append(f'<tr><td style="{_TD}font-weight:600;white-space:nowrap">{p["short"]}</td>')
        if p["status"] != "ok":
            h.append(f'<td style="{_TD}" colspan="3">'
                     f'<span style="{_BAD}padding:2px 6px;border-radius:4px">'
                     f'{_esc(p["note"])}</span></td></tr>')
            continue
        h.append(f'<td style="{_TD}{_MUTED}white-space:nowrap">{_dmy_display(p["date"])}</td>')
        h.append(_summary_cell(p["rr4"]))
        h.append(_summary_cell(p["tm30"], p.get("our_tm30_window", "")))
        h.append("</tr>")

    tr, tt = result["totals"]["rr4"], result["totals"]["tm30"]
    h.append(f'<tr style="background:#f1f5f9"><td style="{_TD}font-weight:700">Total</td>'
             f'<td style="{_TD}{_MUTED}">{result["compared"]} properties</td>')
    h.append(_summary_cell(tr))
    h.append(_summary_cell(tt))
    h.append("</tr></table>")
    return _scroll(
        h,
        f'{_TICK} the sheet and NHGOne hold the <b>same number of guests</b> &nbsp;·&nbsp; '
        f'{_CROSS} that number differs, shown as <b>Google Sheet / NHGOne</b>. '
        f'Anything still differing inside rows that DO line up is noted in grey '
        f'and listed in full in the next table. '
        f'Rows are paired by passport/ID number, then every column of each paired row is compared '
        f'(except the row number, which both sides renumber on their own). '
        f'Rows with no name (a MEWS-booked slot not yet linked to a guest profile) are counted here '
        f'because the sheet keeps them too, but are dropped from the filed .xlsx.')


def _diff_groups(result: dict) -> list:
    """Every difference in the whole comparison as one group each, ordered so
    the ones somebody has to act on today come first.

    A group is one (property, register, thing that differs) with up to a few
    named guests under it. Three kinds go in, and they used to live in two
    separate tables plus a summary note:
      real      a column whose value differs, or a guest one side has and the
                other doesn't - nobody has explained these yet
      expected  a guest the sheet holds and we don't BECAUSE of the property's
                own configured TM30 window - the shortfall the setting is for
      drift     _KNOWN_DRIFT: already chased down, already explained
    """
    groups = []
    for p in result["properties"]:
        if p["status"] != "ok":
            continue
        window = p.get("our_tm30_window", "")
        shifted = bool(window) and window != "00:00"
        for kind, label in (("rr4", "RR4"), ("tm30", "TM30")):
            b = p[kind]
            for key, n in sorted(b["cols"].items(), key=lambda kv: -kv[1]):
                groups.append({
                    "short": p["short"], "reg": label, "what": key, "n": n,
                    "ex": b["col_examples"].get(key, []),
                    "why": "Needs review", "tone": "real",
                })
            if b["only_sheet"]:
                why = "In the sheet but not in our register — needs review"
                tone = "real"
                if kind == "tm30" and shifted:
                    why = (f"Expected — our TM30 day starts at {window}, so a guest arriving "
                           f"before that is filed on the previous day, not this one")
                    tone = "expected"
                groups.append({
                    "short": p["short"], "reg": label,
                    "what": "Guest only in the sheet", "n": b["only_sheet"],
                    "ex": [(who, "missing", "present") for who in b["only_sheet_rows"]],
                    "why": why, "tone": tone,
                })
            if b["only_ours"]:
                groups.append({
                    "short": p["short"], "reg": label,
                    "what": "Guest only in NHGOne", "n": b["only_ours"],
                    "ex": [(who, "present", "missing") for who in b["only_ours_rows"]],
                    "why": "In our register but not in the sheet — needs review",
                    "tone": "real",
                })
            for key, n in sorted(b["drift_cols"].items(), key=lambda kv: -kv[1]):
                groups.append({
                    "short": p["short"], "reg": label, "what": key, "n": n,
                    "ex": b["drift_col_examples"].get(key, []),
                    "why": _KNOWN_DRIFT.get(key, "Known drift"), "tone": "drift",
                })
    rank = {"real": 0, "expected": 1, "drift": 2}
    # Stable, so within a tone the properties keep SHEETS' own order.
    groups.sort(key=lambda g: rank[g["tone"]])
    return groups


def render_column_table(result: dict) -> str:
    """TABLE 2 - the detail behind every red and amber cell in table 1, the
    <<ColumnTable>> token.

    This is the old ColumnTable ("nationality differs on 2 rows") and
    SampleTable ("Nikolaos Pantotis: GRC vs GRL") merged into one, plus the
    guests that only one side holds, which were previously counted in table 1
    and then named nowhere - so a morning like 02-Sep-2026, whose ONLY
    differences were 6 Chinatown and 3 Patong TM30 guests missing on our side,
    showed red cells above a detail table that said everything matched.
    """
    if result["status"] != "ok":
        return ""

    groups = _diff_groups(result)
    if not groups:
        return ('<p style="font-size:13px;color:#166534;margin:0">'
                f'{_TICK} Every column of every paired row matches the sheet, '
                'and both sides hold exactly the same guests</p>')

    tone_style = {"real": "color:#b91c1c;font-weight:700;",
                  "expected": "color:#92400e;font-weight:700;",
                  "drift": _MUTED}
    h = [f'{_TABLE_OPEN}<tr>'
         f'<th style="{_TH}">Property</th><th style="{_TH}">Register</th>'
         f'<th style="{_TH}">What differs</th><th style="{_TH}">Guest</th>'
         f'<th style="{_TH}">Google Sheet</th><th style="{_TH}">NHGOne</th>'
         f'<th style="{_TH}">Why</th></tr>']
    for g in groups:
        ex = g["ex"] or [(_DASH, "", "")]
        span = f' rowspan="{len(ex)}"' if len(ex) > 1 else ""
        style = tone_style[g["tone"]]
        shown = "" if len(ex) >= g["n"] else f' (showing {len(ex)})'
        for i, (who, ours, sheet) in enumerate(ex):
            h.append("<tr>")
            if i == 0:
                h.append(
                    f'<td style="{_TD}white-space:nowrap"{span}>{g["short"]}</td>'
                    f'<td style="{_TD}{_MUTED}"{span}>{g["reg"]}</td>'
                    f'<td style="{_TD}{style}"{span}>{_esc(g["what"])}'
                    f'<div style="font-weight:400;font-size:11px;color:#64748b">'
                    f'{g["n"]} row{"s" if g["n"] != 1 else ""}{shown}</div></td>')
            h.append(f'<td style="{_TD}">{_esc(_clip(who, 46))}</td>'
                     f'<td style="{_TD}{_MUTED}">{_esc(_clip(sheet)) or _DASH}</td>'
                     f'<td style="{_TD}font-weight:700">{_esc(_clip(ours)) or _DASH}</td>')
            if i == 0:
                h.append(f'<td style="{_TD}font-size:11px;{style}"{span}>{g["why"]}</td>')
            h.append("</tr>")
    h.append("</table>")
    return _scroll(
        h,
        f'<b style="color:#b91c1c">Red</b> needs review today. '
        f'<b style="color:#92400e">Amber</b> has already been explained — either known drift '
        f'(the sheet and MEWS moving on after the export) or the shortfall a property\'s own '
        f'configured window is meant to produce. '
        f'Up to {_EXAMPLES_PER_COLUMN} guests are named per column and '
        f'{_EXAMPLES_PER_MISSING_GROUP} per missing-guest group; '
        f'the row count beside each is always the full number.')


def render_sample_table(result: dict) -> str:
    """Retained deliberately, and empty. The example rows this used to render
    are part of ColumnTable now. An Admin template saved before that change
    still carries <<SampleTable>>, and an unknown token is left in the body
    verbatim - so this keeps that template rendering nothing there instead of
    the literal text "<<SampleTable>>" in somebody's mail."""
    return ""


def render_window_table(result: dict) -> str:
    """TABLE 3 - when each side pulled its data, the <<WindowTable>> token.

    Both sides sweep a 24-hour day that STARTS at the time in these columns,
    so a pair that disagrees means the two registers are counting different
    guests, whatever the row counts in table 1 happen to say. See _windows()
    for why that is worth a look every single morning; the last column is when
    our own import actually ran.
    """
    h = [f'{_TABLE_OPEN}<tr>'
         f'<th style="{_TH}">Property</th>'
         f'<th style="{_TH}">RR4 &mdash; Google Sheet</th><th style="{_TH}">RR4 &mdash; NHGOne</th>'
         f'<th style="{_TH}">TM30 &mdash; Google Sheet</th><th style="{_TH}">TM30 &mdash; NHGOne</th>'
         f'<th style="{_TH}">NHGOne built the file</th></tr>']
    for p in result["properties"]:
        ok = p["sheet_rr4_window"] == p["our_window"] and p["sheet_rr4_window"]
        tm_ok = p["sheet_tm30_window"] == p["our_tm30_window"] and p["sheet_tm30_window"]
        h.append(f'<tr><td style="{_TD}white-space:nowrap">{p["short"]}</td>'
                 f'<td style="{_TD}{_MUTED}">{p["sheet_rr4_window"] or _DASH}</td>'
                 f'<td style="{_TD}{_OK if ok else _BAD}">'
                 f'{_TICK if ok else _CROSS} {p["our_window"] or _DASH}</td>'
                 f'<td style="{_TD}{_MUTED}">{p["sheet_tm30_window"] or _DASH}</td>'
                 f'<td style="{_TD}{_OK if tm_ok else _BAD}">'
                 f'{_TICK if tm_ok else _CROSS} {p["our_tm30_window"] or _DASH}</td>'
                 f'<td style="{_TD}{_MUTED}white-space:nowrap">{p["synced_at"] or _DASH}</td></tr>')
    h.append("</table>")
    return _scroll(
        h,
        'Each side sweeps a 24-hour day that <b>starts</b> at the time shown, so both halves of a '
        'pair have to match or the two registers are counting different guests. '
        'The sheets change these windows without warning, and a stale value on our side silently '
        'undercounts the register. TM30 has its own setting per property, separate from RR4\'s — '
        'Chinatown\'s 12:15 is deliberate. The last column is when our own import ran.')


def render_tokens(result: dict) -> dict:
    """Everything the email template can substitute."""
    t = result.get("totals") or {"rr4": {}, "tm30": {}}
    day = datetime.strptime(result["date"], "%Y-%m-%d") if result.get("date") else None
    return {
        "Date": day.strftime("%d/%m/%Y") if day else _DASH,
        "PropertyCount": str(result.get("compared", 0)),
        "Rr4Diff": str(t["rr4"].get("diff_rows", _DASH)),
        "Tm30Diff": str(t["tm30"].get("diff_rows", _DASH)),
        # Sheet first, then ours - the same order every table in this mail
        # reads in, so the header line and the tables can't contradict.
        "Rr4Rows": f"{t['rr4'].get('sheet', _DASH)} / {t['rr4'].get('ours', _DASH)}",
        "Tm30Rows": f"{t['tm30'].get('sheet', _DASH)} / {t['tm30'].get('ours', _DASH)}",
        "SummaryTable": render_summary_table(result),
        "ColumnTable": render_column_table(result),
        "SampleTable": render_sample_table(result),
        "WindowTable": render_window_table(result),
    }


def subject_summary(result: dict) -> str:
    if result["status"] != "ok":
        return "not comparable yet"
    t = result["totals"]
    bad = t["rr4"]["diff_rows"] + t["tm30"]["diff_rows"] \
        + t["rr4"]["only_ours"] + t["rr4"]["only_sheet"] \
        + t["tm30"]["only_ours"] + t["tm30"]["only_sheet"]
    return "matches sheet completely" if bad == 0 else f"{bad} rows need review"
