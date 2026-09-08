#!/usr/bin/env python
"""Find the ST Arrivals rule by search rather than by guessing one at a time.

Every rule tried so far was inferred from a handful of reservations on one
day, shipped, and then contradicted by the next day's data. This scores a
whole family of candidate rules at once against the ONLY ground truth there
is - each property's own "<Name>-ST" sheet - and does it per SPACE CATEGORY,
not per total, so a rule cannot pass by having two errors cancel out.

    .venv/Scripts/python.exe scripts/st_arrival_sweep.py 2026-09-07

Each candidate is run through the real get_st_files_report via the
_ST_ARRIVAL_RULE seam, so it inherits that function's actual category
resolution, parent/child space expansion and unit counting. Scoring a rule
against a re-implementation of those is exactly how the current rule ended up
matching numbers that were not the report's own.

Only properties whose sheet currently holds the requested date are scored;
the workbooks hold one pasted export each, so the rest simply cannot be
compared and are reported as skipped rather than silently dropped.
"""
import asyncio
import io
import os
import sys
from collections import Counter
from pathlib import Path

import httpx
from openpyxl import load_workbook

_API = Path(__file__).resolve().parent.parent / "api"
os.chdir(_API)
sys.path.insert(0, str(_API))

from app.services import sync_service as ss_mod              # noqa: E402
from app.services.mews_client import mews_client             # noqa: E402
from app.services.st_compare_service import SHEETS           # noqa: E402
from app.services.sync_service import sync_service           # noqa: E402

# Every candidate rule rebuilds the same day for the same property, and the
# MEWS half of that is identical every time - only the arrival predicate
# changes. Without this the sweep is ~20 rules x 8 properties x 7 calls of
# entirely redundant traffic against a live PMS. Memoised on the exact
# (endpoint, payload, property) triple, so a genuinely different request
# still goes out.
_CACHE = {}
_real_post = mews_client.post


async def _cached_post(endpoint, payload=None, property_name=None, **kw):
    key = (endpoint, property_name, repr(sorted((payload or {}).items(), key=str)))
    if key not in _CACHE:
        _CACHE[key] = await _real_post(endpoint, payload, property_name=property_name, **kw)
    return _CACHE[key]


mews_client.post = _cached_post

DATE = sys.argv[1] if len(sys.argv) > 1 else None
if not DATE:
    sys.exit("usage: st_arrival_sweep.py YYYY-MM-DD")


def sheet_tabs(sheet_id):
    """The sheet's own Arrivals/Departures per category, plus the date it is
    currently holding."""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    r = httpx.get(url, follow_redirects=True, timeout=180)
    r.raise_for_status()
    wb = load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb["Master"]
    date = None
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip() == "Date":
            v = ws.cell(row, 2).value
            date = v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else str(v)
    out = {}
    for tab in ("Arrivals", "Departures"):
        if tab not in wb.sheetnames:
            continue
        w = wb[tab]
        counts = {}
        for row in range(2, w.max_row + 1):
            svc, cat, val = w.cell(row, 1).value, w.cell(row, 2).value, w.cell(row, 3).value
            if not cat or str(svc).strip().lower() == "total":
                continue
            counts[str(cat).strip()] = int(val or 0)
        out[tab] = counts
    return date, out


# --------------------------------------------------------------- candidates
#
# Each rule answers one question: does THIS reservation count as an arrival on
# the day being built? The signature is what the seam in get_st_files_report
# passes. `actual` is ActualStartUtc or None.

def _hour(ts, parse_utc, tz):
    t = parse_utc(ts)
    return t.astimezone(tz).hour if t else None


def _walk_in(res, rates_by_id):
    name = (rates_by_id.get(res.get("RateId"), {}) or {}).get("Name") or ""
    return bool(ss_mod._ST_WALK_IN_RATE_RE.search(name))


def make_rules():
    rules = {}

    def add(name, fn):
        rules[name] = fn

    # --- which instant decides the day, with no day-use handling at all
    add("sched, no day-use rule",
        lambda res, actual, in_window, pu, tz, rates, ds, de:
            in_window(res.get("StartUtc")))
    add("actual-pref, no day-use rule",
        lambda res, actual, in_window, pu, tz, rates, ds, de:
            in_window(actual or res.get("StartUtc")))
    add("union, no day-use rule",
        lambda res, actual, in_window, pu, tz, rates, ds, de:
            in_window(res.get("StartUtc")) or in_window(actual))

    # --- scheduled instant + a night-tail cutoff on zero-night stays
    for h in (2, 3, 4, 6):
        add(f"sched, drop day-use before {h:02d}:00",
            lambda res, actual, in_window, pu, tz, rates, ds, de, h=h:
                in_window(res.get("StartUtc"))
                and not (in_window(res.get("EndUtc"))
                         and (_hour(res.get("StartUtc"), pu, tz) or 0) < h))
        add(f"sched, drop day-use before {h:02d}:00 unless walk-in",
            lambda res, actual, in_window, pu, tz, rates, ds, de, h=h:
                in_window(res.get("StartUtc"))
                and not (in_window(res.get("EndUtc"))
                         and not _walk_in(res, rates)
                         and (_hour(res.get("StartUtc"), pu, tz) or 0) < h))

    # --- actual-preferred instant + the same cutoffs
    for h in (2, 3, 4, 6):
        add(f"actual-pref, drop day-use before {h:02d}:00",
            lambda res, actual, in_window, pu, tz, rates, ds, de, h=h:
                in_window(actual or res.get("StartUtc"))
                and not (in_window(res.get("EndUtc"))
                         and (_hour(actual or res.get("StartUtc"), pu, tz) or 0) < h))
        add(f"actual-pref, drop day-use before {h:02d}:00 unless walk-in",
            lambda res, actual, in_window, pu, tz, rates, ds, de, h=h:
                in_window(actual or res.get("StartUtc"))
                and not (in_window(res.get("EndUtc"))
                         and not _walk_in(res, rates)
                         and (_hour(actual or res.get("StartUtc"), pu, tz) or 0) < h))

    # --- zero-night stays never count as arrivals / always count
    add("sched, drop ALL day-use",
        lambda res, actual, in_window, pu, tz, rates, ds, de:
            in_window(res.get("StartUtc")) and not in_window(res.get("EndUtc")))
    add("actual-pref, drop ALL day-use",
        lambda res, actual, in_window, pu, tz, rates, ds, de:
            in_window(actual or res.get("StartUtc")) and not in_window(res.get("EndUtc")))

    return rules


async def main():
    print(f"loading {len(SHEETS)} sheets ...", flush=True)
    truth, skipped = {}, []
    for prop, (short, sid) in SHEETS.items():
        try:
            held, tabs = sheet_tabs(sid)
        except Exception as e:
            skipped.append(f"{short} (sheet unreadable: {e})")
            continue
        if held != DATE:
            skipped.append(f"{short} (sheet holds {held})")
            continue
        if "Arrivals" not in tabs:
            skipped.append(f"{short} (no Arrivals tab)")
            continue
        truth[prop] = tabs["Arrivals"]

    print(f"scoring against {len(truth)} properties; skipped: {', '.join(skipped) or 'none'}\n")
    if not truth:
        return

    rules = make_rules()
    results = []
    for name, fn in rules.items():
        ss_mod._ST_ARRIVAL_RULE = fn
        exact_props, cat_hits, cat_total, detail = 0, 0, 0, []
        try:
            for prop, sheet_cats in truth.items():
                short = SHEETS[prop][0]
                try:
                    rep = await sync_service.get_st_files_report(prop, DATE)
                except Exception as e:
                    detail.append(f"{short}:ERR({str(e)[:30]})")
                    cat_total += len(sheet_cats)
                    continue
                ours = Counter()
                for a in rep.get("arrivals", []):
                    ours[(a.get("category") or "?").strip()] += a.get("spaces", 0)
                cats = set(ours) | set(sheet_cats)
                bad = [c for c in cats if ours.get(c, 0) != sheet_cats.get(c, 0)]
                cat_total += len(cats)
                cat_hits += len(cats) - len(bad)
                if not bad:
                    exact_props += 1
                else:
                    detail.append(f"{short}:" + ",".join(
                        f"{c}{ours.get(c,0)-sheet_cats.get(c,0):+d}" for c in sorted(bad)))
        finally:
            ss_mod._ST_ARRIVAL_RULE = None
        results.append((exact_props, cat_hits, cat_total, name, detail))

    results.sort(key=lambda r: (-r[0], -r[1], r[3]))
    print(f"{'rule':<48} {'props':<8} {'categories':<12} misses")
    print("-" * 120)
    for exact, hits, total, name, detail in results:
        print(f"{name:<48} {exact}/{len(truth):<6} {hits}/{total:<10} "
              + "  ".join(detail[:4]))


asyncio.run(main())
